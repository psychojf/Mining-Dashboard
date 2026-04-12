# -*- coding: utf-8 -*-
import os
import re
import ctypes
import glob
import json
import sys
import zipfile
import tempfile
import threading
import urllib.request
import urllib.error
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Tuple
from functools import lru_cache
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import time

# winsound is built-in on Windows - no external dependency needed
import winsound

# Watchdog for event-driven gamelog monitoring (replaces 500ms polling loop)
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# Conditional imports
try:
    from plyer import notification
    HAS_NOTIFICATION = True
except ImportError:
    HAS_NOTIFICATION = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from PIL import Image, ImageDraw
    import pystray
    HAS_PYSTRAY = True
except ImportError:
    HAS_PYSTRAY = False

# config defaults
DOCS = os.path.expanduser(r"~\Documents\EVE\logs\Gamelogs\*")
CRIT_SOUND_FILE = "alert_crit.wav"
PLAY_CRIT_SOUND = True  # Default to playing the sound
CONFIG_FILE = "mining_config.json"
UPDATE_INTERVAL_MS = 500
HISTORY_DAYS = 60
CRITICAL_HIT_KEYWORD = "Critical mining success"
MAX_MODULES = 5  # Maximum mining modules per ship

# auto-pause keywords (notify events that should pause session)
AUTO_PAUSE_KEYWORDS = [
    "Targeting attempt failed as the designated object is no longer present",
    "cargo hold is full","The asteroid is depleted",
]

# ---------------------------------------------------------------------------
# THEME ENGINE (Imported from Ratting Dashboard)
# ---------------------------------------------------------------------------
# Éclaircit une couleur hex par un montant fixe
def _lighten(hx, amt):
    h = hx.lstrip('#')
    r = min(255, int(h[0:2], 16) + amt)
    g = min(255, int(h[2:4], 16) + amt)
    b = min(255, int(h[4:6], 16) + amt)
    return f"#{r:02x}{g:02x}{b:02x}"

# Assombrit une couleur hex par un facteur multiplicatif
def _dim(hx, factor=0.6):
    h = hx.lstrip('#')
    r = int(int(h[0:2], 16) * factor)
    g = int(int(h[2:4], 16) * factor)
    b = int(int(h[4:6], 16) * factor)
    return f"#{r:02x}{g:02x}{b:02x}"

# Mélange deux couleurs hex selon un ratio t (0=h1, 1=h2)
def _blend(h1, h2, t=0.5):
    a = h1.lstrip('#'); b = h2.lstrip('#')
    r = int(int(a[0:2], 16) * (1 - t) + int(b[0:2], 16) * t)
    g = int(int(a[2:4], 16) * (1 - t) + int(b[2:4], 16) * t)
    bl = int(int(a[4:6], 16) * (1 - t) + int(b[4:6], 16) * t)
    return f"#{min(255,r):02x}{min(255,g):02x}{min(255,bl):02x}"

# Génère un dictionnaire de thème complet à partir d'une couleur de base et d'accent
def _gen_theme(base, accent):
    return {
        "BG": base, "BG_P": _lighten(base, 10), "BG_H": _lighten(base, 18),
        "BG_C": base, "BG_POP": _lighten(base, 10),
        "BD": _lighten(base, 30), "BDG": _lighten(base, 42),
        "T0": accent, "T1": _dim(accent, 0.7),
        "TB": "#e5e5e5", "TD": "#777777",
        "CD": accent, "CR": "#cc3325", "CG": "#d4b45d", "CI": "#55a34f",
        "CT": "#c45b47", "CK": "#896a9e", "CW": "#c48b47", "CM": "#777777",
        "CA": "#55a34f", "CP": "#b89645", "CS": "#cc3325",
        "CH": _blend(accent, "#5c7b8c"), "C_DETACH": accent,
        "C_MSN": _blend(accent, "#5b9bd5"), "C_ALERT": "#e07040",
        "C_ESCAL": "#d4b45d", "C_ANOM": _blend(accent, "#5b8fa8"),
    }

THEMES = {
    "EVE Online (Default)": {
        "BG": "#0b0e17", "BG_P": "#111827", "BG_H": "#111827",
        "BG_C": "#0b0e17", "BG_POP": "#111827",
        "BD": "#1e3a4a", "BDG": "#3dd8e0",
        "T0": "#3dd8e0", "T1": "#5a7085",
        "TB": "#ffffff", "TD": "#5a7085",
        "CD": "#3dd8e0", "CR": "#cc3325", "CG": "#ffd700", "CI": "#2ecc40",
        "CT": "#c45b47", "CK": "#896a9e", "CW": "#c48b47", "CM": "#5a7085",
        "CA": "#2ecc40", "CP": "#b89645", "CS": "#cc3325",
        "CH": "#3dd8e0", "C_DETACH": "#3dd8e0",
        "C_MSN": "#5b9bd5", "C_ALERT": "#e07040",
        "C_ESCAL": "#d4b45d", "C_ANOM": "#5b8fa8",
    },
    "Caldari":                  _gen_theme("#191919", "#3C5F73"),
    "Caldari II":               _gen_theme("#0F1114", "#8A8F9A"),
    "Minmatar":                 _gen_theme("#161414", "#5A3737"),
    "Minmatar II":              _gen_theme("#140D0F", "#8C5055"),
    "Amarr":                    _gen_theme("#191714", "#BBA183"),
    "Amarr II":                 _gen_theme("#12110A", "#9A6928"),
    "Gallente":                 _gen_theme("#0F1414", "#576866"),
    "Gallente II":              _gen_theme("#0A0F0F", "#9EAE95"),
    "Guristas Pirates":         _gen_theme("#261500", "#FF9100"),
    "Blood Raiders":            _gen_theme("#260505", "#BE0000"),
    "Angel Cartel":             _gen_theme("#26110E", "#FF4D00"),
    "Serpentis":                _gen_theme("#060A0C", "#BBC400"),
    "Sansha's Nation":          _gen_theme("#0a0a0a", "#218000"),
    "Triglavian Collective":    _gen_theme("#262218", "#DE1400"),
    "Sisters of EVE":           _gen_theme("#262626", "#B60000"),
    "EDENCOM":                  _gen_theme("#001926", "#039DFF"),
    "Intaki Syndicate":         _gen_theme("#060A0C", "#393780"),
    "ORE":                      _gen_theme("#1A1A1A", "#D9A600"),
    "Mordu's Legion":           _gen_theme("#1A1F22", "#4B6B78"),
    "Thukker Tribe":            _gen_theme("#1F1A17", "#B35900"),
    "CONCORD":                  _gen_theme("#0A1428", "#0088FF"),
    "Society of Conscious Thought": _gen_theme("#0A111A", "#00E8FF"),
}
THEME_NAMES = list(THEMES.keys())

# Window transparency (0.2 – 1.0) — overridden by saved config on startup
WIN_ALPHA: float = 0.85

# Color palette variables (will be dynamically overridden by theme)
BG = "#0b0e17"
BG_PANEL = "#111827"
BORDER = "#1e3a4a"
CYAN = "#3dd8e0"
RED = "#cc3325"
GREEN = "#2ecc40"
GOLD = "#ffd700"
DIM = "#5a7085"
WHITE = "#ffffff"

# Applique les couleurs du thème choisi aux variables globales de couleur
def apply_theme_colors(name):
    global BG, BG_PANEL, BORDER, CYAN, RED, GREEN, GOLD, DIM, WHITE
    t = THEMES.get(name, THEMES["EVE Online (Default)"])
    BG = t["BG"]
    BG_PANEL = t["BG_P"]
    BORDER = t["BD"]
    CYAN = t["T0"]
    RED = t["CR"]
    GREEN = t["CI"]
    GOLD = t["CG"]
    DIM = t["TD"]
    WHITE = t["TB"]

# static geometry defaults
DEFAULT_WIN_W = 280   # base window width for control hub
DEFAULT_WIN_H = 320   # base window height

# colors per character
CHAR_ACCENTS = ["#3dd8e0", "#ff9f43", "#a29bfe", "#e056fd", "#26de81", "#fc5c65", "#45aaf2", "#fed330"]

# ---------------------------------------------------------------------------
# NEON PROGRESS BAR DRAWING HELPERS
# ---------------------------------------------------------------------------
# Dessine une barre de progression style néon avec lueur et segments sur un canvas Tkinter
def draw_neon_bar(canvas, pct, bar_color=None, glow=True, segments=True):
    bar_color = bar_color or CYAN
    canvas.delete("all")
    canvas.update_idletasks()
    w = canvas.winfo_width()
    h = canvas.winfo_height()
    if w <= 1: return
    pad = 2

    canvas.create_rectangle(0, 0, w, h, fill="#0a1520", outline="#1a2a3a", width=1)
    if pct <= 0: return
    fill_w = max(4, int((w - pad * 2) * min(1.0, pct)))

    if glow:
        glow_colors = ["#0a2530", "#0c2d3a", "#0e3545", "#103d50"]
        for i, gc in enumerate(glow_colors):
            expand = len(glow_colors) - i
            y1 = max(pad, pad + 1 - expand)
            y2 = min(h - pad, h - pad - 1 + expand)
            canvas.create_rectangle(pad, y1, pad + fill_w, y2, fill=gc, outline="")

    canvas.create_rectangle(pad, pad + 2, pad + fill_w, h - pad - 2, fill=bar_color, outline="")
    canvas.create_rectangle(pad, pad + 2, pad + fill_w, pad + 4, fill="#7eeef5" if bar_color == CYAN else "#ff8a80", outline="")

    if segments and fill_w > 20:
        seg_w, seg_gap, seg_y1, seg_y2, x = 6, 4, pad + 4, h - pad - 3, pad + 4
        while x + seg_w < pad + fill_w - 2:
            canvas.create_oval(x, seg_y1, x + seg_w, seg_y2, fill="#7eeef5" if bar_color == CYAN else "#ff8a80", outline="")
            x += seg_w + seg_gap

    canvas.create_rectangle(pad, h - pad - 2, pad + fill_w, h - pad, fill="#062030", outline="")

# ---------------------------------------------------------------------------
# ORE / ICE / GAS DATA  (SDE-aware, auto-updatable)
# Source: EVE Online SDE build 3294658 (Apr 12, 2026)
# ---------------------------------------------------------------------------
ORE_DATA_CACHE_FILE = "ore_data_cache.json"
SDE_LATEST_URL = "https://developers.eveonline.com/static-data/eve-online-static-data-latest-jsonl.zip"
SDE_ASTEROID_CATEGORY_ID = 25
SDE_GAS_GROUP_ID = 711
SDE_SKIP_GROUPS = {
    "Deadspace Asteroids", "Empire Asteroids", "Non-Interactable Asteroids",
    "Scalable Decorative Asteroid", "Ancient Compressed Ice",
    "AIR Ore Asteroid Resources"
}

# Source: EVE Online SDE build 3294658 (Apr 12, 2026)
# To update before building the exe, run ore_data.py and copy its _SEED_VOLUMES /
# _build_seed_ratios() output into these two dicts.
_DEFAULT_ORE_VOLUMES: Dict[str, float] = {
    # 0.1 m³
    "Banidine": 0.1, "Mordunium": 0.1, "Mordunium II-Grade": 0.1, "Mordunium III-Grade": 0.1,
    "Mordunium IV-Grade": 0.1, "Veldspar": 0.1, "Veldspar 0-Grade": 0.1, "Veldspar II-Grade": 0.1,
    "Veldspar III-Grade": 0.1, "Veldspar IV-Grade": 0.1,
    # 0.15 m³
    "Scordite": 0.15, "Scordite 0-Grade": 0.15, "Scordite II-Grade": 0.15, "Scordite III-Grade": 0.15,
    "Scordite IV-Grade": 0.15,
    # 0.3 m³
    "Augumene": 0.3, "Pyroxeres": 0.3, "Pyroxeres II-Grade": 0.3, "Pyroxeres III-Grade": 0.3,
    "Pyroxeres IV-Grade": 0.3,
    # 0.35 m³
    "Plagioclase": 0.35, "Plagioclase II-Grade": 0.35, "Plagioclase III-Grade": 0.35, "Plagioclase IV-Grade": 0.35,
    # 0.5 m³
    "Nesosilicate Rakovene": 0.5,
    # 0.6 m³
    "Mercium": 0.6, "Omber": 0.6, "Omber II-Grade": 0.6, "Omber III-Grade": 0.6,
    "Omber IV-Grade": 0.6, "Tyranite": 0.6, "Ytirium": 0.6, "Ytirium II-Grade": 0.6,
    "Ytirium III-Grade": 0.6, "Ytirium IV-Grade": 0.6,
    # 0.8 m³
    "Griemeer": 0.8, "Griemeer II-Grade": 0.8, "Griemeer III-Grade": 0.8, "Griemeer IV-Grade": 0.8,
    # 1.0 m³
    "Fullerite-C50": 1.0, "Fullerite-C60": 1.0, "Fullerite-C70": 1.0,
    # 1.2 m³
    "Kernite": 1.2, "Kernite II-Grade": 1.2, "Kernite III-Grade": 1.2, "Kernite IV-Grade": 1.2,
    "Kylixium": 1.2, "Kylixium II-Grade": 1.2, "Kylixium III-Grade": 1.2, "Kylixium IV-Grade": 1.2,
    "Lyavite": 1.2,
    # 2.0 m³
    "Fullerite-C28": 2.0, "Fullerite-C72": 2.0, "Fullerite-C84": 2.0, "Jaspet": 2.0,
    "Jaspet II-Grade": 2.0, "Jaspet III-Grade": 2.0, "Jaspet IV-Grade": 2.0, "Pithix": 2.0,
    # 3.0 m³
    "Hedbergite": 3.0, "Hedbergite II-Grade": 3.0, "Hedbergite III-Grade": 3.0, "Hedbergite IV-Grade": 3.0,
    "Hemorphite": 3.0, "Hemorphite II-Grade": 3.0, "Hemorphite III-Grade": 3.0, "Hemorphite IV-Grade": 3.0,
    # 4.0 m³
    "Admixti Mutanite": 4.0, "Amperum Mutanite": 4.0, "Conflagrati Mutanite": 4.0, "Nocxite": 4.0,
    "Nocxite II-Grade": 4.0, "Nocxite III-Grade": 4.0, "Nocxite IV-Grade": 4.0, "Peregrinus Mutanite": 4.0,
    "Solis Mutanite": 4.0, "Tenebraet Mutanite": 4.0,
    # 5.0 m³
    "Fullerite-C32": 5.0, "Fullerite-C320": 5.0, "Gneiss": 5.0, "Gneiss II-Grade": 5.0,
    "Gneiss III-Grade": 5.0, "Gneiss IV-Grade": 5.0, "Green Arisite": 5.0, "Hezorime": 5.0,
    "Hezorime II-Grade": 5.0, "Hezorime III-Grade": 5.0, "Hezorime IV-Grade": 5.0, "Ueganite": 5.0,
    "Ueganite II-Grade": 5.0, "Ueganite III-Grade": 5.0, "Ueganite IV-Grade": 5.0,
    # 8.0 m³
    "Dark Ochre": 8.0, "Dark Ochre II-Grade": 8.0, "Dark Ochre III-Grade": 8.0, "Dark Ochre IV-Grade": 8.0,
    "Oeryl": 8.0,
    # 10.0 m³
    "Amber Cytoserocin": 10.0, "Amber Mykoserocin": 10.0, "Azure Cytoserocin": 10.0, "Azure Mykoserocin": 10.0,
    "Bitumens": 10.0, "Bountiful Loparite": 10.0, "Bountiful Monazite": 10.0, "Bountiful Xenotime": 10.0,
    "Bountiful Ytterbite": 10.0, "Brimful Bitumens": 10.0, "Brimful Coesite": 10.0, "Brimful Sylvite": 10.0,
    "Brimful Zeolites": 10.0, "Carnotite": 10.0, "Celadon Cytoserocin": 10.0, "Celadon Mykoserocin": 10.0,
    "Chartreuse Cytoserocin": 10.0, "Chromite": 10.0, "Cinnabar": 10.0, "Cobaltite": 10.0,
    "Coesite": 10.0, "Copious Cobaltite": 10.0, "Copious Euxenite": 10.0, "Copious Scheelite": 10.0,
    "Copious Titanite": 10.0, "Euxenite": 10.0, "Fullerite-C540": 10.0, "Gamboge Cytoserocin": 10.0,
    "Glistening Bitumens": 10.0, "Glistening Coesite": 10.0, "Glistening Sylvite": 10.0, "Glistening Zeolites": 10.0,
    "Glowing Carnotite": 10.0, "Glowing Cinnabar": 10.0, "Glowing Pollucite": 10.0, "Glowing Zircon": 10.0,
    "Golden Cytoserocin": 10.0, "Golden Mykoserocin": 10.0, "Hiemal Tricarboxyl Vapor": 10.0, "Lavish Chromite": 10.0,
    "Lavish Otavite": 10.0, "Lavish Sperrylite": 10.0, "Lavish Vanadinite": 10.0, "Lime Cytoserocin": 10.0,
    "Lime Mykoserocin": 10.0, "Loparite": 10.0, "Malachite Cytoserocin": 10.0, "Malachite Mykoserocin": 10.0,
    "Monazite": 10.0, "Otavite": 10.0, "Pollucite": 10.0, "Replete Carnotite": 10.0,
    "Replete Cinnabar": 10.0, "Replete Pollucite": 10.0, "Replete Zircon": 10.0, "Scheelite": 10.0,
    "Shimmering Chromite": 10.0, "Shimmering Otavite": 10.0, "Shimmering Sperrylite": 10.0, "Shimmering Vanadinite": 10.0,
    "Shining Loparite": 10.0, "Shining Monazite": 10.0, "Shining Xenotime": 10.0, "Shining Ytterbite": 10.0,
    "Sperrylite": 10.0, "Sylvite": 10.0, "Titanite": 10.0, "Twinkling Cobaltite": 10.0,
    "Twinkling Euxenite": 10.0, "Twinkling Scheelite": 10.0, "Twinkling Titanite": 10.0, "Vanadinite": 10.0,
    "Vermillion Cytoserocin": 10.0, "Vermillion Mykoserocin": 10.0, "Viridian Cytoserocin": 10.0, "Viridian Mykoserocin": 10.0,
    "Xenotime": 10.0, "Ytterbite": 10.0, "Zeolites": 10.0, "Zircon": 10.0,
    # 16.0 m³
    "Arkonor": 16.0, "Arkonor II-Grade": 16.0, "Arkonor III-Grade": 16.0, "Arkonor IV-Grade": 16.0,
    "Bezdnacine": 16.0, "Bezdnacine II-Grade": 16.0, "Bezdnacine III-Grade": 16.0, "Bistot": 16.0,
    "Bistot II-Grade": 16.0, "Bistot III-Grade": 16.0, "Bistot IV-Grade": 16.0, "Crokite": 16.0,
    "Crokite II-Grade": 16.0, "Crokite III-Grade": 16.0, "Crokite IV-Grade": 16.0, "Ducinium": 16.0,
    "Ducinium II-Grade": 16.0, "Ducinium III-Grade": 16.0, "Ducinium IV-Grade": 16.0, "Eifyrium": 16.0,
    "Eifyrium II-Grade": 16.0, "Eifyrium III-Grade": 16.0, "Eifyrium IV-Grade": 16.0, "Geodite": 16.0,
    "Polygypsum": 16.0, "Rakovene": 16.0, "Rakovene II-Grade": 16.0, "Rakovene III-Grade": 16.0,
    "Spodumain": 16.0, "Spodumain II-Grade": 16.0, "Spodumain III-Grade": 16.0, "Spodumain IV-Grade": 16.0,
    "Talassonite": 16.0, "Talassonite II-Grade": 16.0, "Talassonite III-Grade": 16.0,
    # 40.0 m³
    "Mercoxit": 40.0, "Mercoxit II-Grade": 40.0, "Mercoxit III-Grade": 40.0, "Zuthrine": 40.0,
    # 1000.0 m³
    "Azure Ice": 1000.0, "Blue Ice": 1000.0, "Blue Ice IV-Grade": 1000.0, "Clear Icicle": 1000.0,
    "Clear Icicle IV-Grade": 1000.0, "Crystalline Icicle": 1000.0, "Dark Glitter": 1000.0, "Gelidus": 1000.0,
    "Glacial Mass": 1000.0, "Glacial Mass IV-Grade": 1000.0, "Glare Crust": 1000.0, "Krystallos": 1000.0,
    "White Glaze": 1000.0, "White Glaze IV-Grade": 1000.0,
}

_DEFAULT_COMPRESSION_RATIOS: Dict[str, int] = {
    # Standard ore — 100:1
    "Arkonor": 100, "Arkonor II-Grade": 100, "Arkonor III-Grade": 100, "Arkonor IV-Grade": 100,
    "Bezdnacine": 100, "Bezdnacine II-Grade": 100, "Bezdnacine III-Grade": 100, "Bistot": 100,
    "Bistot II-Grade": 100, "Bistot III-Grade": 100, "Bistot IV-Grade": 100, "Bitumens": 100,
    "Bountiful Loparite": 100, "Bountiful Monazite": 100, "Bountiful Xenotime": 100, "Bountiful Ytterbite": 100,
    "Brimful Bitumens": 100, "Brimful Coesite": 100, "Brimful Sylvite": 100, "Brimful Zeolites": 100,
    "Carnotite": 100, "Chromite": 100, "Cinnabar": 100, "Cobaltite": 100,
    "Coesite": 100, "Copious Cobaltite": 100, "Copious Euxenite": 100, "Copious Scheelite": 100,
    "Copious Titanite": 100, "Crokite": 100, "Crokite II-Grade": 100, "Crokite III-Grade": 100,
    "Crokite IV-Grade": 100, "Dark Ochre": 100, "Dark Ochre II-Grade": 100, "Dark Ochre III-Grade": 100,
    "Dark Ochre IV-Grade": 100, "Ducinium": 100, "Ducinium II-Grade": 100, "Ducinium III-Grade": 100,
    "Ducinium IV-Grade": 100, "Eifyrium": 100, "Eifyrium II-Grade": 100, "Eifyrium III-Grade": 100,
    "Eifyrium IV-Grade": 100, "Euxenite": 100, "Glistening Bitumens": 100, "Glistening Coesite": 100,
    "Glistening Sylvite": 100, "Glistening Zeolites": 100, "Glowing Carnotite": 100, "Glowing Cinnabar": 100,
    "Glowing Pollucite": 100, "Glowing Zircon": 100, "Gneiss": 100, "Gneiss II-Grade": 100,
    "Gneiss III-Grade": 100, "Gneiss IV-Grade": 100, "Griemeer": 100, "Griemeer II-Grade": 100,
    "Griemeer III-Grade": 100, "Griemeer IV-Grade": 100, "Hedbergite": 100, "Hedbergite II-Grade": 100,
    "Hedbergite III-Grade": 100, "Hedbergite IV-Grade": 100, "Hemorphite": 100, "Hemorphite II-Grade": 100,
    "Hemorphite III-Grade": 100, "Hemorphite IV-Grade": 100, "Hezorime": 100, "Hezorime II-Grade": 100,
    "Hezorime III-Grade": 100, "Hezorime IV-Grade": 100, "Jaspet": 100, "Jaspet II-Grade": 100,
    "Jaspet III-Grade": 100, "Jaspet IV-Grade": 100, "Kernite": 100, "Kernite II-Grade": 100,
    "Kernite III-Grade": 100, "Kernite IV-Grade": 100, "Kylixium": 100, "Kylixium II-Grade": 100,
    "Kylixium III-Grade": 100, "Kylixium IV-Grade": 100, "Lavish Chromite": 100, "Lavish Otavite": 100,
    "Lavish Sperrylite": 100, "Lavish Vanadinite": 100, "Loparite": 100, "Mercoxit": 100,
    "Mercoxit II-Grade": 100, "Mercoxit III-Grade": 100, "Monazite": 100, "Mordunium": 100,
    "Mordunium II-Grade": 100, "Mordunium III-Grade": 100, "Mordunium IV-Grade": 100, "Nocxite": 100,
    "Nocxite II-Grade": 100, "Nocxite III-Grade": 100, "Nocxite IV-Grade": 100, "Omber": 100,
    "Omber II-Grade": 100, "Omber III-Grade": 100, "Omber IV-Grade": 100, "Otavite": 100,
    "Plagioclase": 100, "Plagioclase II-Grade": 100, "Plagioclase III-Grade": 100, "Plagioclase IV-Grade": 100,
    "Pollucite": 100, "Pyroxeres": 100, "Pyroxeres II-Grade": 100, "Pyroxeres III-Grade": 100,
    "Pyroxeres IV-Grade": 100, "Rakovene": 100, "Rakovene II-Grade": 100, "Rakovene III-Grade": 100,
    "Replete Carnotite": 100, "Replete Cinnabar": 100, "Replete Pollucite": 100, "Replete Zircon": 100,
    "Scheelite": 100, "Scordite": 100, "Scordite 0-Grade": 100, "Scordite II-Grade": 100,
    "Scordite III-Grade": 100, "Scordite IV-Grade": 100, "Shimmering Chromite": 100, "Shimmering Otavite": 100,
    "Shimmering Sperrylite": 100, "Shimmering Vanadinite": 100, "Shining Loparite": 100, "Shining Monazite": 100,
    "Shining Xenotime": 100, "Shining Ytterbite": 100, "Sperrylite": 100, "Spodumain": 100,
    "Spodumain II-Grade": 100, "Spodumain III-Grade": 100, "Spodumain IV-Grade": 100, "Sylvite": 100,
    "Talassonite": 100, "Talassonite II-Grade": 100, "Talassonite III-Grade": 100, "Titanite": 100,
    "Twinkling Cobaltite": 100, "Twinkling Euxenite": 100, "Twinkling Scheelite": 100, "Twinkling Titanite": 100,
    "Ueganite": 100, "Ueganite II-Grade": 100, "Ueganite III-Grade": 100, "Ueganite IV-Grade": 100,
    "Vanadinite": 100, "Veldspar": 100, "Veldspar 0-Grade": 100, "Veldspar II-Grade": 100,
    "Veldspar III-Grade": 100, "Veldspar IV-Grade": 100, "Xenotime": 100, "Ytirium": 100,
    "Ytirium II-Grade": 100, "Ytirium III-Grade": 100, "Ytirium IV-Grade": 100, "Ytterbite": 100,
    "Zeolites": 100, "Zircon": 100,
    # Ice / Gas — 10:1
    "Amber Cytoserocin": 10, "Amber Mykoserocin": 10, "Azure Cytoserocin": 10, "Azure Mykoserocin": 10,
    "Blue Ice": 10, "Blue Ice IV-Grade": 10, "Celadon Cytoserocin": 10, "Celadon Mykoserocin": 10,
    "Clear Icicle": 10, "Clear Icicle IV-Grade": 10, "Dark Glitter": 10, "Fullerite-C28": 10,
    "Fullerite-C32": 10, "Fullerite-C320": 10, "Fullerite-C50": 10, "Fullerite-C540": 10,
    "Fullerite-C60": 10, "Fullerite-C70": 10, "Fullerite-C72": 10, "Fullerite-C84": 10,
    "Gelidus": 10, "Glacial Mass": 10, "Glacial Mass IV-Grade": 10, "Glare Crust": 10,
    "Golden Cytoserocin": 10, "Golden Mykoserocin": 10, "Krystallos": 10, "Lime Cytoserocin": 10,
    "Lime Mykoserocin": 10, "Malachite Cytoserocin": 10, "Malachite Mykoserocin": 10, "Vermillion Cytoserocin": 10,
    "Vermillion Mykoserocin": 10, "Viridian Cytoserocin": 10, "Viridian Mykoserocin": 10, "White Glaze": 10,
    "White Glaze IV-Grade": 10,
    # Not compressible — 1:1
    "Admixti Mutanite": 1, "Amperum Mutanite": 1, "Augumene": 1, "Azure Ice": 1,
    "Banidine": 1, "Chartreuse Cytoserocin": 1, "Conflagrati Mutanite": 1, "Crystalline Icicle": 1,
    "Gamboge Cytoserocin": 1, "Geodite": 1, "Green Arisite": 1, "Hiemal Tricarboxyl Vapor": 1,
    "Lyavite": 1, "Mercium": 1, "Nesosilicate Rakovene": 1, "Oeryl": 1,
    "Peregrinus Mutanite": 1, "Pithix": 1, "Polygypsum": 1, "Solis Mutanite": 1,
    "Tenebraet Mutanite": 1, "Tyranite": 1, "Zuthrine": 1,
}

# Charge les données de minerais depuis le fichier cache JSON local
def _load_ore_data_from_cache():
    try:
        if os.path.exists(ORE_DATA_CACHE_FILE):
            with open(ORE_DATA_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception: pass
    return None

# Sauvegarde les données de minerais analysées dans le fichier cache JSON
def _save_ore_data_cache(data):
    try:
        with open(ORE_DATA_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as e: print(f"Warning: could not save ore data cache: {e}")

# Analyse les fichiers JSONL du SDE pour extraire les volumes et ratios de compression des minerais
def _parse_sde_ore_data(sde_dir):
    categories = {}
    with open(os.path.join(sde_dir, "categories.jsonl"), "r", encoding="utf-8") as f:
        for line in f:
            obj = json.loads(line)
            categories[obj["_key"]] = obj.get("name", {}).get("en", "")

    groups = {}
    with open(os.path.join(sde_dir, "groups.jsonl"), "r", encoding="utf-8") as f:
        for line in f:
            obj = json.loads(line)
            groups[obj["_key"]] = {
                "name": obj.get("name", {}).get("en", ""),
                "categoryID": obj.get("categoryID", 0),
                "published": obj.get("published", False)
            }

    compress_map = {}
    with open(os.path.join(sde_dir, "compressibleTypes.jsonl"), "r", encoding="utf-8") as f:
        for line in f:
            obj = json.loads(line)
            compress_map[obj["_key"]] = obj["compressedTypeID"]

    types_by_id = {}
    with open(os.path.join(sde_dir, "types.jsonl"), "r", encoding="utf-8") as f:
        for line in f:
            obj = json.loads(line)
            types_by_id[obj["_key"]] = obj

    asteroid_groups = {}
    for gid, g in groups.items():
        if g["categoryID"] == SDE_ASTEROID_CATEGORY_ID and g["published"]:
            if g["name"] not in SDE_SKIP_GROUPS:
                asteroid_groups[gid] = g["name"]

    ore_volumes = {}
    compression_ratios = {}
    for tid, t in types_by_id.items():
        if not t.get("published"): continue
        name = t.get("name", {}).get("en", "")
        vol = t.get("volume", 0)
        gid = t.get("groupID", 0)
        if "Compressed" in name: continue
        if gid not in asteroid_groups and gid != SDE_GAS_GROUP_ID: continue
        comp_ratio = 1
        if tid in compress_map:
            comp_type = types_by_id.get(compress_map[tid])
            if comp_type and vol > 0:
                cv = comp_type.get("volume", 0)
                if cv > 0: comp_ratio = round(vol / cv)
        ore_volumes[name] = vol
        compression_ratios[name] = comp_ratio

    sde_version = ""
    sde_meta = os.path.join(sde_dir, "_sde.jsonl")
    if os.path.exists(sde_meta):
        with open(sde_meta, "r", encoding="utf-8") as f:
            for line in f:
                obj = json.loads(line)
                sde_version = str(obj.get("buildNumber", obj.get("_key", "")))

    return {
        "sde_version": sde_version,
        "updated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "ore_count": len(ore_volumes),
        "ore_volumes": ore_volumes,
        "compression_ratios": compression_ratios
    }

# Télécharge le SDE de CCP, extrait les fichiers nécessaires et retourne les données de minerais
def download_and_parse_sde(progress_callback=None):
    if progress_callback: progress_callback("Downloading SDE from CCP...")
    with tempfile.TemporaryDirectory() as tmp_dir:
        zip_path = os.path.join(tmp_dir, "sde.zip")
        req = urllib.request.Request(SDE_LATEST_URL, headers={"User-Agent": "EVE-Mining-Dashboard/1.0"})
        response = urllib.request.urlopen(req, timeout=120)
        total = int(response.headers.get("Content-Length", 0))
        downloaded = 0
        with open(zip_path, "wb") as f:
            while True:
                chunk = response.read(256 * 1024)
                if not chunk: break
                f.write(chunk)
                downloaded += len(chunk)
                if progress_callback and total > 0:
                    pct = int(downloaded * 100 / total)
                    mb = downloaded / (1024 * 1024)
                    progress_callback(f"Downloading SDE... {mb:.1f} MB ({pct}%)")

        if progress_callback: progress_callback("Extracting SDE data...")
        needed = ["types.jsonl", "groups.jsonl", "categories.jsonl", "compressibleTypes.jsonl", "_sde.jsonl"]
        extract_dir = os.path.join(tmp_dir, "sde")
        os.makedirs(extract_dir, exist_ok=True)
        with zipfile.ZipFile(zip_path, "r") as zf:
            for name in needed:
                if name in zf.namelist(): zf.extract(name, extract_dir)

        if progress_callback: progress_callback("Parsing ore data...")
        return _parse_sde_ore_data(extract_dir)

ORE_VOLUMES: Dict[str, float] = {}
COMPRESSION_RATIOS: Dict[str, int] = {}
SDE_INFO: Dict[str, str] = {"version": "built-in", "updated_at": "n/a", "ore_count": "0"}

_cached = _load_ore_data_from_cache()
if _cached and "ore_volumes" in _cached:
    ORE_VOLUMES = {k: float(v) for k, v in _cached["ore_volumes"].items()}
    COMPRESSION_RATIOS = {k: int(v) for k, v in _cached["compression_ratios"].items()}
    SDE_INFO["version"] = _cached.get("sde_version", "cached")
    SDE_INFO["updated_at"] = _cached.get("updated_at", "unknown")
    SDE_INFO["ore_count"] = str(_cached.get("ore_count", len(ORE_VOLUMES)))
else:
    ORE_VOLUMES = dict(_DEFAULT_ORE_VOLUMES)
    COMPRESSION_RATIOS = dict(_DEFAULT_COMPRESSION_RATIOS)
    SDE_INFO["version"] = "3215400 (built-in)"
    SDE_INFO["ore_count"] = str(len(ORE_VOLUMES))

MINING_LINE = re.compile(r'^\[.*?\]\s+\(mining\)', re.IGNORECASE)
REGULAR_MINE_PATTERN = re.compile(r"You mined <font size=12><color=[^>]+>(?P<amount>\d+)<color=[^>]+><font size=10> units of <color=[^>]+><font size=12>(?P<ore_type>[^\r\n<]+)", re.IGNORECASE)
CRIT_MINE_PATTERN = re.compile(r"You mined an additional <color=[^>]+><font size=12>(?P<amount>\d+)<color=[^>]+><font size=10> units of <color=[^>]+><font size=12>(?P<ore_type>[^\r\n<]+)", re.IGNORECASE | re.DOTALL)
COMPRESSION_PATTERN = re.compile(r'Successfully compressed (?P<ore_type>[^\s]+) into (?P<amount>[\d,]+) Compressed', re.IGNORECASE)
RESIDUE_PATTERN = re.compile(r"Additional <font size=12><color=[^>]+>(?P<amount>\d+)<color=[^>]+><font size=10> units depleted from asteroid as residue", re.IGNORECASE)
LISTENER_LINE = re.compile(r'Listener:\s*(.+)', re.IGNORECASE)
LOG_TIMESTAMP = re.compile(r'^\[\s*(\d{4}\.\d{2}\.\d{2})\s+\d{2}:\d{2}:\d{2}\s*\]')
# SDE download progress messages embed a percentage like "(34%)" — pre-compiled to avoid
# creating a new pattern object on every progress callback invocation
SDE_PROGRESS_PCT = re.compile(r'\((?P<pct>\d+)%\)')

_ORE_CATEGORIES = {
    "Veldspar": "2ecc40", "Scordite": "2ecc40", "Pyroxeres": "2ecc40", "Plagioclase": "2ecc40", "Omber": "2ecc40", "Kernite": "2ecc40",
    "Jaspet": "f1c40f", "Hemorphite": "f1c40f", "Hedbergite": "f1c40f",
    "Gneiss": "ff9f43", "Dark Ochre": "ff9f43", "Spodumain": "ff9f43", "Crokite": "ff9f43", "Bistot": "ff9f43", "Arkonor": "ff9f43", "Mercoxit": "cc3325",
    "Zeolites": "a29bfe", "Sylvite": "a29bfe", "Bitumens": "a29bfe", "Coesite": "a29bfe",
    "Cobaltite": "9b59b6", "Euxenite": "9b59b6", "Titanite": "9b59b6", "Scheelite": "9b59b6",
    "Otavite": "e056fd", "Sperrylite": "e056fd", "Vanadinite": "e056fd", "Chromite": "e056fd",
    "Carnotite": "fd79a8", "Zircon": "fd79a8", "Pollucite": "fd79a8", "Cinnabar": "fd79a8",
    "Xenotime": "ffd700", "Monazite": "ffd700", "Loparite": "ffd700", "Ytterbite": "ffd700",
    "Blue Ice": "74b9ff", "Clear Icicle": "74b9ff", "Glacial Mass": "74b9ff", "White Glaze": "74b9ff", "Glare Crust": "74b9ff", "Dark Glitter": "74b9ff", "Gelidus": "74b9ff", "Krystallos": "74b9ff",
    "Bezdnacine": "00cec9", "Rakovene": "00cec9", "Talassonite": "00cec9",
    "Mordunium": "00d2d3", "Ytirium": "00d2d3", "Eifyrium": "00d2d3", "Griemeer": "00d2d3", "Hezorime": "00d2d3", "Kylixium": "00d2d3", "Nocxite": "00d2d3", "Tyranite": "00d2d3",
    "Ducinium": "ffeaa7", "Ueganite": "ffeaa7", "Mutanite": "ffeaa7",
}

@lru_cache(maxsize=512)
# Retourne la couleur Excel (hex sans #) correspondant à la famille du minerai
def _get_ore_excel_color(ore_name: str) -> str:
    ore_lower = ore_name.lower()
    for base_name, color in _ORE_CATEGORIES.items():
        if base_name.lower() in ore_lower: return color
    if "cytoserocin" in ore_lower or "mykoserocin" in ore_lower: return "55efc4"
    if "fullerite" in ore_lower: return "00b894"
    return "ffffff"

# Widget d'info-bulle affiché avec délai au survol d'un élément Tkinter
class ToolTip:
    # Initialise le tooltip et lie les événements souris au widget cible
    def __init__(self, widget, text=""):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        self._after_id = None
        widget.bind("<Enter>", self._on_enter, add="+")
        widget.bind("<Leave>", self._on_leave, add="+")
        widget.bind("<ButtonPress>", self._on_leave, add="+")

    # Met à jour le texte affiché dans le tooltip
    def update_text(self, new_text):
        self.text = new_text

    # Déclenche l'affichage du tooltip avec un délai de 400ms
    def _on_enter(self, event=None):
        self._cancel()
        self._after_id = self.widget.after(400, self._show)

    # Annule et cache le tooltip quand la souris quitte le widget
    def _on_leave(self, event=None):
        self._cancel()
        self._hide()

    # Annule le minuteur d'affichage en attente
    def _cancel(self):
        if self._after_id:
            self.widget.after_cancel(self._after_id)
            self._after_id = None

    # Crée et positionne la fenêtre popup du tooltip
    def _show(self):
        if not self.text: return
        x = self.widget.winfo_rootx() + self.widget.winfo_width() // 2
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_attributes("-topmost", 1)
        try: tw.wm_attributes("-alpha", 0.92)
        except Exception: pass
        tw.geometry(f"+{x}+{y}")
        tk.Label(tw, text=self.text, bg="#1a2332", fg="#c0d8e8", font=("Consolas", 8), relief="solid", borderwidth=1, padx=6, pady=3, wraplength=260, justify="left").pack()

    # Détruit la fenêtre popup du tooltip
    def _hide(self):
        if self.tipwindow:
            self.tipwindow.destroy()
            self.tipwindow = None

# Représente un module de minage (laser/foreuse) avec rendement et temps de cycle
class MiningModule:
    # Initialise le module avec nom, rendement par cycle et temps de cycle
    def __init__(self, name: str = "", yield_per_cycle: float = 0.0, cycle_time: float = 0.0, enabled: bool = True):
        self.name = name
        self.yield_per_cycle = yield_per_cycle
        self.cycle_time = cycle_time
        self.enabled = enabled

    # Retourne le débit du module en m³/s (0 si non configuré)
    def get_m3_per_sec(self) -> float:
        if self.yield_per_cycle > 0 and self.cycle_time > 0: return self.yield_per_cycle / self.cycle_time
        return 0.0

    # Vérifie si le module a un rendement et un cycle valides
    def is_configured(self) -> bool:
        return self.yield_per_cycle > 0 and self.cycle_time > 0

    # Sérialise le module en dictionnaire pour la sauvegarde JSON
    def to_dict(self) -> Dict:
        return {"name": self.name, "yield_per_cycle": self.yield_per_cycle, "cycle_time": self.cycle_time, "enabled": self.enabled}

    @staticmethod
    # Crée un MiningModule depuis un dictionnaire de configuration
    def from_dict(data: Dict) -> 'MiningModule':
        return MiningModule(name=data.get("name", ""), yield_per_cycle=data.get("yield_per_cycle", 0.0), cycle_time=data.get("cycle_time", 0.0), enabled=data.get("enabled", True))

# Représente le groupe de drones de minage d'un personnage (max 5)
class MiningDrone:
    MAX_DRONES = 5
    # Initialise les drones avec quantité (max 5), rendement par cycle et temps de cycle
    def __init__(self, count: int = 0, yield_per_cycle: float = 0.0, cycle_time: float = 0.0):
        self.count = max(0, min(count, self.MAX_DRONES))
        self.yield_per_cycle = yield_per_cycle
        self.cycle_time = cycle_time

    # Retourne le débit total de tous les drones combinés en m³/s
    def get_total_m3_per_sec(self) -> float:
        if self.count > 0 and self.yield_per_cycle > 0 and self.cycle_time > 0:
            return (self.yield_per_cycle / self.cycle_time) * self.count
        return 0.0

    # Vérifie si les drones ont une configuration complète et valide
    def is_configured(self) -> bool:
        return self.count > 0 and self.yield_per_cycle > 0 and self.cycle_time > 0

    # Sérialise la configuration des drones en dictionnaire
    def to_dict(self) -> Dict:
        return {"count": self.count, "yield_per_cycle": self.yield_per_cycle, "cycle_time": self.cycle_time}

    @staticmethod
    # Crée un MiningDrone depuis un dictionnaire de configuration
    def from_dict(data: Dict) -> 'MiningDrone':
        return MiningDrone(count=data.get("count", 0), yield_per_cycle=data.get("yield_per_cycle", 0.0), cycle_time=data.get("cycle_time", 0.0))

# Suit toutes les données de minage et l'état de session d'un personnage EVE
class CharacterTracker:
    # Initialise les compteurs, profils de vaisseau et état de session du personnage
    def __init__(self, char_id: str, char_name: str):
        self.char_id = char_id
        self.char_name = char_name
        self.log_path: Optional[str] = None
        self.log_pos: int = 0
        self.crit_count: int = 0
        self.crit_m3: float = 0.0
        self.total_m3: float = 0.0
        self.ore_summary: Dict[str, float] = {}
        self.compression_log: Dict[str, float] = {}
        self.total_residue_m3: float = 0.0
        self.residue_summary: Dict[str, float] = {}
        self.current_cargo: float = 0.0
        self.ship_profiles: Dict[str, List[MiningModule]] = {"Default": []}
        self.drone_profiles: Dict[str, MiningDrone] = {"Default": MiningDrone()}
        self.implant_profiles: Dict[str, bool] = {"Default": False}
        self.cargo_profiles: Dict[str, float] = {"Default": 0.0}
        self.active_profile: str = "Default"
        self.session_start_time: float = time.time()
        self.session_start_m3: float = 0.0
        self.session_elapsed_offset: float = 0.0
        self.session_active: bool = False

    # Retourne la durée active cumulée de la session en secondes
    def get_session_active_duration(self) -> float:
        if self.session_active: return self.session_elapsed_offset + (time.time() - self.session_start_time)
        return self.session_elapsed_offset

    # Retourne les modules du profil actif
    def get_active_modules(self) -> List[MiningModule]: return self.ship_profiles.get(self.active_profile, [])
    # Définit les modules du profil actif
    def set_active_modules(self, modules: List[MiningModule]): self.ship_profiles[self.active_profile] = modules
    # Retourne les drones du profil actif
    def get_active_drones(self) -> MiningDrone: return self.drone_profiles.get(self.active_profile, MiningDrone())
    # Définit les drones du profil actif
    def set_active_drones(self, drone: MiningDrone): self.drone_profiles[self.active_profile] = drone
    # Retourne l'état de l'implant de minage pour le profil actif
    def get_active_implant(self) -> bool: return self.implant_profiles.get(self.active_profile, False)
    # Active ou désactive l'implant pour le profil actif
    def set_active_implant(self, enabled: bool): self.implant_profiles[self.active_profile] = enabled
    # Retourne la capacité cargo configurée pour le profil actif
    def get_active_capacity(self) -> float: return self.cargo_profiles.get(self.active_profile, 0.0)
    # Définit la capacité cargo pour le profil actif
    def set_active_capacity(self, capacity: float): self.cargo_profiles[self.active_profile] = capacity

    # Calcule le débit théorique total en m³/s (modules + drones + bonus implant)
    def get_total_theoretical_m3_per_sec(self) -> float:
        total_yield_sec = 0.0
        for module in self.get_active_modules():
            if module.enabled and module.is_configured():
                drain_sec = module.get_m3_per_sec()
                yield_multiplier = 1.054 if self.get_active_implant() else 1.0
                total_yield_sec += drain_sec * yield_multiplier

        drone = self.get_active_drones()
        if drone.is_configured(): total_yield_sec += drone.get_total_m3_per_sec()
        return round(total_yield_sec, 1)

    # Compte le nombre de modules actifs et correctement configurés
    def get_active_module_count(self) -> int: return sum(1 for m in self.get_active_modules() if m.enabled and m.is_configured())
    # Vérifie si au moins un module ou drone est configuré dans le profil actif
    def has_any_configured_module(self) -> bool:
        has_module = any(m.is_configured() for m in self.get_active_modules())
        has_drone = self.get_active_drones().is_configured()
        return has_module or has_drone

    # Retourne la liste des noms de profils de vaisseau
    def get_profile_names(self) -> List[str]: return list(self.ship_profiles.keys())
    # Crée un nouveau profil vide pour tous les types de données (modules, drones, implant, cargo)
    def create_profile(self, name: str):
        if name and name not in self.ship_profiles:
            self.ship_profiles[name] = []
            self.drone_profiles[name] = MiningDrone()
            self.implant_profiles[name] = False
            self.cargo_profiles[name] = 0.0
            return True
        return False
    
    # Supprime un profil et bascule sur un autre si c'était le profil actif (min. 1 requis)
    def delete_profile(self, name: str) -> bool:
        if name in self.ship_profiles and len(self.ship_profiles) > 1:
            if self.active_profile == name:
                for profile_name in self.ship_profiles:
                    if profile_name != name:
                        self.active_profile = profile_name
                        break
            del self.ship_profiles[name]
            if name in self.drone_profiles: del self.drone_profiles[name]
            if name in self.implant_profiles: del self.implant_profiles[name]
            if name in self.cargo_profiles: del self.cargo_profiles[name]
            return True
        return False
    
    # Renomme un profil en mettant à jour toutes les tables de données associées
    def rename_profile(self, old_name: str, new_name: str) -> bool:
        if old_name in self.ship_profiles and new_name and new_name not in self.ship_profiles:
            self.ship_profiles[new_name] = self.ship_profiles.pop(old_name)
            if old_name in self.drone_profiles: self.drone_profiles[new_name] = self.drone_profiles.pop(old_name)
            if old_name in self.implant_profiles: self.implant_profiles[new_name] = self.implant_profiles.pop(old_name)
            if old_name in self.cargo_profiles: self.cargo_profiles[new_name] = self.cargo_profiles.pop(old_name)
            if self.active_profile == old_name: self.active_profile = new_name
            return True
        return False

# ---------------------------------------------------------------------------
# GAMELOG WATCHDOG — fires update_loop only when a log file actually changes,
# replacing the old 500ms fixed-interval root.after() polling loop.
# Runs in a background thread; all UI calls are routed via root.after(0, …).
# ---------------------------------------------------------------------------
class _GamelogHandler(FileSystemEventHandler):
    def __init__(self, app):
        self._app = app

    def on_modified(self, event):
        # New bytes written to an existing log file
        if not event.is_directory:
            self._app.root.after(0, self._app.update_loop)

    def on_created(self, event):
        # EVE started a new session log — need to pick up the new file path
        if not event.is_directory:
            self._app.root.after(0, self._app.update_loop)


# Contrôleur principal de l'application : GUI Tkinter, lecture des logs EVE et état global
class MiningDashboard:
    # Initialise l'app, charge la config, construit l'UI et démarre la boucle de mise à jour
    def __init__(self):
        try:
            myappid = 'eve.mining.dashboard.v2'
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
        except: pass

        self.tray_icon = None
        self.root = tk.Tk()
        self.root.withdraw()
        self.root.title("EVE Mining Dashboard")
    
        try: self.root.iconbitmap(self.get_resource_path("mining_icon.ico"))
        except: pass
        
        self.app_config = self.load_config()
        
        # Apply themes logic
        self.app_theme = self.app_config.get("theme", "EVE Online (Default)")
        apply_theme_colors(self.app_theme)

        self.root.attributes("-topmost", True)
        self.root.configure(bg=BORDER)
        self.root.overrideredirect(True)
        self.root.attributes("-alpha", WIN_ALPHA)
        self.root.resizable(False, False)

        self._drag_x = 0
        self._drag_y = 0
        self._is_rolled_up = False      # EVE-style window shade: collapsed to title bar only
        self._full_height = 0           # Full height saved before collapsing
        self._hidden_widgets = []       # Content widgets hidden during rollup
        self._last_toggle_time = 0.0    # Cooldown to prevent rapid double-toggle
        self._apply_saved_app_settings()
    
        fleet_cfg = self.app_config.get("fleet", {})
        self.fleet_mode = fleet_cfg.get("enabled", False)
        self.fleet_webhook_url = fleet_cfg.get("webhook_url", "")
    
        saved_geom = self.app_config.get("win_geom", "")
        if saved_geom and "x" in saved_geom:
            try: self.root.geometry(saved_geom)
            except: self.root.geometry(f"{DEFAULT_WIN_W}x{DEFAULT_WIN_H}+100+100")
        else:
            self.root.geometry(f"{DEFAULT_WIN_W}x{DEFAULT_WIN_H}+100+100")
    
        self._glob_cache: List[str] = []
        self._glob_cache_time: float = 0.0
        self._glob_cache_ttl: float = 5.0
        self._history_cache = None          # (per_char_ores, per_char_m3, combined_m3, days)
        self._history_cache_time: float = 0.0
    
        self.all_characters = self.discover_all_characters()
        self.characters = self.get_visible_characters()
        self.load_ship_configs()
    
        for tracker in self.all_characters.values():
            tracker.log_path = self._get_latest_log_for_char(tracker.char_id)
            if tracker.log_path:
                tracker.log_pos = os.path.getsize(tracker.log_path)
    
        self.char_widgets: Dict[str, Dict] = {}
        self.floating_windows: Dict[str, tk.Toplevel] = {}
        # Restore hidden state from config so show/hide survives restarts
        self.hidden_windows: set = set(self.app_config.get("hidden_windows", []))
        self.chars_container = None
        
        self.history_window = None
        self.ship_config_dialogs: Dict[str, tk.Toplevel] = {}
        self.config_dialog: Optional[tk.Toplevel] = None
        self.update_loop_running = True
    
        self.setup_ui()
        self.sync_floating_windows()
        
        self.root.bind("<Button-1>", self._start_drag)
        self.root.bind("<B1-Motion>", self._do_drag)

        if HAS_PYSTRAY: self.setup_tray()

        self.update_loop()

        # Start watching the gamelog directory for file changes; this replaces
        # the old UPDATE_INTERVAL_MS timer — updates now fire only on real I/O
        _gamelog_dir = os.path.dirname(os.path.expanduser(DOCS))
        self._gamelog_observer = None
        if os.path.isdir(_gamelog_dir):
            try:
                _handler = _GamelogHandler(self)
                self._gamelog_observer = Observer()
                self._gamelog_observer.schedule(_handler, _gamelog_dir, recursive=False)
                self._gamelog_observer.start()
            except Exception as _e:
                print(f"[error] MiningDashboard.__init__: watchdog failed to start on {_gamelog_dir}: {_e}")

        self.root.deiconify()
        self.root.after(10, self.set_app_window)
        self.root.mainloop()

    # Reconstruit toute l'interface après un changement de thème en préservant les géométries
    def rebuild_all_ui(self):
        """Called when themes change to regenerate all colors dynamically"""
        # Save geometries before destroying windows
        self.app_config["win_geom"] = self.root.winfo_geometry()
        for cid, top in self.floating_windows.items():
            if top.winfo_exists():
                self.app_config.setdefault("detached_geoms", {})[cid] = top.geometry()
        
        if self.history_window and self.history_window.winfo_exists():
            self.app_config["history_win_geom"] = self.history_window.geometry()
            self.history_window.destroy()
            self.history_window = None

        # Destroy all floating windows
        for top in self.floating_windows.values():
            top.destroy()
        self.floating_windows.clear()
        self.char_widgets.clear()
        # Persist hidden state before clearing so it survives the rebuild
        self._save_hidden_windows()
        hidden_before_rebuild = set(self.hidden_windows)
        self.hidden_windows.clear()
        
        # Destroy Main UI inner frame
        for widget in self.root.winfo_children():
            widget.destroy()

        self.root.configure(bg=BORDER)
        self.setup_ui()
        self.sync_floating_windows()

        # Re-apply hidden state — withdraw any windows that were hidden before the theme change
        for cid in hidden_before_rebuild:
            if cid in self.floating_windows:
                top = self.floating_windows[cid]
                if top.winfo_exists():
                    top.withdraw()
                self.hidden_windows.add(cid)
        self.rebuild_dashboard()

    # Force l'apparition de la fenêtre dans la barre des tâches Windows (WS_EX_APPWINDOW)
    def set_app_window(self):
        GWL_EXSTYLE = -20
        WS_EX_APPWINDOW = 0x00040000
        WS_EX_TOOLWINDOW = 0x00000080
        hwnd = ctypes.windll.user32.GetParent(self.root.winfo_id())
        style = ctypes.windll.user32.GetWindowLongW(hwnd, GWL_EXSTYLE)
        style = style & ~WS_EX_TOOLWINDOW
        style = style | WS_EX_APPWINDOW
        ctypes.windll.user32.SetWindowLongW(hwnd, GWL_EXSTYLE, style)

        self.root.withdraw()
        self.root.deiconify()
        self.root.wm_attributes("-topmost", True)

    # Scanne tous les fichiers logs EVE pour détecter les personnages actifs
    def discover_all_characters(self) -> Dict[str, CharacterTracker]:
        char_names: Dict[str, str] = {}
        char_counts: Dict[str, int] = {}
        for filepath in self._get_all_log_files():
            char_id = self._get_char_id_from_file(filepath)
            if char_id:
                char_counts[char_id] = char_counts.get(char_id, 0) + 1
                if char_id not in char_names:
                    name = self._read_listener_name(filepath)
                    if name: char_names[char_id] = name

        sorted_ids = sorted(char_names.keys(), key=lambda cid: char_counts.get(cid, 0), reverse=True)
        result: Dict[str, CharacterTracker] = {}
        for char_id in sorted_ids:
            result[char_id] = CharacterTracker(char_id, char_names[char_id])
        return result

    # Retourne uniquement les personnages marqués comme visibles dans la config
    def get_visible_characters(self) -> Dict[str, CharacterTracker]:
        # Default to empty instead of loading all characters on first run
        visible_chars = self.app_config.get("visible_characters", None)
        if visible_chars is None:
            return {}
        result = {}
        for char_id, tracker in self.all_characters.items():
            if char_id in visible_chars: result[char_id] = tracker
        return result

    # Sauvegarde la liste des personnages visibles et synchronise les fenêtres
    def save_visible_characters(self, visible_char_ids: List[str]):
        self.app_config["visible_characters"] = visible_char_ids
        self.save_config()
        self.characters = self.get_visible_characters()
        self.sync_floating_windows()
        self.rebuild_dashboard()

    # Persiste l'état des fenêtres masquées dans la config pour survivre aux redémarrages
    def _save_hidden_windows(self):
        # Persist the current hidden set to config so it survives restarts and theme rebuilds
        self.app_config["hidden_windows"] = list(self.hidden_windows)
        self.save_config()

    # Crée ou détruit les fenêtres flottantes selon les personnages visibles actuels
    def sync_floating_windows(self):
        # Destroy windows for chars removed from config (visible_characters unchecked)
        for cid in list(self.floating_windows.keys()):
            if cid not in self.characters:
                top = self.floating_windows[cid]
                self.app_config.setdefault("detached_geoms", {})[cid] = top.geometry()
                top.destroy()
                del self.floating_windows[cid]
                if cid in self.char_widgets:
                    del self.char_widgets[cid]
                # Also clear hidden state since window is gone
                self.hidden_windows.discard(cid)
                self._save_hidden_windows()

        # Create windows for chars newly added to config
        offset_i = len(self.floating_windows)
        for cid, tracker in self.characters.items():
            if cid not in self.floating_windows:
                self.create_floating_window(cid, offset_i)
                offset_i += 1
                # If this char was hidden before (restored from config), withdraw immediately
                if cid in self.hidden_windows:
                    top = self.floating_windows.get(cid)
                    if top and top.winfo_exists():
                        top.withdraw()

    # Crée la fenêtre flottante dédiée à un personnage avec barre de titre, contenu et poignée de redimensionnement
    def create_floating_window(self, char_id: str, offset_i: int):
        tracker = self.all_characters[char_id]
        top = tk.Toplevel(self.root)
        top.configure(bg=BORDER)
        top.overrideredirect(True)
        top.attributes("-topmost", True)
        top.attributes("-alpha", WIN_ALPHA)
        self.floating_windows[char_id] = top
        
        top._user_resized = False

        drag_data = {"x": 0, "y": 0}
        def start_drag(event): drag_data["x"] = event.x; drag_data["y"] = event.y
        def do_drag(event):
            x = top.winfo_x() + event.x - drag_data["x"]
            y = top.winfo_y() + event.y - drag_data["y"]
            top.geometry(f"+{x}+{y}")

        def reset_size(event):
            top._user_resized = False
            top.geometry("")

        top_bar = tk.Frame(top, bg=BG_PANEL, cursor="fleur")
        top_bar.pack(fill="x")
        top_bar.bind("<Button-1>", start_drag)
        top_bar.bind("<B1-Motion>", do_drag)
        top_bar.bind("<Double-Button-1>", reset_size)

        try: idx = list(self.characters.keys()).index(char_id)
        except ValueError: idx = list(self.all_characters.keys()).index(char_id)
        accent = CHAR_ACCENTS[idx % len(CHAR_ACCENTS)]

        title = tk.Label(top_bar, text=f"★ {tracker.char_name.upper()}", fg=accent, bg=BG_PANEL, font=("Consolas", 9, "bold"))
        title.pack(side="left", padx=5, pady=2)
        title.bind("<Button-1>", start_drag)
        title.bind("<B1-Motion>", do_drag)
        title.bind("<Double-Button-1>", reset_size)

        def hide_character():
            # Withdraw the window and mark as hidden — does NOT remove from fleet/config
            # To remove from fleet entirely, use the Config UI checkbox
            top.withdraw()
            self.hidden_windows.add(char_id)
            self._save_hidden_windows()
            self.rebuild_dashboard()  # refresh hub row button to show SHOW

        close_btn = tk.Label(top_bar, text="✕", fg=DIM, bg=BG_PANEL, font=("Consolas", 12, "bold"), cursor="hand2")
        close_btn.pack(side="right", padx=(0, 5))
        close_btn.bind("<Button-1>", lambda e: hide_character())
        close_btn.bind("<Enter>", lambda e, b=close_btn: b.config(fg=RED))
        close_btn.bind("<Leave>", lambda e, b=close_btn: b.config(fg=DIM))

        content_frame = tk.Frame(top, bg=BG)
        content_frame.pack(fill="both", expand=True)

        col_frame, widgets = self._create_char_column(content_frame, tracker, accent, char_id)
        col_frame.pack(fill="both", expand=True)

        self.char_widgets[char_id] = widgets
        self.update_ship_indicator(char_id)

        if tracker.session_active:
            widgets['start_stop_btn'].config(text="■ STOP", fg=RED)

        # ---- RESIZE GRIP ----
        grip = tk.Label(top, text="◢", fg=DIM, bg=BG_PANEL, font=("Consolas", 10), cursor="sizing")
        grip.place(relx=1.0, rely=1.0, anchor="se")

        def start_resize(event):
            top._resize_x = event.x_root
            top._resize_y = event.y_root
            top._orig_width = top.winfo_width()
            top._orig_height = top.winfo_height()

        def do_resize(event):
            dx = event.x_root - top._resize_x
            dy = event.y_root - top._resize_y
            new_w = max(240, top._orig_width + dx)
            new_h = max(150, top._orig_height + dy)
            top.geometry(f"{new_w}x{new_h}")
            top._user_resized = True

        grip.bind("<Button-1>", start_resize)
        grip.bind("<B1-Motion>", do_resize)
        grip.bind("<Enter>", lambda e, g=grip: g.config(fg=CYAN))
        grip.bind("<Leave>", lambda e, g=grip: g.config(fg=DIM))

        # ---- LOAD OR CASCADE GEOMETRY ----
        saved_geom = self.app_config.get("detached_geoms", {}).get(char_id)
        if saved_geom:
            top.geometry(saved_geom)
            if 'x' in saved_geom: top._user_resized = True
        else:
            x = self.root.winfo_x() + 50 + (offset_i * 35)
            y = self.root.winfo_y() + 50 + (offset_i * 35)
            top.geometry(f"+{x}+{y}")

    # Reconstruit le panneau hub principal avec la liste des personnages et leurs boutons
    def rebuild_dashboard(self):
        if self.chars_container:
            for widget in self.chars_container.winfo_children():
                widget.destroy()

        if not self.characters:
            tk.Label(
                self.chars_container,
                text="No characters selected\nClick ⚙ to select characters",
                fg=DIM, bg=BG, font=("Consolas", 9), justify="center"
            ).pack(pady=40)
        else:
            tk.Label(self.chars_container, text="ACTIVE MINING FLEET", fg=DIM, bg=BG, font=("Consolas", 9, "bold")).pack(pady=(5, 5))
            
            hub_outer = tk.Frame(self.chars_container, bg=BORDER, padx=1, pady=1)
            hub_outer.pack(fill="both", expand=True, padx=5, pady=(0, 10))
            
            # Removed static height, letting it expand dynamically based on window size
            char_canvas = tk.Canvas(hub_outer, bg=BG_PANEL, highlightthickness=0)
            char_scrollbar = tk.Scrollbar(hub_outer, orient="vertical", command=char_canvas.yview)
            
            hub_frame = tk.Frame(char_canvas, bg=BG_PANEL)
            char_window = char_canvas.create_window((0, 0), window=hub_frame, anchor="nw")
            
            def on_frame_configure(event):
                char_canvas.configure(scrollregion=char_canvas.bbox("all"))
            def on_canvas_configure(event):
                char_canvas.itemconfig(char_window, width=event.width)
                
            hub_frame.bind("<Configure>", on_frame_configure)
            char_canvas.bind("<Configure>", on_canvas_configure)
            char_canvas.configure(yscrollcommand=char_scrollbar.set)
            
            char_canvas.pack(side="left", fill="both", expand=True)
            char_scrollbar.pack(side="right", fill="y")
            
            def _on_mousewheel(event):
                char_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            
            char_canvas.bind("<Enter>", lambda e: char_canvas.bind_all("<MouseWheel>", _on_mousewheel))
            char_canvas.bind("<Leave>", lambda e: char_canvas.unbind_all("<MouseWheel>"))

            for cid, tracker in self.characters.items():
                try: idx = list(self.characters.keys()).index(cid)
                except ValueError: idx = 0
                accent = CHAR_ACCENTS[idx % len(CHAR_ACCENTS)]
                
                # TIGHT padding
                row_f = tk.Frame(hub_frame, bg=BG_PANEL, padx=4, pady=0)
                row_f.pack(fill="x", pady=0)
                
                tk.Label(row_f, text=f"★ {tracker.char_name.upper()}", fg=accent, bg=BG_PANEL, font=("Consolas", 10, "bold")).pack(side="left")
                
                is_hidden = cid in self.hidden_windows

                def show_window(c_id):
                    # Bring back the withdrawn floating window
                    top = self.floating_windows.get(c_id)
                    if top and top.winfo_exists():
                        top.deiconify()
                        top.lift()
                    self.hidden_windows.discard(c_id)
                    self._save_hidden_windows()
                    self.rebuild_dashboard()

                def hide_window(c_id):
                    # Withdraw floating window — stays in fleet, can be shown again
                    top = self.floating_windows.get(c_id)
                    if top and top.winfo_exists():
                        top.withdraw()
                    self.hidden_windows.add(c_id)
                    self._save_hidden_windows()
                    self.rebuild_dashboard()

                if is_hidden:
                    btn = tk.Label(row_f, text="◉ SHOW", fg=GREEN, bg=BG_PANEL, font=("Consolas", 8, "bold"), cursor="hand2")
                    btn.pack(side="right")
                    btn.bind("<Button-1>", lambda e, c=cid: show_window(c))
                    btn.bind("<Enter>", lambda e, b=btn: b.config(fg=WHITE))
                    btn.bind("<Leave>", lambda e, b=btn: b.config(fg=GREEN))
                else:
                    btn = tk.Label(row_f, text="◉ HIDE", fg=DIM, bg=BG_PANEL, font=("Consolas", 8, "bold"), cursor="hand2")
                    btn.pack(side="right")
                    btn.bind("<Button-1>", lambda e, c=cid: hide_window(c))
                    btn.bind("<Enter>", lambda e, b=btn: b.config(fg=RED))
                    btn.bind("<Leave>", lambda e, b=btn: b.config(fg=DIM))

        # Intentionally removed the geometry snapping so it stays where you resized it!

    # Génère une icône simple (disque coloré) pour la barre système
    def create_tray_image(self):
        image = Image.new('RGBA', (64, 64), (0, 0, 0, 0))
        d = ImageDraw.Draw(image)
        d.ellipse((8, 8, 56, 56), fill=CYAN)
        return image

    # Retourne le chemin absolu d'une ressource, compatible avec PyInstaller (_MEIPASS)
    def get_resource_path(self, relative_path):
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, relative_path)
        return os.path.join(os.path.abspath("."), relative_path)

    # Initialise l'icône de la barre système et son menu contextuel (pystray)
    def setup_tray(self):
        icon_path = self.get_resource_path("mining_icon.ico")
        try:
            if os.path.exists(icon_path): icon_img = Image.open(icon_path)
            else: icon_img = Image.new('RGB', (64, 64), CYAN)
        except Exception:
            icon_img = Image.new('RGB', (64, 64), CYAN)

        menu = pystray.Menu(
            pystray.MenuItem("Show Dashboard", self.show_window),
            pystray.MenuItem("EXIT", self.on_close)
        )
        self.tray_icon = pystray.Icon("mining_dash", icon_img, "EVE Mining Dashboard", menu)
        threading.Thread(target=self.tray_icon.run, daemon=True).start()

    # Affiche la fenêtre principale depuis la barre système
    def show_window(self, icon=None, item=None):
        self.root.after(0, self.root.deiconify)
        self.root.after(0, self.root.lift)
        self.root.after(0, lambda: self.root.attributes("-topmost", True))

    # Retourne tous les fichiers .txt dans le répertoire de logs EVE (scan récursif)
    def _get_all_log_files(self) -> List[str]:
        base_dir = DOCS.rstrip('\\').rstrip('/').rstrip('*')
        pattern = os.path.join(base_dir, '**', '*')
        all_files = glob.glob(pattern, recursive=True)
        return [f for f in all_files if f.lower().endswith('.txt')]

    @staticmethod
    # Extrait l'ID numérique du personnage depuis le nom de fichier log (3e segment séparé par _)
    def _get_char_id_from_file(filepath: str) -> Optional[str]:
        basename = os.path.splitext(os.path.basename(filepath))[0]
        parts = basename.split('_')
        if len(parts) >= 3:
            char_id = parts[2]
            if char_id.isdigit(): return char_id
        return None

    @staticmethod
    # Lit les 15 premières lignes du log pour trouver et retourner le nom du personnage (Listener:)
    def _read_listener_name(filepath: str) -> Optional[str]:
        try:
            with open(filepath, 'r', encoding='utf-8-sig', errors='ignore') as f:
                for i, line in enumerate(f):
                    if i > 15: break
                    match = LISTENER_LINE.search(line)
                    if match: return match.group(1).strip()
        except Exception: pass
        return None

    # Retourne la liste des fichiers log avec mise en cache TTL (5s) pour limiter les appels glob
    def _get_cached_log_files(self) -> List[str]:
        now = time.time()
        if now - self._glob_cache_time > self._glob_cache_ttl:
            base_dir = DOCS.rstrip('\\').rstrip('/').rstrip('*')
            pattern = os.path.join(base_dir, '**', '*')
            self._glob_cache = [f for f in glob.glob(pattern, recursive=True) if f.lower().endswith('.txt')]
            self._glob_cache_time = now
        return self._glob_cache

    # Retourne le fichier log le plus récent pour un personnage donné
    def _get_latest_log_for_char(self, char_id: str) -> Optional[str]:
        files = self._get_cached_log_files()
        char_files = [f for f in files if self._get_char_id_from_file(f) == char_id]
        return max(char_files, key=os.path.getmtime) if char_files else None

    # Enregistre la position initiale pour le glissement de la fenêtre principale
    def _start_drag(self, event):
        widget = event.widget
        if isinstance(widget, tk.Button): return
        # Ignore clicks on clickable labels AND the resize grip
        if isinstance(widget, tk.Label) and widget.cget("cursor") in ["hand2", "sizing"]: return
        self._drag_x = event.x
        self._drag_y = event.y

    # Déplace la fenêtre principale en suivant le mouvement de la souris
    def _do_drag(self, event):
        widget = event.widget
        if isinstance(widget, tk.Button): return
        # Ignore drags on clickable labels AND the resize grip
        if isinstance(widget, tk.Label) and widget.cget("cursor") in ["hand2", "sizing"]: return
        x = self.root.winfo_x() + event.x - self._drag_x
        y = self.root.winfo_y() + event.y - self._drag_y
        self.root.geometry(f"+{x}+{y}")

    # Cache la fenêtre principale dans la barre système
    def minimize_to_tray(self, event=None):
        self.root.withdraw()

    # Bascule le mode "toujours au-dessus" et met à jour l'icône de punaise
    def toggle_pin(self, event=None):
        is_top = self.root.attributes("-topmost")
        new_state = not is_top
        self.root.attributes("-topmost", new_state)
        if new_state: self.pin_icon.config(fg=CYAN)
        else: self.pin_icon.config(fg=DIM)

    # Bascule l'enroulement style EVE : double-clic réduit la fenêtre à la barre de titre seulement
    def toggle_rollup(self, event=None):
        # EVE-style window shade: double-click title bar collapses to just the title bar
        # Guard: ignore double-clicks on interactive widgets (buttons, icons, grip)
        if event and isinstance(event.widget, tk.Label) and event.widget.cget("cursor") in ["hand2", "sizing"]:
            return

        # 500ms cooldown — prevents rapid double-toggle from two quick double-clicks
        now = time.time()
        if now - self._last_toggle_time < 0.5:
            return
        self._last_toggle_time = now

        if not self.inner_frame:
            return

        # Content widgets = everything in inner_frame except the top_bar (first child)
        children = self.inner_frame.winfo_children()
        if not children:
            return
        content_widgets = children[1:]  # skip the top_bar at index 0

        if self._is_rolled_up:
            # Expand: restore all hidden content widgets using their saved pack options
            for widget in self._hidden_widgets:
                info = getattr(widget, "_rollup_pack_info", None)
                if info:
                    try:
                        widget.pack(**info)
                    except Exception:
                        widget.pack(fill="x")

            # Restore min/max size constraints
            self.root.minsize(240, 200)
            self.root.maxsize(9999, 9999)

            # Restore full height, keep current width and position
            if self._full_height > 0:
                w = self.root.winfo_width()
                x = self.root.winfo_x()
                y = self.root.winfo_y()
                self.root.geometry(f"{w}x{self._full_height}+{x}+{y}")

            self._hidden_widgets = []
            self._is_rolled_up = False

        else:
            # Collapse: hide all content widgets below the title bar
            self._full_height = self.root.winfo_height()

            self._hidden_widgets = []
            for widget in content_widgets:
                try:
                    widget._rollup_pack_info = widget.pack_info()
                    self._hidden_widgets.append(widget)
                except Exception:
                    widget._rollup_pack_info = None
                widget.pack_forget()

            self.root.update_idletasks()

            # Lock size to just the title bar — remove min/max constraints first
            collapsed_h = 38
            w = self.root.winfo_width()
            x = self.root.winfo_x()
            y = self.root.winfo_y()
            self.root.minsize(1, 1)
            self.root.maxsize(9999, 9999)
            self.root.geometry(f"{w}x{collapsed_h}+{x}+{y}")
            self.root.minsize(w, collapsed_h)
            self.root.maxsize(w, collapsed_h)

            self._is_rolled_up = True

    # Construit l'interface principale : barre de titre, panneau personnages et bouton historique
    def setup_ui(self) -> None:
        border_frame = tk.Frame(self.root, bg=BORDER, padx=1, pady=1)
        border_frame.pack(fill="both", expand=True)

        self.inner_frame = tk.Frame(border_frame, bg=BG)
        self.inner_frame.pack(fill="both", expand=True)

        top_bar = tk.Frame(self.inner_frame, bg=BG, pady=8, padx=10)
        top_bar.pack(fill="x")
        top_bar.bind("<Double-Button-1>", self.toggle_rollup)

        # Shorter title to leave room for resizing
        title_lbl = tk.Label(top_bar, text="★ MINING ★", fg=CYAN, bg=BG, font=("Consolas", 11, "bold"))
        title_lbl.pack(side="left")
        title_lbl.bind("<Double-Button-1>", self.toggle_rollup)

        close_btn = tk.Label(top_bar, text="✕", fg=DIM, bg=BG, font=("Consolas", 14, "bold"), cursor="hand2")
        close_btn.pack(side="right", padx=(5, 0))
        close_btn.bind("<Button-1>", lambda e: self.on_close())
        close_btn.bind("<Enter>", lambda e: close_btn.config(fg=RED))
        close_btn.bind("<Leave>", lambda e: close_btn.config(fg=DIM))

        min_btn = tk.Label(top_bar, text="–", fg=DIM, bg=BG, font=("Consolas", 14, "bold"), cursor="hand2")
        min_btn.pack(side="right", padx=(5, 0))
        min_btn.bind("<Button-1>", self.minimize_to_tray)
        min_btn.bind("<Enter>", lambda e: min_btn.config(fg=WHITE))
        min_btn.bind("<Leave>", lambda e: min_btn.config(fg=DIM))

        self.config_icon = tk.Label(top_bar, text="⚙", fg=DIM, bg=BG, font=("Consolas", 13, "bold"), cursor="hand2")
        self.config_icon.pack(side="right", padx=(5, 0))
        self.config_icon.bind("<Button-1>", lambda e: self.show_config_dialog())
        self.config_icon.bind("<Enter>", lambda e: self.config_icon.config(fg=CYAN))
        self.config_icon.bind("<Leave>", lambda e: self.config_icon.config(fg=DIM))

        self.pin_icon = tk.Label(top_bar, text="📌", fg=CYAN, bg=BG, font=("Consolas", 11), cursor="hand2")
        self.pin_icon.pack(side="right", padx=(0, 5))
        self.pin_icon.bind("<Button-1>", self.toggle_pin)
        self.pin_icon.bind("<Enter>", lambda e: self.pin_icon.config(bg="#1a2332"))
        self.pin_icon.bind("<Leave>", lambda e: self.pin_icon.config(bg=BG))

        # Pack the history button FIRST and anchor it to the bottom so it never gets clipped
        self.history_button = tk.Button(
            self.inner_frame, text="◈ HISTORY", command=self.show_history, bg=BG_PANEL, fg=CYAN,
            font=("Consolas", 9, "bold"), relief="flat", cursor="hand2", activebackground=BORDER, activeforeground=CYAN
        )
        self.history_button.pack(side="bottom", fill="x", padx=20, pady=(12, 15))

        # Pack the character container SECOND so it dynamically shrinks to fit the remaining space
        self.chars_container = tk.Frame(self.inner_frame, bg=BG)
        self.chars_container.pack(side="top", fill="both", expand=True, padx=5, pady=(5, 0))

        self.rebuild_dashboard()

        # Add Resize Grip for the Main UI Controller
        grip = tk.Label(self.root, text="◢", fg=DIM, bg=BG, font=("Consolas", 10), cursor="sizing")
        grip.place(relx=1.0, rely=1.0, anchor="se")

        def start_resize(event):
            self._resize_x = event.x_root
            self._resize_y = event.y_root
            self._orig_width = self.root.winfo_width()
            self._orig_height = self.root.winfo_height()

        def do_resize(event):
            dx = event.x_root - self._resize_x
            dy = event.y_root - self._resize_y
            new_w = max(240, self._orig_width + dx)
            new_h = max(200, self._orig_height + dy)
            self.root.geometry(f"{new_w}x{new_h}")

        def end_resize(event):
            self.app_config["win_geom"] = self.root.winfo_geometry()
            self.save_config()

        grip.bind("<Button-1>", start_resize)
        grip.bind("<B1-Motion>", do_resize)
        grip.bind("<ButtonRelease-1>", end_resize)
        grip.bind("<Enter>", lambda e, g=grip: g.config(fg=CYAN))
        grip.bind("<Leave>", lambda e, g=grip: g.config(fg=DIM))

    # Crée tous les widgets de la colonne d'un personnage (stats, cargo, boutons, breakdown)
    def _create_char_column(self, parent, tracker: CharacterTracker, accent_color: str, char_id: str):
        col_outer = tk.Frame(parent, bg=BORDER, padx=1, pady=1)
        col_inner = tk.Frame(col_outer, bg=BG_PANEL, padx=10, pady=8)
        col_inner.pack(fill="both", expand=True)

        def show_context_menu(event):
            context_menu = tk.Menu(self.root, tearoff=0, bg=BG_PANEL, fg=WHITE, activebackground=BORDER, activeforeground=CYAN, relief="flat", bd=1)
            context_menu.add_command(label="⚙ Ship Config", command=lambda: self.show_ship_config(char_id))
            try: context_menu.tk_popup(event.x_root, event.y_root)
            finally: context_menu.grab_release()

        col_inner.bind("<Button-3>", show_context_menu)

        name_frame = tk.Frame(col_inner, bg=BG_PANEL)
        name_frame.pack(fill="x", pady=(0, 5))
        name_frame.bind("<Button-3>", show_context_menu)

        profile_label = tk.Label(name_frame, text=f"\u3008{tracker.active_profile}\u3009", fg=GOLD, bg=BG_PANEL, font=("Consolas", 8), cursor="hand2")
        profile_label.pack(side="left", padx=(5, 0))
        profile_label.bind("<Button-1>", lambda e, cid=char_id: self.show_profile_picker(cid, e))
        profile_label.bind("<Button-3>", show_context_menu)
        profile_label.bind("<Enter>", lambda e: profile_label.config(fg=CYAN))
        profile_label.bind("<Leave>", lambda e: profile_label.config(fg=GOLD))

        ship_indicator = tk.Label(name_frame, text="◆", fg=DIM, bg=BG_PANEL, font=("Consolas", 10, "bold"))
        ship_indicator.pack(side="right")
        ship_indicator.bind("<Button-3>", show_context_menu)

        stats_frame = tk.Frame(col_inner, bg=BG_PANEL)
        stats_frame.pack(fill="x")
        stats_frame.bind("<Button-3>", show_context_menu)

        crit_label = tk.Label(stats_frame, text="Crit Bonus: 0.0 m³", fg=GOLD, bg=BG_PANEL, font=("Consolas", 11, "bold"))
        crit_label.pack(anchor="w", pady=2)
        crit_label.bind("<Button-3>", show_context_menu)

        ore_label = tk.Label(stats_frame, text="Total: 0.0 m3", fg=GREEN, bg=BG_PANEL, font=("Consolas", 11, "bold"))
        ore_label.pack(anchor="w", pady=2)
        ore_label.bind("<Button-3>", show_context_menu)

        residue_label = tk.Label(stats_frame, text="Residue: 0.0 m3", fg=RED, bg=BG_PANEL, font=("Consolas", 9))
        residue_label.pack(anchor="w", pady=(0, 2))
        residue_label.bind("<Button-3>", show_context_menu)

        cargo_frame = tk.Frame(col_inner, bg=BG_PANEL)
        cargo_frame.pack(fill="x", pady=(4, 0))

        cargo_text_label = tk.Label(cargo_frame, text="Cargo: 0 / 0 m3", fg=DIM, bg=BG_PANEL, font=("Consolas", 8))
        cargo_text_label.pack(anchor="w")

        cargo_bar_border = tk.Frame(cargo_frame, bg=CYAN, padx=1, pady=1)
        cargo_bar_border.pack(fill="x", pady=(2, 0))
        cargo_canvas = tk.Canvas(cargo_bar_border, height=18, bg="#0a1520", highlightthickness=0)
        cargo_canvas.pack(fill="x")

        cycles_label = tk.Label(cargo_frame, text="Full in: --", fg=DIM, bg=BG_PANEL, font=("Consolas", 8))
        cycles_label.pack(anchor="w", pady=(2, 0))

        control_frame = tk.Frame(col_inner, bg=BG_PANEL)
        control_frame.pack(fill="x", pady=(5, 0))
        control_frame.bind("<Button-3>", show_context_menu)

        start_stop_btn = tk.Button(control_frame, text="▶ START", command=lambda: self.toggle_session(char_id), bg=BG, fg=GREEN, font=("Consolas", 8, "bold"), relief="flat", cursor="hand2", width=10)
        start_stop_btn.pack(side="left", padx=(0, 5))

        empty_btn = tk.Button(control_frame, text="⏏ EMPTY", command=lambda: self.empty_cargo(char_id), bg=BG, fg=CYAN, font=("Consolas", 8, "bold"), relief="flat", cursor="hand2", width=8)
        empty_btn.pack(side="left", padx=(0, 3))

        reset_btn = tk.Button(control_frame, text="↺ RESET", command=lambda: self.reset_session(char_id), bg=BG, fg=RED, font=("Consolas", 8, "bold"), relief="flat", cursor="hand2", width=10)
        reset_btn.pack(side="left")

        fleet_outer = tk.Frame(col_inner, bg=BORDER, padx=1, pady=1)
        fleet_outer.pack(fill="x", pady=(6, 0))

        fleet_frame = tk.Frame(fleet_outer, bg=BG_PANEL, padx=6, pady=4)
        fleet_frame.pack(fill="x")

        has_webhook = self._is_valid_webhook_url()

        copy_btn = tk.Button(fleet_frame, text="⎘ Copy to Clipboard", command=lambda: self.copy_session_report(char_id), bg=BG, fg=GOLD, font=("Consolas", 8, "bold"), relief="flat", cursor="hand2", width=18, state="disabled", disabledforeground=DIM)
        copy_btn.pack(side="left", padx=(0, 4))
        copy_tip = ToolTip(copy_btn, "No mining data yet \u2014 start mining to enable")

        send_btn = tk.Button(fleet_frame, text="▲ Send to Discord", command=lambda: self.show_send_report_dialog(char_id), bg=BG, fg=CYAN, font=("Consolas", 8, "bold"), relief="flat", cursor="hand2", width=18, state="disabled", disabledforeground=DIM)
        send_btn.pack(side="left")

        send_tip_text = "No webhook URL configured \u2014 set it in \u2699 Config" if not has_webhook else "No mining data yet \u2014 start mining to enable"
        send_tip = ToolTip(send_btn, send_tip_text)

        rate_frame = tk.Frame(col_inner, bg=BG_PANEL)
        rate_frame.pack(fill="x", pady=(5, 0))
        rate_frame.bind("<Button-3>", show_context_menu)

        theoretical_label = tk.Label(rate_frame, text="◈ Theoretical: -- m3/s", fg=CYAN, bg=BG_PANEL, font=("Consolas", 9))
        theoretical_label.pack(anchor="w", pady=1)
        theoretical_label.bind("<Button-3>", show_context_menu)

        actual_label = tk.Label(rate_frame, text="◉ Actual: -- m3/s", fg=accent_color, bg=BG_PANEL, font=("Consolas", 9))
        actual_label.pack(anchor="w", pady=1)
        actual_label.bind("<Button-3>", show_context_menu)

        toggle_btn = tk.Button(col_inner, text="v  SESSION BREAKDOWN  v", bg=BG, fg=DIM, font=("Consolas", 8, "bold"), relief="flat", cursor="hand2")
        toggle_btn.pack(fill="x", pady=(5, 0))

        session_container = tk.Frame(col_inner, bg=BG_PANEL)

        def toggle_session_breakdown():
            if session_container.winfo_ismapped():
                session_container.pack_forget()
                toggle_btn.config(text="v  SESSION BREAKDOWN  v")
            else:
                session_container.pack(fill="both", expand=True)
                toggle_btn.config(text="^  HIDE BREAKDOWN  ^")
            
            top = col_outer.winfo_toplevel()
            if not getattr(top, '_user_resized', False):
                top.geometry("")

        toggle_btn.config(command=toggle_session_breakdown)

        summary_outer = tk.Frame(session_container, bg=BORDER, padx=1, pady=1)
        summary_outer.pack(fill="both", pady=(0, 3))
        summary_outer.bind("<Button-3>", show_context_menu)

        summary_box = tk.Label(summary_outer, text="Waiting...", fg=WHITE, bg=BG_PANEL, font=("Consolas", 9), justify="left", padx=8, pady=8)
        summary_box.pack(fill="both")
        summary_box.bind("<Button-3>", show_context_menu)

        widgets = {
            'crit': crit_label, 'ore': ore_label, 'residue': residue_label, 'summary': summary_box,
            'theoretical': theoretical_label, 'actual': actual_label,
            'start_stop_btn': start_stop_btn, 'ship_indicator': ship_indicator,
            'profile_label': profile_label, 'fleet_frame': fleet_frame,
            'copy_btn': copy_btn, 'send_btn': send_btn,
            'copy_tip': copy_tip, 'send_tip': send_tip,
            'cargo_text': cargo_text_label, 'cargo_canvas': cargo_canvas,
            'cycles_label': cycles_label
        }
        return col_outer, widgets

    # Ouvre la fenêtre d'historique de minage avec sélection de la période
    def show_history(self) -> None:
        if self.history_window is None or not self.history_window.winfo_exists():
            self.history_button.config(state="disabled")
            self.history_window = tk.Toplevel(self.root)
            self.history_window.overrideredirect(True)
            self.history_window.configure(bg=BORDER)
            self.history_window.attributes("-topmost", True)
            self.history_window.attributes("-alpha", WIN_ALPHA)

            self._history_drag_x = 0
            self._history_drag_y = 0

            def history_start_drag(event):
                if isinstance(event.widget, tk.Entry): return
                self._history_drag_x = event.x
                self._history_drag_y = event.y

            def history_do_drag(event):
                if isinstance(event.widget, tk.Entry): return
                x = self.history_window.winfo_x() + event.x - self._history_drag_x
                y = self.history_window.winfo_y() + event.y - self._history_drag_y
                self.history_window.geometry(f"+{x}+{y}")

            self.history_window.bind("<Button-1>", history_start_drag)
            self.history_window.bind("<B1-Motion>", history_do_drag)

            saved_geom = self.app_config.get("history_win_geom", "+400+100")
            try:
                if '+' in saved_geom:
                    parts = saved_geom.split('+')
                    if len(parts) >= 3: self.history_window.geometry(f"+{parts[1]}+{parts[2]}")
            except: pass

            border_frame = tk.Frame(self.history_window, bg=BORDER, padx=1, pady=1)
            border_frame.pack(fill="both", expand=True)

            inner_frame = tk.Frame(border_frame, bg=BG)
            inner_frame.pack(fill="both", expand=True)

            top_bar = tk.Frame(inner_frame, bg=BG, pady=10, padx=10)
            top_bar.pack(fill="x")

            tk.Label(top_bar, text="★ MINING HISTORY ★", fg=CYAN, bg=BG, font=("Consolas", 12, "bold")).pack(side="left")

            close_btn = tk.Label(top_bar, text="X", fg=DIM, bg=BG, font=("Consolas", 14, "bold"), cursor="hand2")
            close_btn.pack(side="right")
            close_btn.bind("<Button-1>", lambda e: self.on_history_close())
            close_btn.bind("<Enter>", lambda e: close_btn.config(fg=RED))
            close_btn.bind("<Leave>", lambda e: close_btn.config(fg=DIM))

            control_outer = tk.Frame(inner_frame, bg=BORDER, padx=1, pady=1)
            control_outer.pack(fill="x", padx=10, pady=10)

            control_frame = tk.Frame(control_outer, bg=BG_PANEL, padx=12, pady=12)
            control_frame.pack(fill="x")

            days_frame = tk.Frame(control_frame, bg=BG_PANEL)
            days_frame.pack(fill="x", pady=(0, 10))

            tk.Label(days_frame, text="◆ Days to analyze:", fg=CYAN, bg=BG_PANEL, font=("Consolas", 9, "bold")).pack(side="left", padx=(0, 10))

            max_days = self.get_max_history_days()

            self.history_days_var = tk.StringVar(value=str(HISTORY_DAYS))
            days_entry = tk.Entry(days_frame, textvariable=self.history_days_var, width=10, font=("Consolas", 10), bg=BG, fg=WHITE, insertbackground=CYAN, relief="flat", justify="center")
            days_entry.pack(side="left", padx=5)
            days_entry.bind("<Return>", lambda e: self.calculate_and_display_history(text_widget))

            tk.Label(days_frame, text=f"(max: {HISTORY_DAYS})", fg=GOLD, bg=BG_PANEL, font=("Consolas", 9)).pack(side="left", padx=5)

            refresh_button = tk.Button(control_frame, text="↻ REFRESH STATS", command=lambda: self.calculate_and_display_history(text_widget), bg=BG, fg=GREEN, font=("Consolas", 9, "bold"), relief="flat", cursor="hand2", activebackground=BORDER, activeforeground=GREEN)
            refresh_button.pack(side="left", fill="x", expand=True)

            export_button = tk.Button(control_frame, text="◈ EXPORT EXCEL", command=lambda: self.show_export_menu(export_button), bg=BG, fg=GOLD, font=("Consolas", 9, "bold"), relief="flat", cursor="hand2", activebackground=BORDER, activeforeground=GOLD, state="normal" if HAS_OPENPYXL else "disabled")
            export_button.pack(side="left", fill="x", expand=True, padx=(5, 0))

            text_outer = tk.Frame(inner_frame, bg=BORDER, padx=1, pady=1)
            text_outer.pack(fill="both", expand=True, padx=10, pady=(0, 10))

            text_frame = tk.Frame(text_outer, bg=BG_PANEL)
            text_frame.pack(fill="both", expand=True)

            text_widget = tk.Text(text_frame, bg=BG_PANEL, fg=WHITE, font=("Consolas", 10), relief="flat", padx=10, pady=10, wrap="word", width=40, height=20)
            text_widget.pack(fill="both", expand=True)

            self.calculate_and_display_history(text_widget)

    # Calcule le nombre de jours d'historique disponibles selon les fichiers log présents
    def get_max_history_days(self) -> int:
        try:
            all_files = self._get_all_log_files()
            if not all_files: return 0
            oldest_file = min(all_files, key=os.path.getmtime)
            oldest_timestamp = os.path.getmtime(oldest_file)
            oldest_date = datetime.fromtimestamp(oldest_timestamp)
            days_available = (datetime.now() - oldest_date).days
            return max(1, days_available)
        except Exception: return HISTORY_DAYS

    # Gère la fermeture de la fenêtre historique et réactive le bouton
    def on_history_close(self):
        if self.history_window:
            self.app_config["history_win_geom"] = self.history_window.geometry()
            self.save_config()
            self.history_window.destroy()
            self.history_window = None
            self.history_button.config(state="normal")

    # Lit les logs, calcule les totaux par personnage/minerai et affiche dans le widget texte
    # Résultats mis en cache 30 s pour éviter de relire tous les fichiers à chaque clic
    def calculate_and_display_history(self, text_widget: tk.Text):
        text_widget.config(state="normal")
        text_widget.delete("1.0", tk.END)

        try:
            days = int(self.history_days_var.get())
            if days < 1: days = 1
            max_days = self.get_max_history_days()
            if days > max_days: days = max_days
            self.history_days_var.set(str(days))
        except ValueError:
            days = HISTORY_DAYS
            self.history_days_var.set(str(days))

        now = time.time()
        if (self._history_cache is not None
                and now - self._history_cache_time < 30
                and self._history_cache[3] == days):
            per_char_ores, per_char_m3, combined_m3, days = self._history_cache
        else:
            per_char_ores, per_char_m3, combined_m3, days = self._gather_history_data(days)
            self._history_cache = (per_char_ores, per_char_m3, combined_m3, days)
            self._history_cache_time = now

        w = 38
        result = ""
        total_str = f" ALL CHARS ({days}d): {combined_m3:,.1f} m3"
        pad = max(0, w - len(total_str))
        result += f"+{'=' * w}+\n|{total_str}{' ' * pad}|\n+{'=' * w}+\n\n"

        has_any_data = False
        for char_id, tracker in self.all_characters.items():
            char_name = tracker.char_name.upper()
            char_total = per_char_m3.get(char_id, 0.0)
            ores = per_char_ores.get(char_id, {})

            header = f" {char_name}: {char_total:,.1f} m3 "
            dashes = max(0, w - len(header)) // 2
            result += f"{'-' * dashes}{header}{'-' * dashes}\n"

            if ores:
                has_any_data = True
                for ore_name, volume in sorted(ores.items(), key=lambda x: x[1], reverse=True):
                    result += f"  * {ore_name}: {volume:,.1f} m3\n"
            else: result += "  No mining data.\n"
            result += "\n"

        if not has_any_data and not self.all_characters: result += "No mining data found in this period.\n"

        text_widget.insert("1.0", result)
        text_widget.config(state="disabled")

    # Rassemble et retourne les données de minage agrégées sur N jours pour l'export
    def _gather_history_data(self, days: int):
        try:
            days = int(days)
            if days < 1: days = 1
            max_days = self.get_max_history_days()
            if days > max_days: days = max_days
        except ValueError: days = HISTORY_DAYS

        threshold = datetime.now() - timedelta(days=days)
        per_char_ores: Dict[str, Dict[str, float]] = {}
        per_char_m3: Dict[str, float] = {}
        combined_m3 = 0.0

        all_files = self._get_all_log_files()
        for log_file in all_files:
            if os.path.getmtime(log_file) > threshold.timestamp():
                char_id = self._get_char_id_from_file(log_file)
                if not char_id or char_id not in self.all_characters: continue
                if char_id not in per_char_ores:
                    per_char_ores[char_id] = {}
                    per_char_m3[char_id] = 0.0
                try:
                    with open(log_file, "r", encoding="utf-8-sig", errors="ignore") as f:
                        for line in f:
                            match = REGULAR_MINE_PATTERN.search(line) or CRIT_MINE_PATTERN.search(line)
                            if match:
                                units = float(match.group('amount').replace(",", ""))
                                volume, ore_name = self.get_ore_volume(match.group('ore_type'))
                                total_volume = units * volume
                                per_char_ores[char_id][ore_name] = per_char_ores[char_id].get(ore_name, 0) + total_volume
                                per_char_m3[char_id] = per_char_m3.get(char_id, 0) + total_volume
                                combined_m3 += total_volume
                except Exception: continue
        return per_char_ores, per_char_m3, combined_m3, days

    # Rassemble les données de minage groupées par jour et par personnage pour l'export journalier
    def _gather_daily_history_data(self, days: int):
        try:
            days = int(days)
            if days < 1: days = 1
            max_days = self.get_max_history_days()
            if days > max_days: days = max_days
        except ValueError: days = HISTORY_DAYS

        threshold = datetime.now() - timedelta(days=days)
        per_char_daily: Dict[str, Dict[str, Dict[str, float]]] = {}
        all_ore_names = set()
        all_dates = set()

        all_files = self._get_all_log_files()
        for log_file in all_files:
            if os.path.getmtime(log_file) > threshold.timestamp():
                char_id = self._get_char_id_from_file(log_file)
                if not char_id or char_id not in self.all_characters: continue
                if char_id not in per_char_daily: per_char_daily[char_id] = {}
                try:
                    with open(log_file, "r", encoding="utf-8-sig", errors="ignore") as f:
                        for line in f:
                            match = REGULAR_MINE_PATTERN.search(line) or CRIT_MINE_PATTERN.search(line)
                            if match:
                                ts_match = LOG_TIMESTAMP.match(line)
                                if ts_match: date_str = ts_match.group(1).replace(".", "-")
                                else: continue
                                
                                units = float(match.group('amount').replace(",", ""))
                                volume, ore_name = self.get_ore_volume(match.group('ore_type'))
                                total_volume = units * volume
                                
                                all_ore_names.add(ore_name)
                                all_dates.add(date_str)
                                
                                if date_str not in per_char_daily[char_id]:
                                    per_char_daily[char_id][date_str] = {}
                                per_char_daily[char_id][date_str][ore_name] = per_char_daily[char_id][date_str].get(ore_name, 0) + total_volume
                except Exception: continue

        sorted_dates = sorted(all_dates)
        sorted_ores = sorted(all_ore_names)
        return per_char_daily, sorted_ores, sorted_dates, days

    # Génère le chemin de fichier de l'export Excel avec suffixe et nombre de jours
    def _get_export_path(self, suffix: str, days: int) -> str:
        export_dir = self.app_config.get("app_settings", {}).get("export_dir", "")
        if not export_dir or not os.path.isdir(export_dir):
            export_dir = os.path.dirname(os.path.abspath(CONFIG_FILE))
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        filename = f"mining_{suffix}_{timestamp}_{days}d.xlsx"
        return os.path.join(export_dir, filename)

    # Affiche le menu contextuel avec les options d'export Excel
    def show_export_menu(self, button_widget):
        if not HAS_OPENPYXL:
            messagebox.showwarning("Missing Library", "openpyxl is required for Excel export.\n\npip install openpyxl")
            return

        menu = tk.Menu(self.root, tearoff=0, bg=BG_PANEL, fg=WHITE, activebackground=BORDER, activeforeground=CYAN, relief="flat", bd=1, font=("Consolas", 9))
        menu.add_command(label="◆ Summary Export", command=lambda: self._do_export("summary"))
        menu.add_command(label="◆ Daily Breakdown", command=lambda: self._do_export("daily"))
        menu.add_command(label="◆ Ore Pivot (Chars x Ores)", command=lambda: self._do_export("pivot"))
        menu.add_separator()
        menu.add_command(label="◆ Full Export (All Sheets)", command=lambda: self._do_export("full"))
        
        try:
            x = button_widget.winfo_rootx()
            y = button_widget.winfo_rooty() + button_widget.winfo_height()
            menu.tk_popup(x, y)
        finally: menu.grab_release()

    # Lance l'export Excel du type demandé (summary, daily, pivot ou complet)
    def _do_export(self, export_type: str):
        try: days = int(self.history_days_var.get())
        except (ValueError, AttributeError): days = HISTORY_DAYS

        try:
            if export_type == "summary": filepath = self._export_summary(days)
            elif export_type == "daily": filepath = self._export_daily_breakdown(days)
            elif export_type == "pivot": filepath = self._export_ore_pivot(days)
            elif export_type == "full": filepath = self._export_full(days)
            else: return

            if filepath: messagebox.showinfo("Export Complete", f"Saved to:\n{filepath}", parent=self.history_window or self.root)
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export:\n{str(e)}", parent=self.history_window or self.root)

    # Applique le style d'en-tête EVE (couleur, police bold) à une cellule Excel
    def _apply_eve_header(self, ws, row, col, text, width=None, is_title=False):
        cell = ws.cell(row=row, column=col, value=text)
        if is_title:
            cell.font = Font(name="Consolas", size=12, bold=True, color="3DD8E0")
            cell.fill = PatternFill(start_color="0B0E17", end_color="0B0E17", fill_type="solid")
        else:
            cell.font = Font(name="Consolas", size=10, bold=True, color="3DD8E0")
            cell.fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color="1E3A4A"), top=Side(style="thin", color="1E3A4A"), left=Side(style="thin", color="1E3A4A"), right=Side(style="thin", color="1E3A4A"))
        if width: ws.column_dimensions[get_column_letter(col)].width = width
        return cell

    # Applique le style de cellule de données (nombre, texte, total) à une cellule Excel
    def _apply_eve_data_cell(self, ws, row, col, value, fmt="number", ore_name=None, is_total=False):
        cell = ws.cell(row=row, column=col, value=value)
        if is_total:
            cell.font = Font(name="Consolas", size=10, bold=True, color="FFD700")
            cell.fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        elif ore_name:
            ore_color = _get_ore_excel_color(ore_name)
            cell.font = Font(name="Consolas", size=10, color=ore_color)
            cell.fill = PatternFill(start_color="0B0E17", end_color="0B0E17", fill_type="solid")
        else:
            cell.font = Font(name="Consolas", size=10, color="FFFFFF")
            cell.fill = PatternFill(start_color="0B0E17", end_color="0B0E17", fill_type="solid")

        cell.border = Border(bottom=Side(style="thin", color="1E3A4A"), left=Side(style="thin", color="1E3A4A"), right=Side(style="thin", color="1E3A4A"))
        
        if fmt == "number" and isinstance(value, (int, float)):
            cell.number_format = '#,##0.0'
            cell.alignment = Alignment(horizontal="right")
        elif fmt == "integer" and isinstance(value, (int, float)):
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
        else: cell.alignment = Alignment(horizontal="left")
        return cell

    # Applique la couleur de catégorie du minerai à une cellule d'étiquette Excel
    def _apply_eve_ore_label(self, ws, row, col, ore_name):
        cell = ws.cell(row=row, column=col, value=ore_name)
        ore_color = _get_ore_excel_color(ore_name)
        cell.font = Font(name="Consolas", size=10, bold=True, color=ore_color)
        cell.fill = PatternFill(start_color="0B0E17", end_color="0B0E17", fill_type="solid")
        cell.alignment = Alignment(horizontal="left")
        cell.border = Border(bottom=Side(style="thin", color="1E3A4A"), left=Side(style="thin", color="1E3A4A"), right=Side(style="thin", color="1E3A4A"))
        return cell

    # Applique la couleur d'onglet EVE à une feuille Excel
    def _style_eve_sheet(self, ws): ws.sheet_properties.tabColor = "3DD8E0"

    # Génère le fichier Excel de résumé global par personnage sur N jours
    def _export_summary(self, days: int) -> str:
        per_char_ores, per_char_m3, combined_m3, days = self._gather_history_data(days)
        filepath = self._get_export_path("summary", days)
        wb = Workbook(); ws = wb.active; ws.title = "Combined"
        self._style_eve_sheet(ws)
        
        self._apply_eve_header(ws, 1, 1, f"EVE MINING SUMMARY  --  {days} DAYS", width=30, is_title=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        self._apply_eve_header(ws, 2, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", width=30)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
        self._apply_eve_header(ws, 4, 1, "Character", width=22)
        self._apply_eve_header(ws, 4, 2, "Total m3", width=18)
        self._apply_eve_header(ws, 4, 3, "% Share", width=12)
        
        row = 5
        for char_id, tracker in self.all_characters.items():
            char_total = per_char_m3.get(char_id, 0.0)
            pct = (char_total / combined_m3 * 100) if combined_m3 > 0 else 0
            accent_idx = list(self.all_characters.keys()).index(char_id)
            accent_color = CHAR_ACCENTS[accent_idx % len(CHAR_ACCENTS)].lstrip("#")
            
            cell_name = ws.cell(row=row, column=1, value=tracker.char_name.upper())
            cell_name.font = Font(name="Consolas", size=10, bold=True, color=accent_color)
            cell_name.fill = PatternFill(start_color="0B0E17", end_color="0B0E17", fill_type="solid")
            cell_name.border = Border(bottom=Side(style="thin", color="1E3A4A"), left=Side(style="thin", color="1E3A4A"), right=Side(style="thin", color="1E3A4A"))
            
            self._apply_eve_data_cell(ws, row, 2, char_total)
            pct_cell = self._apply_eve_data_cell(ws, row, 3, pct)
            pct_cell.number_format = '0.0"%"'
            row += 1
        
        row += 1
        total_label = self._apply_eve_data_cell(ws, row, 1, "TOTAL", is_total=True)
        total_label.alignment = Alignment(horizontal="left")
        self._apply_eve_data_cell(ws, row, 2, combined_m3, is_total=True)
        total_pct = self._apply_eve_data_cell(ws, row, 3, 100.0, is_total=True)
        total_pct.number_format = '0.0"%"'
        
        for char_id, tracker in self.all_characters.items():
            ores = per_char_ores.get(char_id, {})
            if not ores: continue
            sheet_name = tracker.char_name[:28].replace("/", "-").replace("\\", "-")
            ws = wb.create_sheet(title=sheet_name)
            self._style_eve_sheet(ws)
            accent_idx = list(self.all_characters.keys()).index(char_id)
            accent_color = CHAR_ACCENTS[accent_idx % len(CHAR_ACCENTS)].lstrip("#")
            
            title_cell = self._apply_eve_header(ws, 1, 1, f"{tracker.char_name.upper()}  --  MINING BREAKDOWN", width=30, is_title=True)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            title_cell.font = Font(name="Consolas", size=12, bold=True, color=accent_color)
            
            char_total = per_char_m3.get(char_id, 0.0)
            self._apply_eve_header(ws, 2, 1, f"Total: {char_total:,.1f} m3", width=30)
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
            self._apply_eve_header(ws, 4, 1, "Ore Type", width=28)
            self._apply_eve_header(ws, 4, 2, "Volume (m3)", width=18)
            
            row = 5
            sorted_ores = sorted(ores.items(), key=lambda x: x[1], reverse=True)
            for ore_name, volume in sorted_ores:
                self._apply_eve_ore_label(ws, row, 1, ore_name)
                self._apply_eve_data_cell(ws, row, 2, volume, ore_name=ore_name)
                row += 1
            row += 1
            self._apply_eve_data_cell(ws, row, 1, "TOTAL", is_total=True)
            self._apply_eve_data_cell(ws, row, 2, char_total, is_total=True)
        
        wb.save(filepath)
        return filepath

    # Génère le fichier Excel avec une feuille par personnage détaillant les totaux journaliers
    def _export_daily_breakdown(self, days: int) -> str:
        per_char_daily, sorted_ores, sorted_dates, days = self._gather_daily_history_data(days)
        filepath = self._get_export_path("daily", days)
        wb = Workbook(); wb.remove(wb.active)

        for char_id, tracker in self.all_characters.items():
            daily_data = per_char_daily.get(char_id, {})
            if not daily_data: continue
            
            sheet_name = tracker.char_name[:28].replace("/", "-").replace("\\", "-")
            ws = wb.create_sheet(title=sheet_name)
            self._style_eve_sheet(ws)
            accent_idx = list(self.all_characters.keys()).index(char_id)
            accent_color = CHAR_ACCENTS[accent_idx % len(CHAR_ACCENTS)].lstrip("#")
            
            char_ores = set()
            for date_ores in daily_data.values(): char_ores.update(date_ores.keys())
            char_ores_sorted = sorted(char_ores)
            if not char_ores_sorted: continue
            
            title_cell = self._apply_eve_header(ws, 1, 1, f"{tracker.char_name.upper()}  --  DAILY BREAKDOWN", is_title=True)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(char_ores_sorted) + 2, 10))
            title_cell.font = Font(name="Consolas", size=12, bold=True, color=accent_color)
            
            self._apply_eve_header(ws, 3, 1, "Date", width=14)
            for j, ore_name in enumerate(char_ores_sorted):
                header_cell = self._apply_eve_header(ws, 3, j + 2, ore_name, width=16)
                header_cell.font = Font(name="Consolas", size=9, bold=True, color=_get_ore_excel_color(ore_name))
            total_col = len(char_ores_sorted) + 2
            self._apply_eve_header(ws, 3, total_col, "DAILY TOTAL", width=16)
            
            row = 4
            grand_total = 0.0
            ore_totals = {ore: 0.0 for ore in char_ores_sorted}
            
            for date_str in sorted_dates:
                if date_str not in daily_data: continue
                date_ores = daily_data[date_str]
                self._apply_eve_data_cell(ws, row, 1, date_str, fmt="text")
                
                day_total = 0.0
                for j, ore_name in enumerate(char_ores_sorted):
                    vol = date_ores.get(ore_name, 0.0)
                    if vol > 0:
                        self._apply_eve_data_cell(ws, row, j + 2, vol, ore_name=ore_name)
                        ore_totals[ore_name] += vol
                        day_total += vol
                    else:
                        empty = ws.cell(row=row, column=j + 2, value="")
                        empty.fill = PatternFill(start_color="0B0E17", end_color="0B0E17", fill_type="solid")
                        empty.border = Border(bottom=Side(style="thin", color="1E3A4A"), left=Side(style="thin", color="1E3A4A"), right=Side(style="thin", color="1E3A4A"))
                
                self._apply_eve_data_cell(ws, row, total_col, day_total, is_total=True)
                grand_total += day_total
                row += 1
            
            row += 1
            self._apply_eve_data_cell(ws, row, 1, "TOTAL", is_total=True)
            for j, ore_name in enumerate(char_ores_sorted):
                self._apply_eve_data_cell(ws, row, j + 2, ore_totals[ore_name], is_total=True)
            self._apply_eve_data_cell(ws, row, total_col, grand_total, is_total=True)
        
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet(title="No Data")
            ws.cell(row=1, column=1, value="No mining data found in this period.")
        wb.save(filepath)
        return filepath

    # Génère le fichier Excel pivot : minerais en lignes, personnages en colonnes
    def _export_ore_pivot(self, days: int) -> str:
        per_char_ores, per_char_m3, combined_m3, days = self._gather_history_data(days)
        filepath = self._get_export_path("pivot", days)
        wb = Workbook(); ws = wb.active; ws.title = "Ore Pivot"
        self._style_eve_sheet(ws)

        all_ores = set()
        for ores in per_char_ores.values(): all_ores.update(ores.keys())
        sorted_ores = sorted(all_ores)
        active_chars = [(cid, t) for cid, t in self.all_characters.items() if cid in per_char_ores and per_char_ores[cid]]

        if not active_chars or not sorted_ores:
            ws.cell(row=1, column=1, value="No mining data found.")
            wb.save(filepath)
            return filepath

        self._apply_eve_header(ws, 1, 1, f"EVE MINING ORE PIVOT  --  {days} DAYS", is_title=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(active_chars) + 2)

        self._apply_eve_header(ws, 3, 1, "Ore Type", width=28)
        for j, (char_id, tracker) in enumerate(active_chars):
            accent_idx = list(self.all_characters.keys()).index(char_id)
            accent_color = CHAR_ACCENTS[accent_idx % len(CHAR_ACCENTS)].lstrip("#")
            header = self._apply_eve_header(ws, 3, j + 2, tracker.char_name.upper(), width=18)
            header.font = Font(name="Consolas", size=10, bold=True, color=accent_color)
        total_col = len(active_chars) + 2
        self._apply_eve_header(ws, 3, total_col, "TOTAL", width=18)

        row = 4
        char_totals = {cid: 0.0 for cid, _ in active_chars}
        grand_total = 0.0

        for ore_name in sorted_ores:
            self._apply_eve_ore_label(ws, row, 1, ore_name)
            ore_row_total = 0.0
            for j, (char_id, tracker) in enumerate(active_chars):
                vol = per_char_ores.get(char_id, {}).get(ore_name, 0.0)
                if vol > 0:
                    self._apply_eve_data_cell(ws, row, j + 2, vol, ore_name=ore_name)
                    char_totals[char_id] += vol
                    ore_row_total += vol
                else:
                    empty = ws.cell(row=row, column=j + 2, value="")
                    empty.fill = PatternFill(start_color="0B0E17", end_color="0B0E17", fill_type="solid")
                    empty.border = Border(bottom=Side(style="thin", color="1E3A4A"), left=Side(style="thin", color="1E3A4A"), right=Side(style="thin", color="1E3A4A"))
            
            self._apply_eve_data_cell(ws, row, total_col, ore_row_total, is_total=True)
            grand_total += ore_row_total
            row += 1

        row += 1
        self._apply_eve_data_cell(ws, row, 1, "TOTAL", is_total=True)
        for j, (char_id, tracker) in enumerate(active_chars):
            self._apply_eve_data_cell(ws, row, j + 2, char_totals[char_id], is_total=True)
        self._apply_eve_data_cell(ws, row, total_col, grand_total, is_total=True)

        wb.save(filepath)
        return filepath

    # Génère un fichier Excel complet avec tous les onglets (résumé, journalier, pivot)
    def _export_full(self, days: int) -> str:
        per_char_ores, per_char_m3, combined_m3, days_used = self._gather_history_data(days)
        per_char_daily, sorted_ores_daily, sorted_dates, _ = self._gather_daily_history_data(days)
        filepath = self._get_export_path("full", days_used)
        wb = Workbook(); ws = wb.active; ws.title = "Summary"
        self._style_eve_sheet(ws)
        
        self._apply_eve_header(ws, 1, 1, f"EVE MINING FULL REPORT  --  {days_used} DAYS", width=30, is_title=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        self._apply_eve_header(ws, 2, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", width=30)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
        self._apply_eve_header(ws, 4, 1, "Character", width=22)
        self._apply_eve_header(ws, 4, 2, "Total m3", width=18)
        self._apply_eve_header(ws, 4, 3, "% Share", width=12)
        
        row = 5
        for char_id, tracker in self.all_characters.items():
            char_total = per_char_m3.get(char_id, 0.0)
            pct = (char_total / combined_m3 * 100) if combined_m3 > 0 else 0
            accent_idx = list(self.all_characters.keys()).index(char_id)
            accent_color = CHAR_ACCENTS[accent_idx % len(CHAR_ACCENTS)].lstrip("#")
            
            cell_name = ws.cell(row=row, column=1, value=tracker.char_name.upper())
            cell_name.font = Font(name="Consolas", size=10, bold=True, color=accent_color)
            cell_name.fill = PatternFill(start_color="0B0E17", end_color="0B0E17", fill_type="solid")
            cell_name.border = Border(bottom=Side(style="thin", color="1E3A4A"), left=Side(style="thin", color="1E3A4A"), right=Side(style="thin", color="1E3A4A"))
            
            self._apply_eve_data_cell(ws, row, 2, char_total)
            pct_cell = self._apply_eve_data_cell(ws, row, 3, pct)
            pct_cell.number_format = '0.0"%"'
            row += 1
        
        row += 1
        self._apply_eve_data_cell(ws, row, 1, "TOTAL", is_total=True)
        self._apply_eve_data_cell(ws, row, 2, combined_m3, is_total=True)
        total_pct = self._apply_eve_data_cell(ws, row, 3, 100.0, is_total=True)
        total_pct.number_format = '0.0"%"'
        
        row += 3
        for char_id, tracker in self.all_characters.items():
            ores = per_char_ores.get(char_id, {})
            if not ores: continue
            accent_idx = list(self.all_characters.keys()).index(char_id)
            accent_color = CHAR_ACCENTS[accent_idx % len(CHAR_ACCENTS)].lstrip("#")
            
            sep_cell = ws.cell(row=row, column=1, value=f"--- {tracker.char_name.upper()} ---")
            sep_cell.font = Font(name="Consolas", size=10, bold=True, color=accent_color)
            sep_cell.fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
            for ore_name, volume in sorted(ores.items(), key=lambda x: x[1], reverse=True):
                self._apply_eve_ore_label(ws, row, 1, ore_name)
                self._apply_eve_data_cell(ws, row, 2, volume, ore_name=ore_name)
                row += 1
            row += 1

        ws2 = wb.create_sheet(title="Ore Pivot")
        self._style_eve_sheet(ws2)
        all_ores_set = set()
        for ores in per_char_ores.values(): all_ores_set.update(ores.keys())
        all_ores_sorted = sorted(all_ores_set)
        active_chars = [(cid, t) for cid, t in self.all_characters.items() if cid in per_char_ores and per_char_ores[cid]]

        if active_chars and all_ores_sorted:
            self._apply_eve_header(ws2, 1, 1, "ORE PIVOT", is_title=True)
            ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(active_chars) + 2)
            self._apply_eve_header(ws2, 3, 1, "Ore Type", width=28)
            for j, (char_id, tracker) in enumerate(active_chars):
                accent_idx = list(self.all_characters.keys()).index(char_id)
                accent_color = CHAR_ACCENTS[accent_idx % len(CHAR_ACCENTS)].lstrip("#")
                h = self._apply_eve_header(ws2, 3, j + 2, tracker.char_name.upper(), width=18)
                h.font = Font(name="Consolas", size=10, bold=True, color=accent_color)
            total_col = len(active_chars) + 2
            self._apply_eve_header(ws2, 3, total_col, "TOTAL", width=18)
            
            row2 = 4
            for ore_name in all_ores_sorted:
                self._apply_eve_ore_label(ws2, row2, 1, ore_name)
                ore_total = 0.0
                for j, (char_id, _) in enumerate(active_chars):
                    vol = per_char_ores.get(char_id, {}).get(ore_name, 0.0)
                    if vol > 0:
                        self._apply_eve_data_cell(ws2, row2, j + 2, vol, ore_name=ore_name)
                        ore_total += vol
                    else:
                        empty = ws2.cell(row=row2, column=j + 2, value="")
                        empty.fill = PatternFill(start_color="0B0E17", end_color="0B0E17", fill_type="solid")
                        empty.border = Border(bottom=Side(style="thin", color="1E3A4A"), left=Side(style="thin", color="1E3A4A"), right=Side(style="thin", color="1E3A4A"))
                self._apply_eve_data_cell(ws2, row2, total_col, ore_total, is_total=True)
                row2 += 1
            
            row2 += 1
            self._apply_eve_data_cell(ws2, row2, 1, "TOTAL", is_total=True)
            for j, (char_id, _) in enumerate(active_chars):
                self._apply_eve_data_cell(ws2, row2, j + 2, per_char_m3.get(char_id, 0.0), is_total=True)
            self._apply_eve_data_cell(ws2, row2, total_col, combined_m3, is_total=True)

        for char_id, tracker in self.all_characters.items():
            daily_data = per_char_daily.get(char_id, {})
            if not daily_data: continue
            
            sheet_name = f"Daily-{tracker.char_name[:24]}".replace("/", "-").replace("\\", "-")
            ws3 = wb.create_sheet(title=sheet_name)
            self._style_eve_sheet(ws3)
            accent_idx = list(self.all_characters.keys()).index(char_id)
            accent_color = CHAR_ACCENTS[accent_idx % len(CHAR_ACCENTS)].lstrip("#")
            
            char_ores = set()
            for date_ores in daily_data.values(): char_ores.update(date_ores.keys())
            char_ores_sorted = sorted(char_ores)
            if not char_ores_sorted: continue
            
            title_cell = self._apply_eve_header(ws3, 1, 1, f"{tracker.char_name.upper()} - DAILY", is_title=True)
            ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(char_ores_sorted) + 2, 10))
            title_cell.font = Font(name="Consolas", size=12, bold=True, color=accent_color)
            
            self._apply_eve_header(ws3, 3, 1, "Date", width=14)
            for j, ore_name in enumerate(char_ores_sorted):
                h = self._apply_eve_header(ws3, 3, j + 2, ore_name, width=16)
                h.font = Font(name="Consolas", size=9, bold=True, color=_get_ore_excel_color(ore_name))
            tcol = len(char_ores_sorted) + 2
            self._apply_eve_header(ws3, 3, tcol, "DAILY TOTAL", width=16)
            
            row3 = 4
            ore_totals = {ore: 0.0 for ore in char_ores_sorted}
            grand = 0.0
            for date_str in sorted_dates:
                if date_str not in daily_data: continue
                date_ores = daily_data[date_str]
                self._apply_eve_data_cell(ws3, row3, 1, date_str, fmt="text")
                day_total = 0.0
                for j, ore_name in enumerate(char_ores_sorted):
                    vol = date_ores.get(ore_name, 0.0)
                    if vol > 0:
                        self._apply_eve_data_cell(ws3, row3, j + 2, vol, ore_name=ore_name)
                        ore_totals[ore_name] += vol
                        day_total += vol
                    else:
                        empty = ws3.cell(row=row3, column=j + 2, value="")
                        empty.fill = PatternFill(start_color="0B0E17", end_color="0B0E17", fill_type="solid")
                        empty.border = Border(bottom=Side(style="thin", color="1E3A4A"), left=Side(style="thin", color="1E3A4A"), right=Side(style="thin", color="1E3A4A"))
                self._apply_eve_data_cell(ws3, row3, tcol, day_total, is_total=True)
                grand += day_total
                row3 += 1
            
            row3 += 1
            self._apply_eve_data_cell(ws3, row3, 1, "TOTAL", is_total=True)
            for j, ore_name in enumerate(char_ores_sorted):
                self._apply_eve_data_cell(ws3, row3, j + 2, ore_totals[ore_name], is_total=True)
            self._apply_eve_data_cell(ws3, row3, tcol, grand, is_total=True)

        wb.save(filepath)
        return filepath

    @lru_cache(maxsize=256)
    # Retourne le volume en m³ et le nom normalisé d'un type de minerai depuis le cache SDE
    def get_ore_volume(self, raw_name: str) -> Tuple[float, str]:
        clean_name = raw_name.strip().rstrip('.')
        if clean_name in ORE_VOLUMES: return ORE_VOLUMES[clean_name], clean_name
        clean_lower = clean_name.lower()
        for base_ore, volume in ORE_VOLUMES.items():
            if base_ore.lower() in clean_lower: return volume, clean_name
        return 1.0, clean_name

    # Restaure les paramètres sauvegardés (thème, transparence, topmost) depuis la config
    def _apply_saved_app_settings(self):
        global DOCS, CRIT_SOUND_FILE, UPDATE_INTERVAL_MS, HISTORY_DAYS, CRITICAL_HIT_KEYWORD, PLAY_CRIT_SOUND, WIN_ALPHA
        app_settings = self.app_config.get("app_settings", {})
        if not app_settings: return
        if "docs_path" in app_settings: DOCS = app_settings["docs_path"]
        if "crit_sound_file" in app_settings: CRIT_SOUND_FILE = app_settings["crit_sound_file"]
        if "update_interval_ms" in app_settings: UPDATE_INTERVAL_MS = max(250, int(app_settings["update_interval_ms"]))
        if "history_days" in app_settings: HISTORY_DAYS = max(1, int(app_settings["history_days"]))
        if "crit_keyword" in app_settings: CRITICAL_HIT_KEYWORD = app_settings["crit_keyword"]
        if "play_crit_sound" in app_settings: PLAY_CRIT_SOUND = app_settings["play_crit_sound"]
        if "win_alpha" in app_settings:
            WIN_ALPHA = max(0.2, min(1.0, float(app_settings["win_alpha"])))
            self._apply_alpha(WIN_ALPHA)

    # Applique la transparence à toutes les fenêtres ouvertes (root, flottantes, historique, config)
    def _apply_alpha(self, value: float) -> None:
        v = max(0.2, min(1.0, float(value)))
        try: self.root.attributes("-alpha", v)
        except Exception: pass
        for top in getattr(self, "floating_windows", {}).values():
            try:
                if top.winfo_exists(): top.attributes("-alpha", v)
            except Exception: pass
        hw = getattr(self, "history_window", None)
        if hw and hw.winfo_exists():
            try: hw.attributes("-alpha", v)
            except Exception: pass
        for dlg in getattr(self, "ship_config_dialogs", {}).values():
            try:
                if dlg.winfo_exists(): dlg.attributes("-alpha", v)
            except Exception: pass
        cd = getattr(self, "config_dialog", None)
        if cd and cd.winfo_exists():
            try: cd.attributes("-alpha", v)
            except Exception: pass

    # Charge la configuration JSON depuis le disque (retourne un dict vide si absent)
    def load_config(self) -> Dict:
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r") as f: return json.load(f)
            except Exception: return {}
        return {}

    # Sauvegarde la configuration actuelle dans le fichier JSON
    def save_config(self) -> None:
        self.app_config["win_geom"] = self.root.winfo_geometry()
        try:
            with open(CONFIG_FILE, "w") as f:
                json.dump(self.app_config, f, indent=2)
        except Exception: pass

    # Fermeture propre : sauvegarde la géométrie/config, arrête le tray et quitte
    def on_close(self) -> None:
        self.update_loop_running = False
        obs = getattr(self, '_gamelog_observer', None)
        if obs is not None:
            try:
                obs.stop()
                obs.join(timeout=2.0)
            except Exception as e:
                print(f"[error] on_close: could not stop gamelog observer: {e}")
        tray = getattr(self, 'tray_icon', None)
        if tray is not None:
            try: tray.stop()
            except Exception: pass
        
        if getattr(self, 'history_window', None) and self.history_window.winfo_exists():
            self.on_history_close()
        
        for cid, top in self.floating_windows.items():
            if top.winfo_exists():
                self.app_config.setdefault("detached_geoms", {})[cid] = top.geometry()
                
        self.save_config()
        try: self.root.destroy()
        except Exception: pass
        os._exit(0)

    # Boucle principale : lit les nouveaux logs et met à jour l'UI toutes les 500ms
    def update_loop(self) -> None:
        if not self.update_loop_running: return
        try:
            for char_id, tracker in self.characters.items():
                latest_log = self._get_latest_log_for_char(char_id)
                if latest_log and latest_log != tracker.log_path:
                    tracker.log_path = latest_log
                    tracker.log_pos = 0

                if tracker.log_path and os.path.exists(tracker.log_path):
                    try:
                        with open(tracker.log_path, "r", encoding="utf-8-sig", errors="ignore") as f:
                            f.seek(tracker.log_pos)
                            new_data = f.read()
                            new_pos = f.tell()
                            if new_data:
                                was_active = tracker.session_active
                                self._process_log_data(tracker, new_data)
                                if was_active: tracker.log_pos = new_pos
                            elif tracker.session_active:
                                tracker.log_pos = new_pos
                    except Exception: pass
            self._update_ui_labels()
        except Exception: pass
        # No reschedule here — _GamelogHandler.on_modified/on_created calls
        # root.after(0, self.update_loop) whenever a gamelog file changes

    # Parse les lignes de log EVE pour extraire minage normal, critique, compression et résidu
    def _process_log_data(self, tracker: CharacterTracker, data: str) -> None:
        if not tracker.session_active: return
        crit_processed = False
        last_mined_volume = 0.0
        last_mined_ore = "Unknown"

        for line in data.splitlines():
            line_lower = line.lower()   # computed once, reused for notify check
            if "(notify)" in line_lower:
                if any(kw.lower() in line_lower for kw in AUTO_PAUSE_KEYWORDS):
                    tracker.session_elapsed_offset += time.time() - tracker.session_start_time
                    tracker.session_active = False
                    if tracker.char_id in self.char_widgets:
                        self.char_widgets[tracker.char_id]['start_stop_btn'].config(text="▶ START", fg=GREEN)
                    return
                
            compression_match = COMPRESSION_PATTERN.search(line)
            if compression_match:
                ore_type = compression_match.group('ore_type')
                compressed_amount = float(compression_match.group('amount').replace(",", ""))
                compression_ratio = COMPRESSION_RATIOS.get(ore_type, 100)
                original_units = compressed_amount * compression_ratio
                volume_per_unit, ore_name = self.get_ore_volume(ore_type)
                total_raw_volume = original_units * volume_per_unit
                compressed_volume = total_raw_volume / compression_ratio if compression_ratio > 0 else total_raw_volume
                tracker.current_cargo = max(0.0, tracker.current_cargo - total_raw_volume + compressed_volume)
                tracker.compression_log[ore_name] = tracker.compression_log.get(ore_name, 0) + total_raw_volume
                continue

            if not MINING_LINE.match(line): continue

            regular_match = REGULAR_MINE_PATTERN.search(line)
            if regular_match:
                units = float(regular_match.group('amount').replace(",", ""))
                volume, ore_name = self.get_ore_volume(regular_match.group('ore_type'))
                total_volume = units * volume
                tracker.total_m3 += total_volume
                tracker.current_cargo += total_volume
                tracker.ore_summary[ore_name] = tracker.ore_summary.get(ore_name, 0) + total_volume
                last_mined_volume = volume
                last_mined_ore = ore_name

            if CRITICAL_HIT_KEYWORD in line and not crit_processed:
                crit_match = CRIT_MINE_PATTERN.search(line)
                if crit_match:
                    units = float(crit_match.group('amount').replace(",", ""))
                    ore_type_clean = crit_match.group('ore_type').split('<')[0].split('\r')[0].split('\n')[0].strip()
                    volume, ore_name = self.get_ore_volume(ore_type_clean)
                    total_volume = units * volume
                    tracker.total_m3 += total_volume
                    tracker.current_cargo += total_volume
                    tracker.ore_summary[ore_name] = tracker.ore_summary.get(ore_name, 0) + total_volume
                    tracker.crit_count += 1
                    tracker.crit_m3 += total_volume
                    crit_processed = True
                    self.trigger_crit_alert()
                    last_mined_volume = volume
                    last_mined_ore = ore_name

            residue_match = RESIDUE_PATTERN.search(line)
            if residue_match and last_mined_volume > 0:
                units = float(residue_match.group('amount').replace(",", ""))
                total_volume = units * last_mined_volume
                tracker.total_residue_m3 += total_volume
                tracker.residue_summary[last_mined_ore] = tracker.residue_summary.get(last_mined_ore, 0) + total_volume

    # Met à jour tous les labels, barres cargo et stats de débit de l'interface
    def _update_ui_labels(self) -> None:
        char_widgets = self.char_widgets  # Local reference for faster lookup
        has_webhook = self._is_valid_webhook_url()
        
        for char_id, tracker in self.characters.items():
            if char_id not in char_widgets: continue
            w = char_widgets[char_id]
            
            w['crit'].config(text=f"Crit Bonus: {tracker.crit_m3:,.1f} m³ ({tracker.crit_count})")
            session_m3 = tracker.total_m3 - tracker.session_start_m3
            w['ore'].config(text=f"Total: {session_m3:,.1f} m3")
            w['residue'].config(text=f"Residue: {tracker.total_residue_m3:,.1f} m3")
    
            ore_summary = tracker.ore_summary
            if ore_summary:
                summary = "\n".join([f"{ore_name}: {volume:,.1f} m3" for ore_name, volume in ore_summary.items()])
            else:
                summary = "Waiting..."
            w['summary'].config(text=summary)
    
            has_data = bool(ore_summary)
            if has_data:
                w['copy_btn'].config(state="normal", fg=GOLD)
                w['copy_tip'].update_text("Copy session report to clipboard")
                if has_webhook:
                    w['send_btn'].config(state="normal", fg=CYAN)
                    w['send_tip'].update_text("Send session report to Discord webhook")
                else:
                    w['send_btn'].config(state="disabled", fg=DIM)
                    w['send_tip'].update_text("No webhook URL configured \u2014 set it in \u2699 Config")
            else:
                w['copy_btn'].config(state="disabled", fg=DIM)
                w['copy_tip'].update_text("No mining data yet \u2014 start mining to enable")
                w['send_btn'].config(state="disabled", fg=DIM)
                if not has_webhook: w['send_tip'].update_text("No mining data and no webhook URL configured")
                else: w['send_tip'].update_text("No mining data yet \u2014 start mining to enable")

            capacity = tracker.get_active_capacity()
            current = tracker.current_cargo

            if capacity > 0:
                pct = min(1.0, current / capacity)
                w['cargo_text'].config(text=f"Cargo: {current:,.0f} / {capacity:,.0f} m3 ({int(pct*100)}%)")
                bar_color = RED if pct >= 1.0 else CYAN
                draw_neon_bar(w['cargo_canvas'], pct, bar_color=bar_color)

                rate = tracker.get_total_theoretical_m3_per_sec()
                if rate > 0 and pct < 1.0:
                    remaining = capacity - current
                    seconds_left = remaining / rate
                    modules = tracker.get_active_modules()
                    cycle_time = modules[0].cycle_time if modules and modules[0].cycle_time > 0 else 0
                    time_str = f"{int(seconds_left//60)}m {int(seconds_left%60)}s"

                    if cycle_time > 0:
                        cycles_left = seconds_left / cycle_time
                        w['cycles_label'].config(text=f"Full in: ~{cycles_left:.1f} cycles ({time_str})")
                    else: w['cycles_label'].config(text=f"Full in: {time_str}")
                elif pct >= 1.0: w['cycles_label'].config(text="Full in: FULL")
                else: w['cycles_label'].config(text="Full in: --")
            else:
                w['cargo_text'].config(text=f"Cargo: {current:,.0f} m3 (No Cap Set)")
                draw_neon_bar(w['cargo_canvas'], 0)
                w['cycles_label'].config(text="Full in: (Set Capacity in Config)")

            self._update_rate_stats(char_id, tracker, w)

    # Joue le son d'alerte critique et envoie une notification bureau si disponible
    def trigger_crit_alert(self) -> None:
        if HAS_NOTIFICATION:
            try: notification.notify(title="MINING", message="Critical Hit!", timeout=1)
            except Exception: pass
        # Use native winsound instead of playsound dependency
        if PLAY_CRIT_SOUND and os.path.exists(CRIT_SOUND_FILE):
            try:
                winsound.PlaySound(CRIT_SOUND_FILE, winsound.SND_FILENAME | winsound.SND_ASYNC)
            except Exception: pass

    # Démarre ou arrête la session de minage pour un personnage donné
    def toggle_session(self, char_id: str):
        tracker = self.all_characters[char_id]
        widgets = self.char_widgets[char_id]
        tracker.session_active = not tracker.session_active

        if tracker.session_active:
            is_resume = bool(tracker.ore_summary)
            tracker.session_start_time = time.time()
            if not is_resume:
                tracker.session_start_m3 = tracker.total_m3
                tracker.session_elapsed_offset = 0.0

            if tracker.log_path and os.path.exists(tracker.log_path):
                try:
                    with open(tracker.log_path, "r", encoding="utf-8-sig", errors="ignore") as f:
                        f.seek(tracker.log_pos)
                        backlog = f.read()
                        if backlog: self._process_log_data(tracker, backlog)
                        tracker.log_pos = f.tell()
                except Exception: pass

            if not tracker.session_active:
                widgets['start_stop_btn'].config(text="▶ START", fg=GREEN)
                return

            widgets['start_stop_btn'].config(text="■ STOP", fg=RED)
            rate = tracker.get_total_theoretical_m3_per_sec()
            if rate > 0: widgets['actual'].config(text=f"◉ Actual: {rate:.2f} m3/s ({rate * 3600:,.0f} m3/hr)")
        else:
            tracker.session_elapsed_offset += time.time() - tracker.session_start_time
            widgets['start_stop_btn'].config(text="▶ START", fg=GREEN)

    # Remet le cargo à zéro pour simuler un déchargement en station
    def empty_cargo(self, char_id: str):
        tracker = self.all_characters[char_id]
        tracker.current_cargo = 0.0
        self._update_ui_labels()

    # Réinitialise complètement la session : compteurs, cargo et position dans le log
    def reset_session(self, char_id: str):
        tracker = self.all_characters[char_id]
        widgets = self.char_widgets[char_id]

        if tracker.session_active:
            tracker.session_active = False
            widgets['start_stop_btn'].config(text="▶ START", fg=GREEN)

        tracker.current_cargo = 0.0
        tracker.crit_count = 0
        tracker.crit_m3 = 0.0
        tracker.total_m3 = 0.0
        tracker.ore_summary = {}
        tracker.compression_log = {}
        tracker.total_residue_m3 = 0.0
        tracker.residue_summary = {}
        tracker.session_start_time = time.time()
        tracker.session_start_m3 = 0.0
        tracker.session_elapsed_offset = 0.0

        widgets['crit'].config(text="Crit Bonus: 0.0 m³")
        widgets['ore'].config(text="Total: 0.0 m3")
        widgets['residue'].config(text="Residue: 0.0 m3")
        widgets['summary'].config(text="Waiting...")
        widgets['actual'].config(text="◉ Actual: -- m3/s")
        widgets['copy_btn'].config(state="disabled", fg=DIM)
        widgets['copy_tip'].update_text("No mining data yet \u2014 start mining to enable")
        widgets['send_btn'].config(state="disabled", fg=DIM)
        if not self._is_valid_webhook_url(): widgets['send_tip'].update_text("No mining data and no webhook URL configured")
        else: widgets['send_tip'].update_text("No mining data yet \u2014 start mining to enable")

    # Charge les profils de vaisseau (modules, drones, implants, cargo) depuis la config JSON
    def load_ship_configs(self):
        ship_configs = self.app_config.get("ship_configs", {})
        for char_id, tracker in self.all_characters.items():
            if char_id in ship_configs:
                cfg = ship_configs[char_id]
                if "profiles" in cfg:
                    tracker.ship_profiles = {}
                    tracker.drone_profiles = {}
                    tracker.implant_profiles = {}
                    tracker.cargo_profiles = {}
                    for profile_name, profile_data in cfg["profiles"].items():
                        modules = [MiningModule.from_dict(mod_data) for mod_data in profile_data.get("modules", [])]
                        tracker.ship_profiles[profile_name] = modules
                        
                        drone_data = profile_data.get("drones", {})
                        if drone_data: tracker.drone_profiles[profile_name] = MiningDrone.from_dict(drone_data)
                        else: tracker.drone_profiles[profile_name] = MiningDrone()
                        
                        tracker.implant_profiles[profile_name] = profile_data.get("highwall_implant", False)
                        tracker.cargo_profiles[profile_name] = profile_data.get("cargo_capacity", 0.0)

                    tracker.active_profile = cfg.get("active_profile", "Default")
                    if tracker.active_profile not in tracker.ship_profiles:
                        if tracker.ship_profiles: tracker.active_profile = list(tracker.ship_profiles.keys())[0]
                        else:
                            tracker.active_profile = "Default"
                            tracker.ship_profiles["Default"] = []
                            tracker.drone_profiles["Default"] = MiningDrone()
                            tracker.implant_profiles["Default"] = False
                            tracker.cargo_profiles["Default"] = 0.0
                elif "modules" in cfg:
                    modules = [MiningModule.from_dict(mod_data) for mod_data in cfg.get("modules", [])]
                    tracker.ship_profiles = {"Default": modules}
                    tracker.drone_profiles = {"Default": MiningDrone()}
                    tracker.implant_profiles = {"Default": False}
                    tracker.cargo_profiles["Default"] = 0.0
                    tracker.active_profile = "Default"
                elif "yield_per_cycle" in cfg:
                    old_yield = cfg.get("yield_per_cycle", 0.0)
                    old_cycle = cfg.get("cycle_time", 60.0)
                    if old_yield > 0:
                        module = MiningModule(name="Module 1", yield_per_cycle=old_yield, cycle_time=old_cycle, enabled=True)
                        tracker.ship_profiles = {"Default": [module]}
                        tracker.drone_profiles = {"Default": MiningDrone()}
                        tracker.implant_profiles = {"Default": False}
                        tracker.cargo_profiles["Default"] = 0.0
                        tracker.active_profile = "Default"

    # Sauvegarde tous les profils de vaisseau de tous les personnages dans la config JSON
    def save_ship_configs(self):
        ship_configs = {}
        for char_id, tracker in self.all_characters.items():
            profiles_data = {}
            for profile_name, modules in tracker.ship_profiles.items():
                drone = tracker.drone_profiles.get(profile_name, MiningDrone())
                implant = tracker.implant_profiles.get(profile_name, False)
                capacity = tracker.cargo_profiles.get(profile_name, 0.0)
                profiles_data[profile_name] = {
                    "modules": [m.to_dict() for m in modules],
                    "drones": drone.to_dict(),
                    "highwall_implant": implant,
                    "cargo_capacity": capacity
                }
            ship_configs[char_id] = {
                "active_profile": tracker.active_profile,
                "profiles": profiles_data
            }
        self.app_config["ship_configs"] = ship_configs
        self.save_config()

    # Ouvre le dialogue de configuration du vaisseau : modules, drones, implant et cargo
    def show_ship_config(self, char_id: str):
        if char_id in self.ship_config_dialogs and self.ship_config_dialogs[char_id].winfo_exists():
            self.ship_config_dialogs[char_id].lift()
            self.ship_config_dialogs[char_id].focus_force()
            return

        tracker = self.all_characters[char_id]
        dialog = tk.Toplevel(self.root)
        dialog.configure(bg=BORDER)
        dialog.overrideredirect(True)
        dialog.wm_attributes("-topmost", 1)
        dialog.attributes("-alpha", WIN_ALPHA)
        dialog.resizable(False, False)
        self.ship_config_dialogs[char_id] = dialog

        _drag_x = [0]
        _drag_y = [0]

        def start_drag(event):
            if isinstance(event.widget, (tk.Entry, tk.OptionMenu)): return
            _drag_x[0] = event.x
            _drag_y[0] = event.y

        def do_drag(event):
            if isinstance(event.widget, (tk.Entry, tk.OptionMenu)): return
            x = dialog.winfo_x() + event.x - _drag_x[0]
            y = dialog.winfo_y() + event.y - _drag_y[0]
            dialog.geometry(f"+{x}+{y}")

        config_key = f"ship_config_{char_id}_geom"
        saved_geom = self.app_config.get(config_key, "+300+200")

        border_frame = tk.Frame(dialog, bg=BORDER, padx=1, pady=1)
        border_frame.pack(fill="both", expand=True)

        main_frame = tk.Frame(border_frame, bg=BG_PANEL, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        main_frame.bind("<Button-1>", start_drag)
        main_frame.bind("<B1-Motion>", do_drag)

        top_bar = tk.Frame(main_frame, bg=BG_PANEL)
        top_bar.grid(row=0, column=0, columnspan=4, sticky="ew", pady=(0, 15))
        top_bar.bind("<Button-1>", start_drag)
        top_bar.bind("<B1-Motion>", do_drag)

        title_label = tk.Label(top_bar, text=f"⚙ {tracker.char_name.upper()} — SHIP FITTING", fg=CYAN, bg=BG_PANEL, font=("Consolas", 10, "bold"))
        title_label.pack(side="left")
        title_label.bind("<Button-1>", start_drag)
        title_label.bind("<B1-Motion>", do_drag)

        def close_dialog():
            try:
                x, y = dialog.winfo_x(), dialog.winfo_y()
                self.app_config[config_key] = f"+{x}+{y}"
                self.save_config()
            except Exception: pass
            if char_id in self.ship_config_dialogs: del self.ship_config_dialogs[char_id]
            dialog.destroy()

        close_btn = tk.Label(top_bar, text="✕", fg=DIM, bg=BG_PANEL, font=("Consolas", 14, "bold"), cursor="hand2")
        close_btn.pack(side="right")
        close_btn.bind("<Button-1>", lambda e: close_dialog())
        close_btn.bind("<Enter>", lambda e: close_btn.config(fg=RED))
        close_btn.bind("<Leave>", lambda e: close_btn.config(fg=DIM))

        profile_frame = tk.Frame(main_frame, bg=BG_PANEL)
        profile_frame.grid(row=1, column=0, columnspan=4, sticky="ew", pady=(0, 15))
        
        tk.Label(profile_frame, text="◆ SHIP PROFILE:", fg=CYAN, bg=BG_PANEL, font=("Consolas", 9, "bold")).pack(side="left", padx=(0, 10))

        current_profile = tk.StringVar(value=tracker.active_profile)
        profile_menu = tk.OptionMenu(profile_frame, current_profile, *tracker.get_profile_names())
        profile_menu.config(bg=BG, fg=WHITE, font=("Consolas", 9), activebackground=BORDER, activeforeground=CYAN, highlightthickness=0, relief="flat")
        profile_menu["menu"].config(bg=BG_PANEL, fg=WHITE, activebackground=BORDER, activeforeground=CYAN)
        profile_menu.pack(side="left", padx=5)

        btn_new = tk.Button(profile_frame, text="+ NEW", command=lambda: self.create_new_profile(tracker, current_profile, profile_menu, module_vars, update_preview, dialog, drone_vars, implant_var, cargo_cap_var), bg=BG, fg=GREEN, font=("Consolas", 8, "bold"), relief="flat", cursor="hand2", width=6)
        btn_new.pack(side="left", padx=2)

        btn_rename = tk.Button(profile_frame, text="✎ RENAME", command=lambda: self.rename_current_profile(tracker, current_profile, profile_menu, dialog), bg=BG, fg=CYAN, font=("Consolas", 8, "bold"), relief="flat", cursor="hand2", width=8)
        btn_rename.pack(side="left", padx=2)

        btn_delete = tk.Button(profile_frame, text="✕ DELETE", command=lambda: self.delete_current_profile(tracker, current_profile, profile_menu, module_vars, update_preview, dialog, drone_vars, implant_var, cargo_cap_var), bg=BG, fg=RED, font=("Consolas", 8, "bold"), relief="flat", cursor="hand2", width=8)
        btn_delete.pack(side="left", padx=2)

        modules_label = tk.Label(main_frame, text="◆ MINING MODULES", fg=CYAN, bg=BG_PANEL, font=("Consolas", 9, "bold"))
        modules_label.grid(row=2, column=0, columnspan=4, sticky="w", pady=(0, 10))

        tk.Label(main_frame, text="", bg=BG_PANEL, width=3).grid(row=3, column=0)
        tk.Label(main_frame, text="Module Name", fg=DIM, bg=BG_PANEL, font=("Consolas", 8)).grid(row=3, column=1, padx=5)
        tk.Label(main_frame, text="Yield (m3/cycle)", fg=DIM, bg=BG_PANEL, font=("Consolas", 8)).grid(row=3, column=2, padx=5)
        tk.Label(main_frame, text="Cycle Time (s)", fg=DIM, bg=BG_PANEL, font=("Consolas", 8)).grid(row=3, column=3, padx=5)

        module_vars = []

        def load_profile_into_ui(profile_name: str):
            modules = tracker.ship_profiles.get(profile_name, [])
            while len(modules) < MAX_MODULES: modules.append(MiningModule())
            for i, (module, mv) in enumerate(zip(modules[:MAX_MODULES], module_vars)):
                mv['enabled'].set(module.enabled and module.is_configured())
                mv['name'].set(module.name if module.is_configured() else "")
                mv['yield'].set(str(module.yield_per_cycle) if module.yield_per_cycle > 0 else "")
                mv['cycle'].set(str(module.cycle_time) if module.cycle_time > 0 else "")
                if not (module.enabled and module.is_configured()): mv['name_entry'].config(state="disabled")
                else: mv['name_entry'].config(state="normal")
            
            drone = tracker.drone_profiles.get(profile_name, MiningDrone())
            drone_vars['count'].set(str(drone.count) if drone.count > 0 else "")
            drone_vars['yield'].set(str(drone.yield_per_cycle) if drone.yield_per_cycle > 0 else "")
            drone_vars['cycle'].set(str(drone.cycle_time) if drone.cycle_time > 0 else "")
            implant_var.set(tracker.implant_profiles.get(profile_name, False))
            
            active_capacity = tracker.cargo_profiles.get(profile_name, 0.0)
            cargo_cap_var.set(str(active_capacity) if active_capacity > 0 else "")
            
            update_preview()

        active_modules = tracker.get_active_modules()
        while len(active_modules) < MAX_MODULES: active_modules.append(MiningModule())

        for i in range(MAX_MODULES):
            module = active_modules[i] if i < len(active_modules) else MiningModule()
            row = 4 + i
            enabled_var = tk.BooleanVar(value=module.enabled and module.is_configured())
            enabled_cb = tk.Checkbutton(main_frame, variable=enabled_var, bg=BG_PANEL, activebackground=BG_PANEL, selectcolor=WHITE, highlightthickness=0)
            enabled_cb.grid(row=row, column=0, padx=2, pady=3)

            name_var = tk.StringVar(value=module.name if module.is_configured() else "")
            name_entry = tk.Entry(main_frame, textvariable=name_var, width=12, font=("Consolas", 9), bg=BG, fg=WHITE, insertbackground=CYAN, disabledbackground=BG, disabledforeground=DIM)
            name_entry.grid(row=row, column=1, padx=5, pady=3)

            yield_var = tk.StringVar(value=str(module.yield_per_cycle) if module.yield_per_cycle > 0 else "")
            yield_entry = tk.Entry(main_frame, textvariable=yield_var, width=12, font=("Consolas", 9), bg=BG, fg=WHITE, insertbackground=CYAN)
            yield_entry.grid(row=row, column=2, padx=5, pady=3)

            cycle_var = tk.StringVar(value=str(module.cycle_time) if module.cycle_time > 0 else "")
            cycle_entry = tk.Entry(main_frame, textvariable=cycle_var, width=12, font=("Consolas", 9), bg=BG, fg=WHITE, insertbackground=CYAN)
            cycle_entry.grid(row=row, column=3, padx=5, pady=3)

            def update_name_state(name_e=name_entry, enabled_v=enabled_var, name_v=name_var):
                if enabled_v.get(): name_e.config(state="normal")
                else: name_v.set(""); name_e.config(state="disabled")

            if not (module.enabled and module.is_configured()): name_entry.config(state="disabled")
            enabled_var.trace_add('write', lambda *args, fn=update_name_state: fn())

            module_vars.append({'enabled': enabled_var, 'name': name_var, 'yield': yield_var, 'cycle': cycle_var, 'name_entry': name_entry})

        sep_row = 4 + MAX_MODULES

        drones_label = tk.Label(main_frame, text="◆ MINING DRONES", fg=CYAN, bg=BG_PANEL, font=("Consolas", 9, "bold"))
        drones_label.grid(row=sep_row, column=0, columnspan=4, sticky="w", pady=(15, 5))

        drone_row = sep_row + 1
        tk.Label(main_frame, text="Count:", fg=DIM, bg=BG_PANEL, font=("Consolas", 8)).grid(row=drone_row, column=0, columnspan=2, sticky="e", padx=(0, 5), pady=3)
        active_drone = tracker.get_active_drones()
        drone_count_var = tk.StringVar(value=str(active_drone.count) if active_drone.count > 0 else "")
        drone_count_entry = tk.Entry(main_frame, textvariable=drone_count_var, width=6, font=("Consolas", 9), bg=BG, fg=WHITE, insertbackground=CYAN)
        drone_count_entry.grid(row=drone_row, column=2, sticky="w", padx=5, pady=3)
        tk.Label(main_frame, text="(max 5)", fg=GOLD, bg=BG_PANEL, font=("Consolas", 8)).grid(row=drone_row, column=3, sticky="w", padx=5, pady=3)

        drone_row2 = drone_row + 1
        tk.Label(main_frame, text="Yield (m3/cycle):", fg=DIM, bg=BG_PANEL, font=("Consolas", 8)).grid(row=drone_row2, column=0, columnspan=2, sticky="e", padx=(0, 5), pady=3)
        drone_yield_var = tk.StringVar(value=str(active_drone.yield_per_cycle) if active_drone.yield_per_cycle > 0 else "")
        drone_yield_entry = tk.Entry(main_frame, textvariable=drone_yield_var, width=12, font=("Consolas", 9), bg=BG, fg=WHITE, insertbackground=CYAN)
        drone_yield_entry.grid(row=drone_row2, column=2, sticky="w", padx=5, pady=3)
        tk.Label(main_frame, text="per drone", fg=GOLD, bg=BG_PANEL, font=("Consolas", 8)).grid(row=drone_row2, column=3, sticky="w", padx=5, pady=3)

        drone_row3 = drone_row + 2
        tk.Label(main_frame, text="Cycle Time (s):", fg=DIM, bg=BG_PANEL, font=("Consolas", 8)).grid(row=drone_row3, column=0, columnspan=2, sticky="e", padx=(0, 5), pady=3)
        drone_cycle_var = tk.StringVar(value=str(active_drone.cycle_time) if active_drone.cycle_time > 0 else "")
        drone_cycle_entry = tk.Entry(main_frame, textvariable=drone_cycle_var, width=12, font=("Consolas", 9), bg=BG, fg=WHITE, insertbackground=CYAN)
        drone_cycle_entry.grid(row=drone_row3, column=2, sticky="w", padx=5, pady=3)

        drone_vars = {'count': drone_count_var, 'yield': drone_yield_var, 'cycle': drone_cycle_var}

        implant_row = drone_row + 3
        implant_label = tk.Label(main_frame, text="◆ MINING IMPLANT", fg=CYAN, bg=BG_PANEL, font=("Consolas", 9, "bold"))
        implant_label.grid(row=implant_row, column=0, columnspan=4, sticky="w", pady=(15, 5))

        implant_cb_row = implant_row + 1
        implant_var = tk.BooleanVar(value=tracker.get_active_implant())
        implant_cb = tk.Checkbutton(main_frame, variable=implant_var, bg=BG_PANEL, activebackground=BG_PANEL, selectcolor=WHITE, highlightthickness=0)
        implant_cb.grid(row=implant_cb_row, column=0, sticky="e", padx=(0, 0), pady=3)

        implant_text = tk.Label(main_frame, text="Highwall MX-1005", fg=WHITE, bg=BG_PANEL, font=("Consolas", 9))
        implant_text.grid(row=implant_cb_row, column=1, sticky="w", padx=(0, 5), pady=3)

        implant_note = tk.Label(main_frame, text="+5% mining yield (modules only)", fg=GOLD, bg=BG_PANEL, font=("Consolas", 8))
        implant_note.grid(row=implant_cb_row, column=2, columnspan=2, sticky="w", padx=5, pady=3)

        cargo_row = implant_cb_row + 1
        cargo_label = tk.Label(main_frame, text="◆ CARGO HOLD", fg=CYAN, bg=BG_PANEL, font=("Consolas", 9, "bold"))
        cargo_label.grid(row=cargo_row, column=0, columnspan=4, sticky="w", pady=(15, 5))
        
        active_capacity = tracker.get_active_capacity()
        cargo_row1 = cargo_row + 1
        tk.Label(main_frame, text="Capacity (m3):", fg=DIM, bg=BG_PANEL, font=("Consolas", 8)).grid(row=cargo_row1, column=0, columnspan=2, sticky="e", padx=(0, 5), pady=3)
        cargo_cap_var = tk.StringVar(value=str(active_capacity) if active_capacity > 0 else "")
        cargo_cap_entry = tk.Entry(main_frame, textvariable=cargo_cap_var, width=12, font=("Consolas", 9), bg=BG, fg=WHITE, insertbackground=CYAN)
        cargo_cap_entry.grid(row=cargo_row1, column=2, sticky="w", padx=5, pady=3)
        tk.Label(main_frame, text="(e.g. 11500)", fg=GOLD, bg=BG_PANEL, font=("Consolas", 8)).grid(row=cargo_row1, column=3, sticky="w", padx=5, pady=3)

        preview_row = cargo_row1 + 1
        preview_frame = tk.Frame(main_frame, bg=BG, padx=10, pady=8)
        preview_frame.grid(row=preview_row, column=0, columnspan=4, sticky="ew", pady=(15, 10))

        preview_label = tk.Label(preview_frame, text="◈ Theoretical: -- m3/s (-- m3/hr)", fg=CYAN, bg=BG, font=("Consolas", 9, "bold"))
        preview_label.pack()

        def update_preview(*args):
            module_m3_per_sec = 0.0
            active_count = 0
            for mv in module_vars:
                if mv['enabled'].get():
                    try:
                        y = float(mv['yield'].get()) if mv['yield'].get() else 0.0
                        c = float(mv['cycle'].get()) if mv['cycle'].get() else 0.0
                        if y > 0 and c > 0:
                            module_m3_per_sec += (y / c)
                            active_count += 1
                    except ValueError: pass
            
            has_implant = implant_var.get()
            if has_implant and module_m3_per_sec > 0: module_m3_per_sec *= 1.054 
            total_m3_per_sec = module_m3_per_sec

            drone_count = 0
            try:
                dc = int(drone_vars['count'].get()) if drone_vars['count'].get() else 0
                dy = float(drone_vars['yield'].get()) if drone_vars['yield'].get() else 0.0
                dcy = float(drone_vars['cycle'].get()) if drone_vars['cycle'].get() else 0.0
                if dc > 0 and dy > 0 and dcy > 0:
                    dc = max(0, min(dc, MiningDrone.MAX_DRONES))
                    total_m3_per_sec += (dy / dcy) * dc
                    drone_count = dc
            except ValueError: pass
            
            if total_m3_per_sec > 0:
                display_sec = round(total_m3_per_sec, 1)
                parts = []
                if active_count > 0: parts.append(f"{active_count} mod{'s' if active_count > 1 else ''}")
                if drone_count > 0: parts.append(f"{drone_count} drone{'s' if drone_count > 1 else ''}")
                if has_implant: parts.append("HW")
                detail = " + ".join(parts)
                preview_label.config(text=f"◈ Theoretical: {display_sec:.1f} m3/s ({display_sec * 3600:,.0f} m3/hr) [{detail}]")
            else:
                preview_label.config(text="◈ Theoretical: -- m3/s (configure ship)")

        for mv in module_vars:
            mv['enabled'].trace_add('write', update_preview)
            mv['yield'].trace_add('write', update_preview)
            mv['cycle'].trace_add('write', update_preview)

        drone_vars['count'].trace_add('write', update_preview)
        drone_vars['yield'].trace_add('write', update_preview)
        drone_vars['cycle'].trace_add('write', update_preview)
        implant_var.trace_add('write', update_preview)

        def on_profile_change(*args):
            new_profile = current_profile.get()
            if new_profile != tracker.active_profile:
                save_current_profile_to_tracker()
                tracker.active_profile = new_profile
                load_profile_into_ui(new_profile)

        current_profile.trace_add('write', on_profile_change)

        def save_current_profile_to_tracker():
            modules = []
            for mv in module_vars:
                mod = MiningModule(
                    name=mv['name'].get(),
                    yield_per_cycle=float(mv['yield'].get()) if mv['yield'].get() else 0.0,
                    cycle_time=float(mv['cycle'].get()) if mv['cycle'].get() else 0.0,
                    enabled=mv['enabled'].get()
                )
                modules.append(mod)
            tracker.ship_profiles[tracker.active_profile] = modules
            
            try:
                dc = int(drone_vars['count'].get()) if drone_vars['count'].get() else 0
                dy = float(drone_vars['yield'].get()) if drone_vars['yield'].get() else 0.0
                dcy = float(drone_vars['cycle'].get()) if drone_vars['cycle'].get() else 0.0
                dc = max(0, min(dc, MiningDrone.MAX_DRONES))
            except ValueError: dc, dy, dcy = 0, 0.0, 0.0
            tracker.drone_profiles[tracker.active_profile] = MiningDrone(dc, dy, dcy)
            tracker.implant_profiles[tracker.active_profile] = implant_var.get()
            
            try: cap = float(cargo_cap_var.get()) if cargo_cap_var.get() else 0.0
            except ValueError: cap = 0.0
            tracker.cargo_profiles[tracker.active_profile] = cap

        update_preview()

        btn_frame = tk.Frame(main_frame, bg=BG_PANEL)
        btn_frame.grid(row=preview_row + 1, column=0, columnspan=4, pady=(10, 0))

        def save_and_close():
            try:
                save_current_profile_to_tracker()
                self.save_ship_configs()
                if char_id in self.char_widgets:
                    self.update_ship_indicator(char_id)
                    self.update_profile_label(char_id)

                try:
                    x, y = dialog.winfo_x(), dialog.winfo_y()
                    self.app_config[config_key] = f"+{x}+{y}"
                    self.save_config()
                except Exception: pass
                if char_id in self.ship_config_dialogs: del self.ship_config_dialogs[char_id]
                dialog.destroy()
            except ValueError:
                messagebox.showerror("Invalid Input", "Please enter valid numbers")

        tk.Button(btn_frame, text="✔ SAVE", command=save_and_close, bg=BG, fg=GREEN, font=("Consolas", 9, "bold"), relief="flat", cursor="hand2", width=10).pack(side="left", padx=5)
        tk.Button(btn_frame, text="✕ CANCEL", command=close_dialog, bg=BG, fg=RED, font=("Consolas", 9, "bold"), relief="flat", cursor="hand2", width=10).pack(side="left", padx=5)

        dialog.update_idletasks()
        try:
            if '+' in saved_geom:
                parts = saved_geom.split('+')
                if len(parts) >= 3: dialog.geometry(f"+{parts[1]}+{parts[2]}")
                else: dialog.geometry("+300+200")
            else: dialog.geometry("+300+200")
        except Exception: dialog.geometry("+300+200")
        dialog.update()

        def initial_focus():
            if dialog.winfo_exists():
                dialog.lift()
                dialog.focus_force()

        dialog.after(150, initial_focus)

    # Affiche une boîte de dialogue modale centrée pour saisir une chaîne de texte
    def _ask_string_centered(self, title, prompt, parent_dialog, initialvalue=""):
        result = [None]
        dlg = tk.Toplevel(parent_dialog)
        dlg.title(title)
        dlg.configure(bg=BG_PANEL)
        dlg.resizable(False, False)
        dlg.transient(parent_dialog)
        dlg.grab_set()

        tk.Label(dlg, text=prompt, bg=BG_PANEL, fg=WHITE, font=("Consolas", 10)).pack(padx=20, pady=(15, 5))
        entry = tk.Entry(dlg, font=("Consolas", 10), width=30, bg=BG, fg=WHITE, insertbackground=WHITE)
        entry.pack(padx=20, pady=5)
        if initialvalue:
            entry.insert(0, initialvalue)
            entry.select_range(0, tk.END)

        btn_frame = tk.Frame(dlg, bg=BG_PANEL)
        btn_frame.pack(pady=(5, 15))

        def on_ok(event=None):
            result[0] = entry.get().strip()
            dlg.destroy()
        def on_cancel(event=None): dlg.destroy()

        tk.Button(btn_frame, text="OK", command=on_ok, bg=BG, fg=GREEN, font=("Consolas", 9, "bold"), relief="flat", width=8, cursor="hand2").pack(side="left", padx=5)
        tk.Button(btn_frame, text="Cancel", command=on_cancel, bg=BG, fg=RED, font=("Consolas", 9, "bold"), relief="flat", width=8, cursor="hand2").pack(side="left", padx=5)

        entry.bind("<Return>", on_ok)
        entry.bind("<Escape>", on_cancel)

        dlg.update_idletasks()
        pw, ph = parent_dialog.winfo_width(), parent_dialog.winfo_height()
        px, py = parent_dialog.winfo_x(), parent_dialog.winfo_y()
        dw, dh = dlg.winfo_reqwidth(), dlg.winfo_reqheight()
        x, y = px + (pw - dw) // 2, py + (ph - dh) // 2
        dlg.geometry(f"+{x}+{y}")
        dlg.wm_attributes("-topmost", 1)
        entry.focus_set()
        dlg.wait_window()
        return result[0] if result[0] else None

    # Crée un nouveau profil vide via dialogue et réinitialise les champs de l'UI
    def create_new_profile(self, tracker: CharacterTracker, current_profile_var: tk.StringVar, profile_menu: tk.OptionMenu, module_vars: List, update_preview_fn, parent_dialog=None, drone_vars=None, implant_var=None, cargo_cap_var=None):
        parent = parent_dialog or self.root
        new_name = self._ask_string_centered("New Profile", "Enter name for new ship profile:", parent)
        if new_name:
            if tracker.create_profile(new_name):
                menu = profile_menu["menu"]
                menu.delete(0, "end")
                for profile in tracker.get_profile_names():
                    menu.add_command(label=profile, command=lambda value=profile: current_profile_var.set(value))

                current_profile_var.set(new_name)
                tracker.active_profile = new_name

                for mv in module_vars:
                    mv['enabled'].set(False)
                    mv['name'].set("")
                    mv['yield'].set("")
                    mv['cycle'].set("")
                    mv['name_entry'].config(state="disabled")

                if drone_vars:
                    drone_vars['count'].set("")
                    drone_vars['yield'].set("")
                    drone_vars['cycle'].set("")
                if implant_var: implant_var.set(False)
                if cargo_cap_var is not None: cargo_cap_var.set("")
                
                update_preview_fn()
            else:
                messagebox.showerror("Error", "Profile name already exists or is invalid")

    # Renomme le profil actif via dialogue et met à jour le menu déroulant
    def rename_current_profile(self, tracker: CharacterTracker, current_profile_var: tk.StringVar, profile_menu: tk.OptionMenu, parent_dialog=None):
        old_name = tracker.active_profile
        if len(tracker.ship_profiles) == 1:
            messagebox.showwarning("Cannot Rename", "You must have at least one profile", parent=parent_dialog or self.root)
            return
        
        parent = parent_dialog or self.root
        new_name = self._ask_string_centered("Rename Profile", f"Rename '{old_name}' to:", parent, initialvalue=old_name)
        if new_name and new_name != old_name:
            if tracker.rename_profile(old_name, new_name):
                menu = profile_menu["menu"]
                menu.delete(0, "end")
                for profile in tracker.get_profile_names():
                    menu.add_command(label=profile, command=lambda value=profile: current_profile_var.set(value))
                current_profile_var.set(new_name)
            else:
                messagebox.showerror("Error", "Profile name already exists or is invalid", parent=parent)

    # Supprime le profil actif après confirmation et bascule sur le suivant disponible
    def delete_current_profile(self, tracker: CharacterTracker, current_profile_var: tk.StringVar, profile_menu: tk.OptionMenu, module_vars: List, update_preview_fn, parent_dialog=None, drone_vars=None, implant_var=None, cargo_cap_var=None):
        profile_to_delete = tracker.active_profile
        parent = parent_dialog or self.root
        if len(tracker.ship_profiles) == 1:
            messagebox.showwarning("Cannot Delete", "You must have at least one profile", parent=parent)
            return
        
        result = messagebox.askyesno("Delete Profile", f"Are you sure you want to delete profile '{profile_to_delete}'?", parent=parent)
        if result:
            if tracker.delete_profile(profile_to_delete):
                menu = profile_menu["menu"]
                menu.delete(0, "end")
                for profile in tracker.get_profile_names():
                    menu.add_command(label=profile, command=lambda value=profile: current_profile_var.set(value))
                
                current_profile_var.set(tracker.active_profile)
                modules = tracker.get_active_modules()
                while len(modules) < MAX_MODULES: modules.append(MiningModule())
                
                for i, (module, mv) in enumerate(zip(modules[:MAX_MODULES], module_vars)):
                    mv['enabled'].set(module.enabled and module.is_configured())
                    mv['name'].set(module.name if module.is_configured() else "")
                    mv['yield'].set(str(module.yield_per_cycle) if module.yield_per_cycle > 0 else "")
                    mv['cycle'].set(str(module.cycle_time) if module.cycle_time > 0 else "")
                    if not (module.enabled and module.is_configured()): mv['name_entry'].config(state="disabled")
                    else: mv['name_entry'].config(state="normal")
                
                if drone_vars:
                    drone = tracker.get_active_drones()
                    drone_vars['count'].set(str(drone.count) if drone.count > 0 else "")
                    drone_vars['yield'].set(str(drone.yield_per_cycle) if drone.yield_per_cycle > 0 else "")
                    drone_vars['cycle'].set(str(drone.cycle_time) if drone.cycle_time > 0 else "")
                
                if implant_var: implant_var.set(tracker.get_active_implant())
                update_preview_fn()

    # Affiche un menu contextuel pour choisir ou créer un profil depuis l'interface principale
    def show_profile_picker(self, char_id: str, event):
        tracker = self.all_characters.get(char_id)
        if not tracker: return

        menu = tk.Menu(self.root, tearoff=0, bg=BG_PANEL, fg=WHITE, activebackground=BORDER, activeforeground=CYAN, relief="flat", bd=1, font=("Consolas", 9))
        for profile_name in tracker.get_profile_names():
            label = f"\u2714 {profile_name}" if profile_name == tracker.active_profile else f"   {profile_name}"
            menu.add_command(label=label, command=lambda pn=profile_name: self.switch_profile_from_main(char_id, pn))

        menu.add_separator()
        menu.add_command(label="\u2795 Create New Profile\u2026", command=lambda: self.create_profile_from_main(char_id))
        try: menu.tk_popup(event.x_root, event.y_root)
        finally: menu.grab_release()

    # Change le profil actif d'un personnage depuis l'interface principale et met à jour l'UI
    def switch_profile_from_main(self, char_id: str, profile_name: str):
        tracker = self.all_characters.get(char_id)
        if not tracker or profile_name == tracker.active_profile: return
        tracker.active_profile = profile_name
        self.save_ship_configs()
        self.update_profile_label(char_id)
        self.update_ship_indicator(char_id)
        if char_id in self.char_widgets:
            self._update_rate_stats(char_id, tracker, self.char_widgets[char_id])

    # Crée un nouveau profil directement depuis l'interface principale (sans ouvrir ship config)
    def create_profile_from_main(self, char_id: str):
        tracker = self.all_characters.get(char_id)
        if not tracker: return
        new_name = self._ask_string_centered("New Profile", "Enter name for new ship profile:", self.root)
        if new_name:
            if tracker.create_profile(new_name):
                tracker.active_profile = new_name
                self.save_ship_configs()
                self.update_profile_label(char_id)
                self.update_ship_indicator(char_id)
            else:
                messagebox.showerror("Error", "Profile name already exists or is invalid", parent=self.root)

    # Met à jour l'indicateur ◆ selon que le vaisseau est configuré ou non
    def update_ship_indicator(self, char_id: str):
        tracker = self.all_characters[char_id]
        if char_id not in self.char_widgets: return
        widgets = self.char_widgets[char_id]
        if tracker.has_any_configured_module(): widgets['ship_indicator'].config(fg=GREEN)
        else: widgets['ship_indicator'].config(fg=RED)

    # Met à jour l'étiquette du profil actif affiché dans l'interface principale
    def update_profile_label(self, char_id: str):
        tracker = self.all_characters[char_id]
        if char_id not in self.char_widgets: return
        self.char_widgets[char_id]['profile_label'].config(text=f"\u3008{tracker.active_profile}\u3009")

    # Calcule et affiche le débit réel (m³/s) et la durée de session formatée
    def _update_rate_stats(self, char_id: str, tracker: CharacterTracker, widgets: Dict):
        theoretical_m3_per_sec = tracker.get_total_theoretical_m3_per_sec()
        if theoretical_m3_per_sec > 0: widgets['theoretical'].config(text=f"◈ Theoretical: {theoretical_m3_per_sec:.2f} m3/s ({theoretical_m3_per_sec * 3600:,.0f} m3/hr)")
        else: widgets['theoretical'].config(text="◈ Theoretical: -- m3/s (configure ship)")

        if not tracker.session_active: return

        actual_m3_per_sec = 0.0
        session_duration = tracker.get_session_active_duration()
        if session_duration > 10 and tracker.total_m3 > 0:
            actual_m3_per_sec = (tracker.total_m3 - tracker.session_start_m3) / session_duration

        widgets['actual'].config(text=f"◉ Actual: {actual_m3_per_sec:.2f} m3/s ({actual_m3_per_sec * 3600:,.0f} m3/hr)")

    # Ouvre la boîte de dialogue des paramètres : personnages, thème, sons, chemins, webhook
    def show_config_dialog(self):
        global DOCS, UPDATE_INTERVAL_MS, HISTORY_DAYS

        if self.config_dialog is not None and self.config_dialog.winfo_exists():
            self.config_dialog.lift()
            self.config_dialog.focus_force()
            return

        # ── helpers ──────────────────────────────────────────────────────
        def _sep(parent):
            tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", pady=8)

        def _section_label(parent, text):
            tk.Label(parent, text=text, fg=CYAN, bg=BG_PANEL,
                     font=("Consolas", 9, "bold")).pack(anchor="w", pady=(0, 10))

        def _card(parent, pady=(0, 10)):
            outer = tk.Frame(parent, bg=BORDER, padx=1, pady=1)
            outer.pack(fill="x", pady=pady)
            inner = tk.Frame(outer, bg=BG_CARD, padx=14, pady=12)
            inner.pack(fill="both", expand=True)
            return inner

        def _field_row(parent, label, var_or_val, width=28, note=None):
            row = tk.Frame(parent, bg=BG_CARD)
            row.pack(fill="x", pady=3)
            tk.Label(row, text=label, fg=DIM, bg=BG_CARD,
                     font=("Consolas", 9), width=22, anchor="w").pack(side="left")
            if isinstance(var_or_val, tk.StringVar):
                var = var_or_val
            else:
                var = tk.StringVar(value=str(var_or_val))
            e = tk.Entry(row, textvariable=var, width=width,
                         font=("Consolas", 9), bg=BG, fg=WHITE,
                         insertbackground=CYAN, relief="flat",
                         highlightthickness=1,
                         highlightbackground=BORDER, highlightcolor=CYAN)
            e.pack(side="left")
            if note:
                tk.Label(row, text=note, fg=GOLD, bg=BG_CARD,
                         font=("Consolas", 8)).pack(side="left", padx=(6, 0))
            return var

        def _make_scrollable(parent):
            canvas = tk.Canvas(parent, bg=BG_PANEL, highlightthickness=0, bd=0)
            sb = tk.Scrollbar(parent, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=sb.set)
            inner = tk.Frame(canvas, bg=BG_PANEL, padx=18, pady=14)
            win_id = canvas.create_window((0, 0), window=inner, anchor="nw")
            def _resize(e): canvas.itemconfig(win_id, width=e.width)
            def _scroll_region(e): canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.bind("<Configure>", _resize)
            inner.bind("<Configure>", _scroll_region)
            canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>",
                lambda ev: canvas.yview_scroll(int(-1*(ev.delta/120)), "units")))
            canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
            canvas.pack(side="left", fill="both", expand=True)
            sb.pack(side="right", fill="y")
            return inner

        # ── constants ────────────────────────────────────────────────────
        NAV_W    = 140
        BG_CARD  = "#0f1620"

        self.config_icon.config(fg=CYAN)
        self.config_icon.unbind("<Button-1>")
        self.config_icon.unbind("<Enter>")
        self.config_icon.unbind("<Leave>")

        dialog = tk.Toplevel(self.root)
        dialog.configure(bg=BORDER)
        dialog.overrideredirect(True)
        dialog.wm_attributes("-topmost", 1)
        dialog.attributes("-alpha", WIN_ALPHA)
        dialog.resizable(False, False)
        self.config_dialog = dialog

        _drag_x = [0]; _drag_y = [0]
        def start_drag(event):
            if isinstance(event.widget, (tk.Entry, ttk.Combobox)): return
            _drag_x[0] = event.x; _drag_y[0] = event.y
        def do_drag(event):
            if isinstance(event.widget, (tk.Entry, ttk.Combobox)): return
            dialog.geometry(f"+{dialog.winfo_x()+event.x-_drag_x[0]}+{dialog.winfo_y()+event.y-_drag_y[0]}")

        config_key  = "config_dialog_geom"
        saved_geom  = self.app_config.get(config_key, "+250+150")
        app_settings = self.app_config.get("app_settings", {})

        # ── window chrome ────────────────────────────────────────────────
        outer_frame = tk.Frame(dialog, bg=BORDER, padx=1, pady=1)
        outer_frame.pack(fill="both", expand=True)
        main_frame = tk.Frame(outer_frame, bg=BG_PANEL)
        main_frame.pack(fill="both", expand=True)

        # title bar
        title_bar = tk.Frame(main_frame, bg=BG_PANEL, padx=16, pady=10)
        title_bar.pack(fill="x")
        title_bar.bind("<Button-1>", start_drag)
        title_bar.bind("<B1-Motion>", do_drag)
        title_lbl = tk.Label(title_bar, text="⚙  CONFIG", fg=CYAN, bg=BG_PANEL,
                             font=("Consolas", 11, "bold"))
        title_lbl.pack(side="left")
        title_lbl.bind("<Button-1>", start_drag)
        title_lbl.bind("<B1-Motion>", do_drag)

        def close_dialog():
            try:
                x, y = dialog.winfo_x(), dialog.winfo_y()
                self.app_config[config_key] = f"+{x}+{y}"
                self.save_config()
            except Exception: pass
            self.config_dialog = None
            self._enable_config_icon()
            dialog.destroy()

        close_btn = tk.Label(title_bar, text="✕", fg=DIM, bg=BG_PANEL,
                             font=("Consolas", 14, "bold"), cursor="hand2")
        close_btn.pack(side="right")
        close_btn.bind("<Button-1>", lambda e: close_dialog())
        close_btn.bind("<Enter>",    lambda e: close_btn.config(fg=RED))
        close_btn.bind("<Leave>",    lambda e: close_btn.config(fg=DIM))

        # thin line under title
        tk.Frame(main_frame, bg=BORDER, height=1).pack(fill="x")

        # body = sidebar + content
        body = tk.Frame(main_frame, bg=BG_PANEL)
        body.pack(fill="both", expand=True)

        sidebar = tk.Frame(body, bg=BG, width=NAV_W)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)
        tk.Frame(body, bg=BORDER, width=1).pack(side="left", fill="y")

        content_host = tk.Frame(body, bg=BG_PANEL)
        content_host.pack(side="left", fill="both", expand=True)

        # bottom bar
        tk.Frame(main_frame, bg=BORDER, height=1).pack(fill="x")
        bottom = tk.Frame(main_frame, bg=BG_PANEL, padx=16, pady=10)
        bottom.pack(fill="x")

        # ── SECTION: CHARACTERS ──────────────────────────────────────────
        p_chars = tk.Frame(content_host, bg=BG_PANEL)
        inner_chars = _make_scrollable(p_chars)

        _section_label(inner_chars, "◆  ACTIVE CHARACTERS")
        tk.Label(inner_chars,
                 text="Choose which pilots are shown on the dashboard.",
                 fg=DIM, bg=BG_PANEL, font=("Consolas", 8),
                 wraplength=380, justify="left").pack(anchor="w", pady=(0, 10))

        visible_chars = self.app_config.get("visible_characters", [])
        if visible_chars is None: visible_chars = []
        char_vars = {}

        ctrl_frame = tk.Frame(inner_chars, bg=BG_PANEL)
        ctrl_frame.pack(fill="x", pady=(0, 8))
        def set_all_chars(state):
            for v in char_vars.values(): v.set(state)
        tk.Button(ctrl_frame, text="SELECT ALL", command=lambda: set_all_chars(True),
                  bg=BG, fg=CYAN, font=("Consolas", 8, "bold"),
                  relief="flat", cursor="hand2", width=12).pack(side="left", padx=(0, 6))
        tk.Button(ctrl_frame, text="DESELECT ALL", command=lambda: set_all_chars(False),
                  bg=BG, fg=WHITE, font=("Consolas", 8, "bold"),
                  relief="flat", cursor="hand2", width=12).pack(side="left")

        def _make_toggle(parent, var, bg):
            """Custom checkbox — no tkinter Checkbutton flicker."""
            box = tk.Label(parent, font=("Consolas", 12), cursor="hand2", bg=bg)
            def _refresh():
                box.config(text="■", fg=CYAN) if var.get() else box.config(text="□", fg=DIM)
            def _toggle(e=None):
                var.set(not var.get())
                _refresh()
            box.bind("<Button-1>", _toggle)
            _refresh()
            return box

        grid_frame = tk.Frame(inner_chars, bg=BG_PANEL)
        grid_frame.pack(fill="x")
        grid_frame.columnconfigure(0, weight=1)
        grid_frame.columnconfigure(1, weight=1)
        for i, (char_id, tracker) in enumerate(self.all_characters.items()):
            var = tk.BooleanVar(value=char_id in visible_chars)
            char_vars[char_id] = var
            accent = CHAR_ACCENTS[i % len(CHAR_ACCENTS)]
            cell = tk.Frame(grid_frame, bg=BG_CARD, padx=10, pady=8)
            pad_l = 0 if i % 2 == 0 else 4
            pad_r = 4 if i % 2 == 0 else 0
            cell.grid(row=i//2, column=i%2, sticky="ew",
                      padx=(pad_l, pad_r), pady=3)
            tk.Frame(cell, bg=accent, width=3).pack(side="left", fill="y", padx=(0, 8))
            box = _make_toggle(cell, var, BG_CARD)
            box.pack(side="left", padx=(0, 4))
            name_lbl = tk.Label(cell, text=tracker.char_name, fg=accent, bg=BG_CARD,
                                font=("Consolas", 10, "bold"), cursor="hand2")
            name_lbl.pack(side="left", padx=(4, 0))
            # clicking the name also toggles — re-use the same box binding
            name_lbl.bind("<Button-1>", lambda e, b=box: b.event_generate("<Button-1>"))

        # ── SECTION: APPEARANCE ──────────────────────────────────────────
        p_appearance = tk.Frame(content_host, bg=BG_PANEL)
        inner_app = _make_scrollable(p_appearance)

        _section_label(inner_app, "◆  PATHS & FILES")
        c_paths = _card(inner_app)
        docs_var = _field_row(c_paths, "Gamelogs Path:",
                              app_settings.get("docs_path", DOCS), width=26)

        _sep(inner_app)
        _section_label(inner_app, "◆  APPEARANCE")
        c_appear = _card(inner_app)

        # theme row
        theme_row = tk.Frame(c_appear, bg=BG_CARD)
        theme_row.pack(fill="x", pady=3)
        tk.Label(theme_row, text="Dashboard Theme:", fg=DIM, bg=BG_CARD,
                 font=("Consolas", 9), width=22, anchor="w").pack(side="left")
        theme_var = tk.StringVar(value=self.app_theme)
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Dark.TCombobox", fieldbackground=BG, background=BG_PANEL,
                        foreground=WHITE, bordercolor=BORDER, arrowcolor=CYAN)
        style.map("Dark.TCombobox",
                  fieldbackground=[("readonly", BG)],
                  foreground=[("readonly", WHITE)],
                  selectbackground=[("readonly", BORDER)],
                  selectforeground=[("readonly", CYAN)])
        self.root.option_add("*TCombobox*Listbox.background", BG)
        self.root.option_add("*TCombobox*Listbox.foreground", WHITE)
        self.root.option_add("*TCombobox*Listbox.selectBackground", BORDER)
        self.root.option_add("*TCombobox*Listbox.selectForeground", CYAN)
        self.root.option_add("*TCombobox*Listbox.font", ("Consolas", 9))
        theme_cb = ttk.Combobox(theme_row, textvariable=theme_var,
                                values=THEME_NAMES, state="readonly",
                                style="Dark.TCombobox",
                                font=("Consolas", 9), width=26)
        theme_cb.pack(side="left")

        history_var = _field_row(c_appear, "Default History Days:",
                                 app_settings.get("history_days", HISTORY_DAYS), width=8)

        # transparency
        alpha_row = tk.Frame(c_appear, bg=BG_CARD)
        alpha_row.pack(fill="x", pady=3)
        tk.Label(alpha_row, text="Window Transparency:", fg=DIM, bg=BG_CARD,
                 font=("Consolas", 9), width=22, anchor="w").pack(side="left")
        alpha_var    = tk.DoubleVar(value=app_settings.get("win_alpha", WIN_ALPHA))
        alpha_pct_var = tk.StringVar(value=f"{int(alpha_var.get()*100)}%")
        def on_alpha_change(val):
            v = float(val)
            alpha_pct_var.set(f"{int(v*100)}%")
            self._apply_alpha(v)
        tk.Scale(alpha_row, variable=alpha_var, from_=0.2, to=1.0, resolution=0.05,
                 orient="horizontal", length=160, command=on_alpha_change,
                 bg=BG_CARD, fg=WHITE, troughcolor=BG, highlightthickness=0,
                 activebackground=CYAN, sliderrelief="flat",
                 showvalue=False).pack(side="left")
        tk.Label(alpha_row, textvariable=alpha_pct_var, fg=CYAN, bg=BG_CARD,
                 font=("Consolas", 9), width=5).pack(side="left", padx=(6, 0))

        # sound toggle
        crit_sound_var = tk.BooleanVar(value=app_settings.get("play_crit_sound", PLAY_CRIT_SOUND))
        sound_row = tk.Frame(c_appear, bg=BG_CARD)
        sound_row.pack(fill="x", pady=(6, 0))
        sound_box = _make_toggle(sound_row, crit_sound_var, BG_CARD)
        sound_box.pack(side="left", padx=(0, 4))
        sound_lbl = tk.Label(sound_row, text="Play audio on Critical Hit",
                             fg=WHITE, bg=BG_CARD, font=("Consolas", 9), cursor="hand2")
        sound_lbl.pack(side="left", padx=(4, 0))
        sound_lbl.bind("<Button-1>", lambda e: sound_box.event_generate("<Button-1>"))

        # ── SECTION: FLEET ───────────────────────────────────────────────
        p_fleet = tk.Frame(content_host, bg=BG_PANEL)
        inner_fleet = _make_scrollable(p_fleet)

        _section_label(inner_fleet, "◆  FLEET INTEGRATION")
        tk.Label(inner_fleet,
                 text="Paste a Discord webhook URL to broadcast mining stats to your fleet channel.",
                 fg=DIM, bg=BG_PANEL, font=("Consolas", 8),
                 wraplength=380, justify="left").pack(anchor="w", pady=(0, 10))
        fleet_cfg = self.app_config.get("fleet", {})
        c_fleet = _card(inner_fleet)
        webhook_var = _field_row(c_fleet, "Webhook URL:",
                                 fleet_cfg.get("webhook_url", ""), width=34)

        # ── SECTION: DATABASE ────────────────────────────────────────────
        p_db = tk.Frame(content_host, bg=BG_PANEL)
        inner_db = _make_scrollable(p_db)

        _section_label(inner_db, "◆  ORE DATABASE  (SDE)")
        tk.Label(inner_db,
                 text="The ore database is sourced from the EVE Static Data Export.\nUpdate it whenever CCP releases a new expansion.",
                 fg=DIM, bg=BG_PANEL, font=("Consolas", 8),
                 justify="left").pack(anchor="w", pady=(0, 10))

        info_card = _card(inner_db)
        sde_info_var = tk.StringVar()
        def _refresh_sde_label():
            sde_info_var.set(
                f"SDE: {SDE_INFO['version']}  |  {SDE_INFO['ore_count']} ores  |  {SDE_INFO['updated_at']}")
        _refresh_sde_label()
        for lbl_text, key in [("Version:", "version"), ("Ore types:", "ore_count"), ("Last updated:", "updated_at")]:
            r = tk.Frame(info_card, bg=BG_CARD)
            r.pack(fill="x", pady=2)
            tk.Label(r, text=lbl_text, fg=DIM, bg=BG_CARD,
                     font=("Consolas", 9), width=16, anchor="w").pack(side="left")
            tk.Label(r, text=SDE_INFO[key], fg=WHITE, bg=BG_CARD,
                     font=("Consolas", 9)).pack(side="left")

        tk.Frame(inner_db, bg=BG_PANEL, height=8).pack()
        sde_bar_frame = tk.Frame(inner_db, bg=BG_PANEL)
        sde_bar_frame.pack(fill="x")
        sde_bar_border = tk.Frame(sde_bar_frame, bg=CYAN, padx=1, pady=1)
        sde_bar_border.pack(fill="x")
        sde_bar_canvas = tk.Canvas(sde_bar_border, height=18, bg="#0a1520", highlightthickness=0)
        sde_bar_canvas.pack(fill="x")
        sde_bar_pct_label = tk.Label(sde_bar_frame, text="", fg=CYAN, bg=BG_PANEL,
                                     font=("Consolas", 8, "bold"), anchor="center")
        sde_bar_pct_label.pack(fill="x")
        sde_bar_frame.pack_forget()

        sde_status_var = tk.StringVar(value="")
        tk.Label(inner_db, textvariable=sde_status_var, fg=GOLD, bg=BG_PANEL,
                 font=("Consolas", 8)).pack(anchor="w", pady=(6, 0))

        tk.Frame(inner_db, bg=BG_PANEL, height=10).pack()

        def do_sde_update():
            global ORE_VOLUMES, COMPRESSION_RATIOS, SDE_INFO
            update_btn.config(state="disabled", text="↻  UPDATING...")
            sde_bar_frame.pack(fill="x")
            draw_neon_bar(sde_bar_canvas, 0)
            sde_bar_pct_label.config(text="")

            def run_update():
                try:
                    def progress(msg):
                        try:
                            pct_match = SDE_PROGRESS_PCT.search(msg)
                            if pct_match:
                                pct_val = int(pct_match.group("pct")) / 100.0
                                dialog.after(0, lambda p=pct_val, m=msg: _update_sde_progress(p, m))
                            elif "Extracting" in msg: dialog.after(0, lambda m=msg: _update_sde_progress(0.85, m))
                            elif "Parsing"    in msg: dialog.after(0, lambda m=msg: _update_sde_progress(0.95, m))
                            else:                     dialog.after(0, lambda m=msg: _update_sde_progress(0.02, m))
                        except Exception: pass

                    result = download_and_parse_sde(progress_callback=progress)
                    _save_ore_data_cache(result)
                    ORE_VOLUMES         = {k: float(v) for k, v in result["ore_volumes"].items()}
                    COMPRESSION_RATIOS  = {k: int(v)   for k, v in result["compression_ratios"].items()}
                    SDE_INFO["version"]   = result.get("sde_version", "updated")
                    SDE_INFO["updated_at"]= result.get("updated_at", "now")
                    SDE_INFO["ore_count"] = str(result.get("ore_count", len(ORE_VOLUMES)))

                    def on_success():
                        sde_status_var.set(f"✔ Updated! {SDE_INFO['ore_count']} ores loaded.")
                        sde_status_label_ref[0].config(fg=GREEN)
                        draw_neon_bar(sde_bar_canvas, 1.0)
                        sde_bar_pct_label.config(text="100% ─ Complete!", fg=GREEN)
                        update_btn.config(state="normal", text="↻  UPDATE ORE DATA")
                        self._history_cache = None
                    try: dialog.after(0, on_success)
                    except Exception: pass

                except Exception as e:
                    def on_error():
                        sde_status_var.set(f"✖ Error: {str(e)[:60]}")
                        sde_status_label_ref[0].config(fg=RED)
                        draw_neon_bar(sde_bar_canvas, 0, bar_color=RED)
                        sde_bar_pct_label.config(text="Download failed", fg=RED)
                        update_btn.config(state="normal", text="↻  UPDATE ORE DATA")
                    try: dialog.after(0, on_error)
                    except Exception: pass

            threading.Thread(target=run_update, daemon=True).start()

        def _update_sde_progress(pct_val, msg):
            sde_status_var.set(msg)
            draw_neon_bar(sde_bar_canvas, pct_val)
            pct_display = int(pct_val * 100)
            sde_bar_pct_label.config(text=f"{pct_display}%  ─  {msg[:40]}" if msg else f"{pct_display}%")

        sde_status_label_ref = [
            tk.Label(inner_db, textvariable=sde_status_var, fg=GOLD, bg=BG_PANEL,
                     font=("Consolas", 8))
        ]
        update_btn = tk.Button(inner_db, text="↻  UPDATE ORE DATA",
                               command=do_sde_update, bg=BG, fg=CYAN,
                               font=("Consolas", 9, "bold"), relief="flat",
                               cursor="hand2", pady=6, padx=12)
        update_btn.pack(anchor="w")

        # ── NAV SIDEBAR ──────────────────────────────────────────────────
        panels = {
            "CHARACTERS": p_chars,
            "APPEARANCE": p_appearance,
            "FLEET":      p_fleet,
            "DATABASE":   p_db,
        }
        nav_buttons   = {}
        active_section = tk.StringVar(value="CHARACTERS")

        NAV_ITEMS = [
            ("CHARACTERS", "◈"),
            ("APPEARANCE", "◈"),
            ("FLEET",      "◈"),
            ("DATABASE",   "◈"),
        ]

        def show_section(name):
            active_section.set(name)
            for n, panel in panels.items():
                panel.pack_forget()
            panels[name].pack(fill="both", expand=True)
            for n, nd in nav_buttons.items():
                is_active = (n == name)
                nd["btn"].config(fg=CYAN if is_active else DIM,
                                 bg="#0f1e28" if is_active else BG)
                nd["bar"].config(bg=CYAN if is_active else BG)

        for name, icon in NAV_ITEMS:
            nav_row = tk.Frame(sidebar, bg=BG)
            nav_row.pack(fill="x")
            bar = tk.Frame(nav_row, bg=BG, width=3)
            bar.pack(side="left", fill="y")
            btn = tk.Button(nav_row, text=f"{icon}  {name}", fg=DIM, bg=BG,
                            font=("Consolas", 8, "bold"), relief="flat",
                            cursor="hand2", anchor="w", pady=10,
                            command=lambda n=name: show_section(n))
            btn.pack(side="left", fill="both", expand=True)
            btn.bind("<Enter>", lambda e, b=btn, n=name: b.config(
                fg=WHITE if active_section.get() != n else CYAN))
            btn.bind("<Leave>", lambda e, b=btn, n=name: b.config(
                fg=CYAN if active_section.get() == n else DIM))
            nav_buttons[name] = {"btn": btn, "bar": bar}
            tk.Frame(sidebar, bg=BORDER, height=1).pack(fill="x")

        show_section("CHARACTERS")

        # ── THEME PREVIEW (live rebuild on selection) ─────────────────────
        def on_theme_preview(event=None):
            new_theme = theme_var.get()
            if new_theme == self.app_theme: return
            try: history_safe = max(1, int(history_var.get()))
            except: history_safe = HISTORY_DAYS
            self.app_config["app_settings"] = {
                "docs_path": docs_var.get().strip(), "crit_sound_file": CRIT_SOUND_FILE,
                "update_interval_ms": UPDATE_INTERVAL_MS, "history_days": history_safe,
                "max_modules": MAX_MODULES, "play_crit_sound": crit_sound_var.get(),
                "win_alpha": float(alpha_var.get()),
            }
            fleet_cfg_p = self.app_config.get("fleet", {})
            fleet_cfg_p["webhook_url"] = webhook_var.get().strip()
            self.app_config["fleet"] = fleet_cfg_p
            self.app_config["theme"] = new_theme
            self.app_theme = new_theme
            apply_theme_colors(new_theme)
            try:
                cx, cy = dialog.winfo_x(), dialog.winfo_y()
                self.app_config[config_key] = f"+{cx}+{cy}"
            except Exception: pass
            self.config_dialog = None
            self._enable_config_icon()
            dialog.destroy()
            self.rebuild_all_ui()
            self.root.after(80, self.show_config_dialog)

        theme_cb.bind("<<ComboboxSelected>>", on_theme_preview)

        # ── SAVE / CANCEL ────────────────────────────────────────────────
        def save_and_close():
            global DOCS, UPDATE_INTERVAL_MS, HISTORY_DAYS, PLAY_CRIT_SOUND, WIN_ALPHA
            try:
                new_history = int(history_var.get())
                if new_history < 1: new_history = 1
            except ValueError:
                messagebox.showerror("Invalid Input", "Please enter a valid number for history days.")
                return

            selected_theme  = theme_var.get()
            theme_changed   = selected_theme != self.app_theme
            selected_chars  = [cid for cid, v in char_vars.items() if v.get()]
            self.save_visible_characters(selected_chars)

            DOCS             = docs_var.get().strip()
            HISTORY_DAYS     = new_history
            PLAY_CRIT_SOUND  = crit_sound_var.get()
            WIN_ALPHA        = max(0.2, min(1.0, float(alpha_var.get())))
            self._apply_alpha(WIN_ALPHA)

            self.app_config["app_settings"] = {
                "docs_path": DOCS, "crit_sound_file": CRIT_SOUND_FILE,
                "update_interval_ms": UPDATE_INTERVAL_MS, "history_days": HISTORY_DAYS,
                "max_modules": MAX_MODULES, "play_crit_sound": PLAY_CRIT_SOUND,
                "win_alpha": WIN_ALPHA,
            }
            self.app_config["theme"] = selected_theme
            self.app_theme           = selected_theme
            self.fleet_webhook_url   = webhook_var.get().strip()
            fleet_cfg_s = self.app_config.get("fleet", {})
            fleet_cfg_s["webhook_url"] = self.fleet_webhook_url
            self.app_config["fleet"] = fleet_cfg_s
            self.save_config()
            self._update_send_button_states()

            try:
                x, y = dialog.winfo_x(), dialog.winfo_y()
                self.app_config[config_key] = f"+{x}+{y}"
            except Exception: pass

            self.save_config()
            self.config_dialog = None
            self._enable_config_icon()
            dialog.destroy()
            if theme_changed:
                apply_theme_colors(self.app_theme)
                self.rebuild_all_ui()

        tk.Button(bottom, text="✔  SAVE", command=save_and_close,
                  bg=BG, fg=GREEN, font=("Consolas", 9, "bold"),
                  relief="flat", cursor="hand2", width=12, pady=4).pack(side="right", padx=(5, 0))
        tk.Button(bottom, text="✕  CANCEL", command=close_dialog,
                  bg=BG, fg=RED, font=("Consolas", 9, "bold"),
                  relief="flat", cursor="hand2", width=12, pady=4).pack(side="right")

        dialog.update_idletasks()
        try:
            if "+" in saved_geom:
                parts = saved_geom.split("+")
                if len(parts) >= 3: dialog.geometry(f"660x580+{parts[1]}+{parts[2]}")
                else: dialog.geometry("660x580+250+150")
            else: dialog.geometry("660x580+250+150")
        except Exception: dialog.geometry("660x580+250+150")
        dialog.update()

        def initial_focus():
            if dialog.winfo_exists():
                dialog.lift()
                dialog.focus_force()

        dialog.after(150, initial_focus)

    # Vérifie si l'URL du webhook Discord configurée est valide (commence par https://discord.com/api/webhooks/)
    def _is_valid_webhook_url(self) -> bool:
        url = self.fleet_webhook_url.strip()
        if not url: return False
        return url.startswith("https://discord.com/api/webhooks/") or url.startswith("https://discordapp.com/api/webhooks/")

    # Active ou désactive les boutons Copy/Send selon les données disponibles et le webhook configuré
    def _update_send_button_states(self):
        has_webhook = self._is_valid_webhook_url()
        for cid, w in self.char_widgets.items():
            tracker = self.all_characters.get(cid)
            has_data = bool(tracker and tracker.ore_summary)
            if has_data:
                w['copy_btn'].config(state="normal", fg=GOLD)
                w['copy_tip'].update_text("Copy session report to clipboard")
                if has_webhook:
                    w['send_btn'].config(state="normal", fg=CYAN)
                    w['send_tip'].update_text("Send session report to Discord webhook")
                else:
                    w['send_btn'].config(state="disabled", fg=DIM)
                    w['send_tip'].update_text("No webhook URL configured \u2014 set it in \u2699 Config")
            else:
                w['copy_btn'].config(state="disabled", fg=DIM)
                w['copy_tip'].update_text("No mining data yet \u2014 start mining to enable")
                w['send_btn'].config(state="disabled", fg=DIM)
                if not has_webhook: w['send_tip'].update_text("No mining data and no webhook URL configured")
                else: w['send_tip'].update_text("No mining data yet \u2014 start mining to enable")

    # Génère le texte formaté du rapport de session (totaux, minerais, débit, durée)
    def _build_session_report_text(self, tracker: CharacterTracker) -> str:
        session_duration = tracker.get_session_active_duration()
        hours = int(session_duration // 3600)
        minutes = int((session_duration % 3600) // 60)
        duration_str = f"{hours}h {minutes:02d}m" if hours > 0 else f"{minutes}m"

        lines = [
            f"Mining Report — {tracker.char_name}",
            f"Session: {duration_str} | Crit Bonus: {tracker.crit_m3:,.1f} m³ ({tracker.crit_count})",
            ""
        ]

        total_m3 = 0.0
        if tracker.ore_summary:
            for ore_name, volume in sorted(tracker.ore_summary.items(), key=lambda x: x[1], reverse=True):
                vol_per_unit, _ = self.get_ore_volume(ore_name)
                units = int(volume / vol_per_unit) if vol_per_unit > 0 else 0
                lines.append(f"  {ore_name}: {volume:,.1f} m³ ({units:,} units)")
                total_m3 += volume
        else: lines.append("  No ores mined yet.")

        lines.append("")
        lines.append(f"Total: {total_m3:,.1f} m³")
        return "\n".join(lines)

    # Construit le payload JSON (embed Discord) prêt à envoyer au webhook
    def _build_discord_payload(self, tracker: CharacterTracker) -> Dict:
        report_text = self._build_session_report_text(tracker)
        return {"content": report_text}

    # Copie le rapport de session du personnage dans le presse-papiers Windows
    def copy_session_report(self, char_id: str):
        tracker = self.all_characters.get(char_id)
        if not tracker: return

        session_m3 = tracker.total_m3 - tracker.session_start_m3
        if session_m3 <= 0 and not tracker.ore_summary:
            messagebox.showinfo("No Data", "No mining data in current session.", parent=self.root)
            return

        report_text = self._build_session_report_text(tracker)
        self.root.clipboard_clear()
        self.root.clipboard_append(report_text)

        if char_id in self.char_widgets:
            btn = self.char_widgets[char_id].get('copy_btn')
            if btn:
                original_text = btn.cget('text')
                original_fg = btn.cget('fg')
                btn.config(text="✓ Copied!", fg=GREEN)
                btn.after(2000, lambda: btn.config(text=original_text, fg=original_fg))

    # Ouvre le dialogue de prévisualisation et d'envoi du rapport vers Discord
    def show_send_report_dialog(self, char_id: str):
        tracker = self.all_characters.get(char_id)
        if not tracker: return

        session_m3 = tracker.total_m3 - tracker.session_start_m3
        if session_m3 <= 0 and not tracker.ore_summary:
            messagebox.showinfo("No Data", "No mining data in current session.", parent=self.root)
            return

        if not self.fleet_webhook_url:
            messagebox.showwarning("No Webhook", "Webhook URL not configured.\nSet it in ⚙ Config → Fleet section.", parent=self.root)
            return

        report_text = self._build_session_report_text(tracker)

        dlg = tk.Toplevel(self.root)
        dlg.configure(bg=BORDER)
        dlg.overrideredirect(True)
        dlg.wm_attributes("-topmost", 1)
        dlg.attributes("-alpha", 0.90)
        dlg.resizable(False, False)

        _drag_x = [0]
        _drag_y = [0]
        def start_drag(event): _drag_x[0] = event.x; _drag_y[0] = event.y
        def do_drag(event):
            x = dlg.winfo_x() + event.x - _drag_x[0]
            y = dlg.winfo_y() + event.y - _drag_y[0]
            dlg.geometry(f"+{x}+{y}")

        dlg.bind("<Button-1>", start_drag)
        dlg.bind("<B1-Motion>", do_drag)

        border_frame = tk.Frame(dlg, bg=BORDER, padx=1, pady=1)
        border_frame.pack(fill="both", expand=True)

        main_frame = tk.Frame(border_frame, bg=BG_PANEL, padx=15, pady=15)
        main_frame.pack(fill="both", expand=True)

        tk.Label(main_frame, text="▲ Send Mining Report to Discord?", fg=CYAN, bg=BG_PANEL, font=("Consolas", 10, "bold")).pack(anchor="w", pady=(0, 10))

        preview_outer = tk.Frame(main_frame, bg=BORDER, padx=1, pady=1)
        preview_outer.pack(fill="both", pady=(0, 10))

        preview_text = tk.Text(preview_outer, bg=BG, fg=WHITE, font=("Consolas", 9), relief="flat", padx=10, pady=10, wrap="word", width=42, height=12)
        preview_text.pack(fill="both")
        preview_text.insert("1.0", report_text)
        preview_text.config(state="disabled")

        url_display = self.fleet_webhook_url
        if len(url_display) > 50: url_display = url_display[:25] + "..." + url_display[-22:]
        tk.Label(main_frame, text=f"To: {url_display}", fg=DIM, bg=BG_PANEL, font=("Consolas", 8)).pack(anchor="w", pady=(0, 8))

        btn_frame = tk.Frame(main_frame, bg=BG_PANEL)
        btn_frame.pack()

        def do_send():
            dlg.destroy()
            self._send_to_webhook(char_id)

        tk.Button(btn_frame, text="✖ Cancel", command=dlg.destroy, bg=BG, fg=RED, font=("Consolas", 9, "bold"), relief="flat", cursor="hand2", width=10).pack(side="left", padx=5)
        tk.Button(btn_frame, text="✔ Send", command=do_send, bg=BG, fg=GREEN, font=("Consolas", 9, "bold"), relief="flat", cursor="hand2", width=10).pack(side="left", padx=5)

        dlg.update_idletasks()
        pw, ph = self.root.winfo_width(), self.root.winfo_height()
        px, py = self.root.winfo_x(), self.root.winfo_y()
        dw, dh = dlg.winfo_reqwidth(), dlg.winfo_reqheight()
        x, y = px + (pw - dw) // 2, py + (ph - dh) // 2
        dlg.geometry(f"+{x}+{y}")

    # Envoie le rapport de session au webhook Discord configuré via HTTP POST
    def _send_to_webhook(self, char_id: str):
        tracker = self.all_characters.get(char_id)
        if not tracker or not self.fleet_webhook_url: return
        payload = self._build_discord_payload(tracker)

        try:
            data = json.dumps(payload).encode('utf-8')
            req = urllib.request.Request(self.fleet_webhook_url, data=data, headers={"Content-Type": "application/json", "User-Agent": "EVE-Mining-Dashboard/1.0"}, method="POST")
            response = urllib.request.urlopen(req, timeout=10)
            status = response.getcode()

            if status in (200, 204):
                if char_id in self.char_widgets:
                    btn = self.char_widgets[char_id].get('send_btn')
                    if btn:
                        original_text = btn.cget('text')
                        original_fg = btn.cget('fg')
                        btn.config(text="✓ Sent!", fg=GREEN)
                        btn.after(3000, lambda: btn.config(text=original_text, fg=original_fg))
            else:
                messagebox.showerror("Send Failed", f"Discord returned status {status}", parent=self.root)
        except urllib.error.HTTPError as e:
            error_body = ""
            try: error_body = e.read().decode('utf-8', errors='ignore')[:200]
            except Exception: pass
            messagebox.showerror("Send Failed", f"HTTP {e.code}: {e.reason}\n{error_body}", parent=self.root)
        except urllib.error.URLError as e:
            messagebox.showerror("Send Failed", f"Connection error:\n{str(e.reason)}", parent=self.root)
        except Exception as e:
            messagebox.showerror("Send Failed", f"Error: {str(e)}", parent=self.root)

    # Réactive l'icône ⚙ après la fermeture du dialogue de config
    def _enable_config_icon(self):
        self.config_icon.config(fg=DIM)
        self.config_icon.bind("<Button-1>", lambda e: self.show_config_dialog())
        self.config_icon.bind("<Enter>", lambda e: self.config_icon.config(fg=CYAN))
        self.config_icon.bind("<Leave>", lambda e: self.config_icon.config(fg=DIM))

if __name__ == "__main__":
    MiningDashboard()